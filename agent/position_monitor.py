from __future__ import annotations

import os
import sys
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.data import fetch_intraday_frame
from app.telegram_notifications import (
    dashboard_url_from_env,
    format_position_event_message,
    send_telegram_message,
)


DEFAULT_TRACKER = ROOT / "agent_tracker" / "market_lens_agent_portfolio_budget_100k.xlsx"
DEFAULT_RUN_DIR = ROOT / "agent_results"
EVENT_SHEET = "Position Events"


@dataclass
class MonitorSettings:
    excel_path: Path
    run_dir: Path
    period: str
    interval: str
    save_noop: bool
    dashboard_url: str


@dataclass
class PositionEvent:
    ticker: str
    action: str
    triggered_at: str
    trigger_price: float
    high: float
    low: float
    close: float
    quantity: int
    cash_in: float
    note: str


@dataclass
class MonitorResult:
    ticker: str
    status: str
    current_price: float
    event: PositionEvent | None = None
    error: str = ""


def main() -> None:
    settings = load_settings()
    run_id = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    timestamp = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    wb = load_workbook(settings.excel_path)
    ensure_agent_columns(wb)
    ensure_position_events_sheet(wb)

    settings_values = read_settings(wb)
    currency = str(settings_values.get("budget_currency") or "USD").upper()
    currency_rate = 1.0 if currency == "USD" else float(settings_values.get("usd_ils_rate", 3.7))
    starting_capital = float(
        settings_values.get("starting_capital_usd")
        or settings_values.get("starting_capital_ils")
        or 100_000
    )

    open_positions = read_open_positions(wb)
    last_event_times = read_last_event_times(wb)
    results: list[MonitorResult] = []
    event_notifications: list[tuple[dict[str, Any], PositionEvent]] = []

    for ticker, position in list(open_positions.items()):
        since = last_event_times.get(ticker) or parse_timestamp(position.get("entry_date"))
        result = monitor_position(position, settings=settings, since=since, currency_rate=currency_rate)
        results.append(result)
        if result.event:
            event_notifications.append((dict(position), result.event))
            apply_event(wb, open_positions, position, result.event, timestamp, run_id, currency_rate)
        elif result.current_price > 0 and settings.save_noop:
            refresh_position(position, result.current_price, currency_rate)

    events = [result.event for result in results if result.event]
    errors = [f"{result.ticker}: {result.error}" for result in results if result.error]
    should_save = bool(events) or settings.save_noop

    summary_path = None
    if should_save:
        if events:
            for result in results:
                if not result.event and result.current_price > 0 and result.ticker in open_positions:
                    refresh_position(open_positions[result.ticker], result.current_price, currency_rate)
        write_open_positions(wb, open_positions)
        cash = compute_cash(wb, starting_capital)
        exposure = sum(float(pos.get("exposure_ils") or 0) for pos in open_positions.values())
        open_risk = sum(float(pos.get("risk_ils") or 0) for pos in open_positions.values())
        summary_path = write_summary(
            settings=settings,
            run_id=run_id,
            timestamp=timestamp,
            results=results,
            cash=cash,
            exposure=exposure,
            open_risk=open_risk,
            currency=currency,
        )
        append_update_log(
            wb,
            timestamp=timestamp,
            run_id=f"monitor_{run_id}",
            tickers=list(open_positions.keys()) or [result.ticker for result in results],
            actions=events,
            cash=cash,
            exposure=exposure,
            open_risk=open_risk,
            open_positions=len(open_positions),
            summary_path=summary_path,
        )
        wb.save(settings.excel_path)
        if events:
            send_position_event_notifications(
                event_notifications,
                settings=settings,
                run_id=run_id,
                timestamp=timestamp,
            )

    if events:
        print(f"Position monitor completed with {len(events)} event(s).")
    elif errors:
        print(f"Position monitor completed with no portfolio events and {len(errors)} data issue(s).")
    else:
        print("Position monitor completed with no portfolio events.")
    if summary_path:
        print(f"Summary: {summary_path}")


def load_settings() -> MonitorSettings:
    load_dotenv(ROOT / ".env")
    return MonitorSettings(
        excel_path=Path(os.getenv("MARKET_LENS_EXCEL_PATH", DEFAULT_TRACKER)),
        run_dir=Path(os.getenv("MARKET_LENS_RUN_DIR", DEFAULT_RUN_DIR)),
        period=os.getenv("MARKET_LENS_MONITOR_PERIOD", "5d"),
        interval=os.getenv("MARKET_LENS_MONITOR_INTERVAL", "1m"),
        save_noop=os.getenv("MARKET_LENS_MONITOR_SAVE_NOOP", "false").lower() in {"1", "true", "yes"},
        dashboard_url=dashboard_url_from_env(os.getenv("MARKET_LENS_URL", "")),
    )


def send_position_event_notifications(
    event_notifications: list[tuple[dict[str, Any], PositionEvent]],
    *,
    settings: MonitorSettings,
    run_id: str,
    timestamp: str,
) -> None:
    for position, event in event_notifications:
        message = format_position_event_message(
            position=position,
            event=event,
            run_id=run_id,
            timestamp=timestamp,
            dashboard_url=settings.dashboard_url,
        )
        outcome = send_telegram_message(message)
        if outcome.sent:
            print(f"Telegram position-event notification sent for {event.ticker}:{event.action}.")
        elif outcome.status != "not_configured":
            print(f"Telegram position-event notification skipped for {event.ticker}:{event.action}: {outcome.reason}")


def monitor_position(
    position: dict[str, Any],
    *,
    settings: MonitorSettings,
    since: datetime | None,
    currency_rate: float,
) -> MonitorResult:
    ticker = str(position.get("ticker") or "").upper()
    try:
        frame = fetch_intraday_frame(ticker, period=settings.period, interval=settings.interval)
    except Exception as exc:
        return MonitorResult(ticker=ticker, status="DATA_ERROR", current_price=0.0, error=str(exc))

    if frame.empty:
        return MonitorResult(ticker=ticker, status="NO_DATA", current_price=0.0)

    latest_close = float(frame["Close"].iloc[-1])
    bars = frame
    if since:
        cutoff = ensure_utc(since)
        bars = frame[frame.index > cutoff]
    if bars.empty:
        return MonitorResult(ticker=ticker, status="NO_NEW_BARS", current_price=latest_close)

    stop = float(position.get("stop_loss") or 0)
    target_1 = float(position.get("target_1") or 0)
    target_2 = float(position.get("target_2") or 0)
    partial_taken = bool(position.get("partial_taken"))
    quantity = int(position.get("quantity") or 0)
    update_position_excursion(position, bars)

    for index, row in bars.iterrows():
        high = float(row["High"])
        low = float(row["Low"])
        close = float(row["Close"])
        hit_stop = stop > 0 and low <= stop
        hit_target_2 = target_2 > 0 and high >= target_2
        hit_target_1 = target_1 > 0 and high >= target_1 and not partial_taken

        if hit_stop:
            note = "Stop loss touched by intraday low."
            if hit_target_1 or hit_target_2:
                note += " Same candle also touched a target; conservative stop-first policy applied."
            event = build_event(
                position,
                action="EXIT_STOP",
                triggered_at=index.isoformat(),
                trigger_price=stop,
                high=high,
                low=low,
                close=close,
                quantity=quantity,
                currency_rate=currency_rate,
                note=note,
            )
            return MonitorResult(ticker=ticker, status="EVENT", current_price=close, event=event)

        if hit_target_2:
            event = build_event(
                position,
                action="TAKE_PROFIT",
                triggered_at=index.isoformat(),
                trigger_price=target_2,
                high=high,
                low=low,
                close=close,
                quantity=quantity,
                currency_rate=currency_rate,
                note="Target 2 touched by intraday high; closing remaining simulated position.",
            )
            return MonitorResult(ticker=ticker, status="EVENT", current_price=close, event=event)

        if hit_target_1:
            closed_qty = max(1, quantity // 2)
            event = build_event(
                position,
                action="TAKE_PARTIAL_PROFIT",
                triggered_at=index.isoformat(),
                trigger_price=target_1,
                high=high,
                low=low,
                close=close,
                quantity=closed_qty,
                currency_rate=currency_rate,
                note="Target 1 touched by intraday high; taking partial profit and moving stop to breakeven.",
            )
            return MonitorResult(ticker=ticker, status="EVENT", current_price=close, event=event)

    return MonitorResult(ticker=ticker, status="HOLD", current_price=latest_close)


def build_event(
    position: dict[str, Any],
    *,
    action: str,
    triggered_at: str,
    trigger_price: float,
    high: float,
    low: float,
    close: float,
    quantity: int,
    currency_rate: float,
    note: str,
) -> PositionEvent:
    return PositionEvent(
        ticker=str(position.get("ticker") or "").upper(),
        action=action,
        triggered_at=triggered_at,
        trigger_price=round(trigger_price, 4),
        high=round(high, 4),
        low=round(low, 4),
        close=round(close, 4),
        quantity=quantity,
        cash_in=round(quantity * trigger_price * currency_rate, 2),
        note=note,
    )


def apply_event(
    wb: Any,
    open_positions: dict[str, dict[str, Any]],
    position: dict[str, Any],
    event: PositionEvent,
    timestamp: str,
    run_id: str,
    currency_rate: float,
) -> None:
    append_position_event(wb, timestamp, run_id, position, event)
    append_trade_log_row(wb, timestamp, position, event, currency_rate)

    ticker = event.ticker
    if event.action == "TAKE_PARTIAL_PROFIT":
        remaining_qty = max(0, int(position.get("quantity") or 0) - event.quantity)
        if remaining_qty <= 0:
            open_positions.pop(ticker, None)
            return
        position["quantity"] = remaining_qty
        position["current_price"] = event.trigger_price
        position["stop_loss"] = float(position.get("entry_price") or event.trigger_price)
        position["notes"] = "Partial profit taken; stop moved to breakeven."
        refresh_position(position, event.trigger_price, currency_rate)
        return

    open_positions.pop(ticker, None)


def refresh_position(position: dict[str, Any], current_price: float, currency_rate: float) -> None:
    quantity = int(position.get("quantity") or 0)
    entry = float(position.get("entry_price") or 0)
    stop = float(position.get("stop_loss") or 0)
    position["current_price"] = round(current_price, 4)
    position["unrealized_usd"] = round((current_price - entry) * quantity, 2)
    position["unrealized_ils"] = round((current_price - entry) * quantity * currency_rate, 2)
    position["exposure_ils"] = round(current_price * quantity * currency_rate, 2)
    position["risk_ils"] = round(max(0.0, current_price - stop) * quantity * currency_rate, 2)
    position["notes"] = position.get("notes") or "Open position refreshed by monitor."


def update_position_excursion(position: dict[str, Any], bars: Any) -> None:
    if bars is None or bars.empty:
        return
    entry = float(position.get("entry_price") or 0)
    quantity = int(position.get("quantity") or 0)
    if entry <= 0 or quantity <= 0:
        return
    high = float(bars["High"].max())
    low = float(bars["Low"].min())
    decision = parse_decision_json(position.get("decision_json", ""))
    existing_mfe = float(decision.get("mfe") or 0)
    existing_mae = float(decision.get("mae") or 0)
    decision["mfe"] = round(max(existing_mfe, max(0.0, high - entry) * quantity), 2)
    decision["mae"] = round(max(existing_mae, max(0.0, entry - low) * quantity), 2)
    position["decision_json"] = json.dumps(decision, ensure_ascii=False, sort_keys=True, default=str)


def parse_decision_json(value: Any) -> dict[str, Any]:
    if not value:
        return {}
    try:
        parsed = json.loads(str(value))
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def event_r_multiple(position: dict[str, Any], event: PositionEvent) -> float | None:
    entry = float(position.get("entry_price") or 0)
    stop = float(position.get("stop_loss") or 0)
    if entry <= 0 or stop <= 0 or entry <= stop:
        return None
    return round((event.trigger_price - entry) / (entry - stop), 4)


def trade_analytics_values(decision: dict[str, Any]) -> list[Any]:
    return [
        decision.get("trade_id", ""),
        decision.get("setup_score_bucket", ""),
        decision.get("entry_confirmation_status") or decision.get("confirmation_status", ""),
        decision.get("mfe"),
        decision.get("mae"),
        decision.get("r_multiple"),
        decision.get("duration"),
        decision.get("exit_reason"),
        decision.get("outcome_after_1d"),
        decision.get("outcome_after_3d"),
        decision.get("outcome_after_5d"),
        decision.get("outcome_after_10d"),
    ]


def read_settings(wb: Any) -> dict[str, Any]:
    ws = wb["Settings"]
    values = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            values[str(row[0])] = row[1]
    return values


def read_open_positions(wb: Any) -> dict[str, dict[str, Any]]:
    ws = wb["Open Positions"]
    positions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        positions[str(row[0]).upper()] = {
            "ticker": str(row[0]).upper(),
            "entry_date": row[1],
            "entry_price": float(row[2] or 0),
            "current_price": float(row[3] or row[2] or 0),
            "quantity": int(row[4] or 0),
            "stop_loss": float(row[5] or 0),
            "target_1": float(row[6] or 0),
            "target_2": float(row[7] or 0),
            "status": row[8] or "OPEN",
            "unrealized_usd": float(row[9] or 0),
            "unrealized_ils": float(row[10] or 0),
            "exposure_ils": float(row[11] or 0),
            "risk_ils": float(row[12] or 0),
            "notes": str(row[13] or ""),
            "screenshot": str(row[14] or ""),
            "chart_url": str(row[15] or ""),
            "selection_context": str(row[16] or ""),
            "decision_json": str(row[17] or ""),
            "partial_taken": "partial" in str(row[13] or "").lower(),
        }
    return positions


def write_open_positions(wb: Any, positions: dict[str, dict[str, Any]]) -> None:
    ws = wb["Open Positions"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row_idx, position in enumerate(positions.values(), start=2):
        values = [
            position["ticker"],
            position["entry_date"],
            position["entry_price"],
            position["current_price"],
            position["quantity"],
            position["stop_loss"],
            position["target_1"],
            position["target_2"],
            position["status"],
            position["unrealized_usd"],
            position["unrealized_ils"],
            position["exposure_ils"],
            position["risk_ils"],
            position.get("notes", ""),
            position.get("screenshot", ""),
            position.get("chart_url", ""),
            position.get("selection_context", ""),
            position.get("decision_json", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)


def ensure_agent_columns(wb: Any) -> None:
    headers = {
        "Trade Log": {
            18: "Chart URL",
            19: "Selection Context",
            20: "Decision JSON",
            21: "Trade ID",
            22: "Setup Score Bucket",
            23: "Entry Confirmation",
            24: "MFE",
            25: "MAE",
            26: "R Multiple",
            27: "Duration",
            28: "Exit Reason",
            29: "Outcome After 1D",
            30: "Outcome After 3D",
            31: "Outcome After 5D",
            32: "Outcome After 10D",
        },
        "Open Positions": {
            16: "Chart URL",
            17: "Selection Context",
            18: "Decision JSON",
        },
    }
    for sheet_name, sheet_headers in headers.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for col_idx, header in sheet_headers.items():
            ws.cell(1, col_idx, header)


def ensure_position_events_sheet(wb: Any) -> None:
    if EVENT_SHEET in wb.sheetnames:
        return
    ws = wb.create_sheet(EVENT_SHEET)
    headers = [
        "Timestamp",
        "Run ID",
        "Ticker",
        "Action",
        "Triggered At",
        "Trigger Price USD",
        "Bar High USD",
        "Bar Low USD",
        "Bar Close USD",
        "Quantity",
        "Cash In",
        "Stop Loss USD",
        "Target 1 USD",
        "Target 2 USD",
        "Notes",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx, header)


def read_last_event_times(wb: Any) -> dict[str, datetime]:
    if EVENT_SHEET not in wb.sheetnames:
        return {}
    latest: dict[str, datetime] = {}
    ws = wb[EVENT_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = str(row[2] or "").upper()
        triggered_at = parse_timestamp(row[4])
        if ticker and triggered_at and triggered_at > latest.get(ticker, datetime.min.replace(tzinfo=timezone.utc)):
            latest[ticker] = triggered_at
    return latest


def append_position_event(
    wb: Any,
    timestamp: str,
    run_id: str,
    position: dict[str, Any],
    event: PositionEvent,
) -> None:
    ws = wb[EVENT_SHEET]
    row = next_row(ws)
    values = [
        timestamp,
        run_id,
        event.ticker,
        event.action,
        event.triggered_at,
        event.trigger_price,
        event.high,
        event.low,
        event.close,
        event.quantity,
        event.cash_in,
        position.get("stop_loss"),
        position.get("target_1"),
        position.get("target_2"),
        event.note,
    ]
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row, col_idx, value)


def append_trade_log_row(
    wb: Any,
    timestamp: str,
    position: dict[str, Any],
    event: PositionEvent,
    currency_rate: float,
) -> None:
    ws = wb["Trade Log"]
    row = next_row(ws)
    ws.cell(row, 1, timestamp)
    ws.cell(row, 2, event.action)
    ws.cell(row, 3, event.ticker)
    ws.cell(row, 4, None)
    ws.cell(row, 5, event.trigger_price)
    ws.cell(row, 6, event.quantity)
    ws.cell(row, 7, currency_rate)
    ws.cell(row, 8, 0)
    ws.cell(row, 9, event.cash_in)
    ws.cell(row, 10, 0)
    ws.cell(row, 11, event.cash_in)
    ws.cell(row, 12, position.get("stop_loss"))
    ws.cell(row, 13, position.get("target_1"))
    ws.cell(row, 14, position.get("target_2"))
    ws.cell(row, 15, 0)
    ws.cell(row, 16, event.note)
    ws.cell(row, 17, position.get("screenshot", ""))
    ws.cell(row, 18, position.get("chart_url", ""))
    ws.cell(row, 19, position.get("selection_context", ""))
    ws.cell(row, 20, position.get("decision_json", ""))
    analytics = parse_decision_json(position.get("decision_json", ""))
    analytics["exit_reason"] = event.action
    analytics["r_multiple"] = event_r_multiple(position, event)
    for col_idx, value in enumerate(trade_analytics_values(analytics), start=21):
        ws.cell(row, col_idx, value)


def append_update_log(
    wb: Any,
    *,
    timestamp: str,
    run_id: str,
    tickers: list[str],
    actions: list[PositionEvent],
    cash: float,
    exposure: float,
    open_risk: float,
    open_positions: int,
    summary_path: Path,
) -> None:
    ws = wb["Update Log"]
    row = next_row(ws)
    actions_summary = ", ".join(f"{event.ticker}:{event.action}" for event in actions) or "MONITOR:HOLD"
    ws.cell(row, 1, timestamp)
    ws.cell(row, 2, run_id)
    ws.cell(row, 3, " ".join(tickers))
    ws.cell(row, 4, 0)
    ws.cell(row, 5, actions_summary)
    ws.cell(row, 6, cash)
    ws.cell(row, 7, exposure)
    ws.cell(row, 8, open_risk)
    ws.cell(row, 9, open_positions)
    ws.cell(row, 10, portable_path(summary_path))
    ws.cell(row, 11, "")


def compute_cash(wb: Any, starting_capital: float) -> float:
    ws = wb["Trade Log"]
    cash_out = 0.0
    cash_in = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cash_out += float(row[9] or 0)
        cash_in += float(row[10] or 0)
    return round(starting_capital - cash_out + cash_in, 2)


def write_summary(
    *,
    settings: MonitorSettings,
    run_id: str,
    timestamp: str,
    results: list[MonitorResult],
    cash: float,
    exposure: float,
    open_risk: float,
    currency: str,
) -> Path:
    summary_dir = settings.run_dir / "position_monitor"
    summary_dir.mkdir(parents=True, exist_ok=True)
    summary_path = summary_dir / f"position_monitor_{run_id}.md"
    events = [result.event for result in results if result.event]
    errors = [result for result in results if result.error]
    lines = [
        "Market Lens Position Monitor Update",
        "",
        f"Date: {timestamp}",
        f"Run status: {'OK' if not errors else 'ISSUES'}",
        f"Open positions checked: {len(results)}",
        f"Events found: {len(events)}",
        f"Actions taken: {', '.join(f'{event.ticker}:{event.action}' for event in events) or 'None'}",
        f"Cash remaining: {cash:.2f} {currency}",
        f"Current exposure: {exposure:.2f} {currency}",
        f"Total open risk: {open_risk:.2f} {currency}",
        f"Excel updated: {settings.excel_path}",
        f"Errors: {'; '.join(f'{result.ticker}: {result.error}' for result in errors) if errors else 'None'}",
        "Agent feedback:",
    ]
    for result in results:
        if result.event:
            lines.append(f"- {result.event.ticker}: {result.event.action} - {result.event.note}")
        elif result.error:
            lines.append(f"- {result.ticker}: DATA_ERROR - {result.error}")
        else:
            lines.append(f"- {result.ticker}: {result.status} - no target or stop event detected.")
    summary_path.write_text("\n".join(lines), encoding="utf-8")
    return summary_path


def parse_timestamp(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return ensure_utc(value)
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return ensure_utc(datetime.fromisoformat(text.replace("Z", "+00:00")))
    except ValueError:
        try:
            return ensure_utc(datetime.strptime(text[:19], "%Y-%m-%dT%H:%M:%S"))
        except ValueError:
            return None


def ensure_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def next_row(ws: Any) -> int:
    for row in range(2, ws.max_row + 2):
        if ws.cell(row, 1).value is None:
            return row
    return ws.max_row + 1


def portable_path(path: Path) -> str:
    try:
        return path.relative_to(ROOT).as_posix()
    except ValueError:
        return path.as_posix()


if __name__ == "__main__":
    main()
