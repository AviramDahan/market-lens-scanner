from __future__ import annotations

import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import urlencode, urljoin
from urllib.request import Request, urlopen

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright

from app.charts import write_scan_chart
from app.agent_risk import build_agent_run_context, evaluate_agent_candidate
from app.performance_summary import write_performance_summaries
from app.scanner import scan_ticker_detail
from app.shadow_strategies import evaluate_shadow_strategies
from app.smart_universe import base_universe, build_sector_health, build_smart_universe, curated_universe
from app.strategy import StrategyDecision as Decision
from app.strategy import decide_strategy_candidate, normalize_strategy_candidate
from app.telegram_notifications import (
    dashboard_url_from_app_url,
    format_position_opened_message,
    send_telegram_message,
)


ROOT = Path(__file__).resolve().parents[1]
RUN_DIR = Path(os.getenv("MARKET_LENS_RUN_DIR", ROOT / "agent_runs"))
SCREENSHOT_DIR = RUN_DIR / "screenshots"
SUMMARY_DIR = RUN_DIR / "summaries"
CHART_DIR = RUN_DIR / "charts"
DECISION_DIR = RUN_DIR / "decisions"
APP_READY_SELECTOR = '[data-testid="auth-email"], #authStatus'


@dataclass
class Settings:
    url: str
    email: str
    password: str
    excel_path: Path
    universe: str
    tickers: list[str]
    analysis_period: str
    min_rr: float
    headless: bool
    timeout_seconds: int


@dataclass
class SetupResult:
    ticker: str
    setup_type: str
    score: float
    current_price: float
    buy_zone_low: float | None
    buy_zone_high: float | None
    stop_loss: float | None
    target_1: float | None
    target_2: float | None
    risk_reward: float
    reason: str
    raw_text: str
    chart_url: str = ""
    selection_context: str = ""


@dataclass(frozen=True)
class ChartRetentionSettings:
    save_rejected_charts: bool
    rejected_chart_limit: int
    rejected_chart_min_score: float


def main() -> None:
    settings = load_settings()
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    log(f"Agent run started: {run_id}")
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    SUMMARY_DIR.mkdir(parents=True, exist_ok=True)
    CHART_DIR.mkdir(parents=True, exist_ok=True)
    DECISION_DIR.mkdir(parents=True, exist_ok=True)
    screenshot_path = SCREENSHOT_DIR / f"market_lens_agent_{run_id}.png"
    summary_path = SUMMARY_DIR / f"market_lens_agent_{run_id}.md"
    decision_path = DECISION_DIR / f"market_lens_agent_{run_id}.jsonl"

    errors: list[str] = []
    results: list[SetupResult] = []
    decisions: dict[str, Decision] = {}
    login_status = "not_started"
    scan_status = "not_started"
    auth_failed = False
    deadline = time.monotonic() + settings.timeout_seconds

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=settings.headless)
        page = browser.new_page(viewport={"width": 1440, "height": 1200})
        try:
            log("Opening Market Lens UI")
            open_app(page, settings.url, deadline)
            log("Logging in")
            login_status = login(page, settings, deadline)
            log(f"Login status: {login_status}")
            log("Configuring scan")
            scan_tickers = configure_scan(page, settings, deadline)
            log("Running UI scan")
            results = run_scan_batches(page, scan_tickers, deadline)
            log(f"Scan completed: {len(results)} results")
            scan_status = f"completed: {len(results)} results"
            log("Saving screenshot")
            page.screenshot(path=str(screenshot_path), full_page=True)
        except Exception as exc:
            message = str(exc)
            if is_auth_failure(message):
                auth_failed = True
                login_status = "AUTH_FAILED"
                scan_status = "auth_failed"
                errors.append(f"AUTH_FAILED: {message}")
            else:
                errors.append(message)
                scan_status = "failed"
            log(f"Agent UI phase failed: {exc}")
            try:
                page.screenshot(path=str(screenshot_path), full_page=True)
            except Exception:
                pass
        finally:
            browser.close()

    scan_completed = scan_status.startswith("completed:")
    if auth_failed or not scan_completed:
        run_status = "AUTH_FAILED" if auth_failed else "RUN_FAILED"
        log(f"{run_status}; skipping workbook/trade updates")
        workbook_context = workbook_snapshot(settings=settings, run_status=run_status)
        decisions = {}
    else:
        log("Updating workbook")
        workbook_context = update_workbook(
            settings=settings,
            run_id=run_id,
            results=results,
            screenshot_path=screenshot_path,
            summary_path=summary_path,
            decision_path=decision_path,
            errors=errors,
        )
        decisions = workbook_context["decisions"]
    log("Writing summary")
    write_summary(
        settings=settings,
        run_id=run_id,
        login_status=login_status,
        scan_status=scan_status,
        results=results,
        decisions=decisions,
        screenshot_path=screenshot_path,
        summary_path=summary_path,
        workbook_context=workbook_context,
        errors=errors,
    )
    print(f"Agent run complete: {summary_path}")
    if workbook_context.get("run_status") in {"AUTH_FAILED", "RUN_FAILED"}:
        raise SystemExit(2)


def load_settings() -> Settings:
    load_dotenv(ROOT / ".env")
    url = os.getenv("MARKET_LENS_URL", "https://market-lens-scanner-fb63.onrender.com/?v=latest")
    email = required_env("MARKET_LENS_EMAIL")
    password = required_env("MARKET_LENS_PASSWORD")
    excel_path = Path(required_env("MARKET_LENS_EXCEL_PATH"))
    universe = os.getenv("MARKET_LENS_UNIVERSE", "smart-universe")
    tickers = parse_tickers(os.getenv("MARKET_LENS_TICKERS", ""))
    analysis_period = os.getenv("MARKET_LENS_ANALYSIS_PERIOD", "6mo")
    min_rr = float(os.getenv("MARKET_LENS_MIN_RR", "2"))
    headless = os.getenv("MARKET_LENS_HEADLESS", "true").lower() not in {"0", "false", "no"}
    timeout_seconds = int(os.getenv("MARKET_LENS_AGENT_TIMEOUT_SECONDS", "720"))
    return Settings(
        url,
        email,
        password,
        excel_path,
        universe,
        tickers,
        analysis_period,
        min_rr,
        headless,
        timeout_seconds,
    )


def required_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def env_bool(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() not in {"0", "false", "no", "off"}


def chart_retention_settings() -> ChartRetentionSettings:
    return ChartRetentionSettings(
        save_rejected_charts=env_bool("MARKET_LENS_SAVE_REJECTED_CHARTS", False),
        rejected_chart_limit=max(0, int(os.getenv("MARKET_LENS_REJECTED_CHART_LIMIT", "5"))),
        rejected_chart_min_score=float(os.getenv("MARKET_LENS_REJECTED_CHART_MIN_SCORE", "0.40")),
    )


def is_auth_failure(message: str) -> bool:
    text = message.lower()
    auth_markers = (
        "auth not configured",
        "account login is not configured",
        "login did not complete",
        "invalid login credentials",
        "email not confirmed",
        "sign in required",
    )
    return any(marker in text for marker in auth_markers)


def auth_config_is_open(config: Any) -> bool:
    if not isinstance(config, dict):
        return False
    mode = str(config.get("mode", "")).strip().lower()
    return mode in {"open", "disabled", "off", "public"} or config.get("enabled") is False


def open_app(page: Page, url: str, run_deadline: float) -> None:
    deadline = min(time.monotonic() + 180, run_deadline)
    last_error = ""
    while time.monotonic() < deadline:
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=remaining_ms(deadline, 90_000))
            page.wait_for_selector(APP_READY_SELECTOR, timeout=remaining_ms(deadline, 30_000))
            return
        except PlaywrightTimeoutError as exc:
            last_error = str(exc)
            page.wait_for_timeout(remaining_ms(deadline, 10_000))
    raise RuntimeError(f"Market Lens app did not become ready after Render wake-up wait. {last_error}")


def login(page: Page, settings: Settings, deadline: float) -> str:
    page.wait_for_selector(APP_READY_SELECTOR, timeout=remaining_ms(deadline, 30_000))
    try:
        auth_config = page.evaluate(
            """async () => {
                const response = await fetch('/auth/config', { cache: 'no-store' });
                return await response.json();
            }"""
        )
        if auth_config_is_open(auth_config):
            return "open access"
    except Exception as exc:
        log(f"Auth config check unavailable; falling back to UI login flow: {exc}")

    page.wait_for_function(
        "() => document.querySelector('#authStatus')?.textContent?.trim().length > 0",
        timeout=remaining_ms(deadline, 60_000),
    )
    auth_status = page.locator("#authStatus").inner_text(timeout=remaining_ms(deadline, 10_000))
    if "open access" in auth_status.lower() or "auth not configured" in auth_status.lower():
        return "open access"
    if settings.email.lower() in auth_status.lower():
        return "already signed in"

    last_message = ""
    for attempt in range(1, 4):
        page.locator('[data-testid="auth-email"]').fill(settings.email)
        page.locator('[data-testid="auth-password"]').fill(settings.password)
        page.locator('[data-testid="sign-in-button"]').click()
        try:
            page.wait_for_function(
                "email => document.querySelector('#authStatus')?.textContent?.toLowerCase().includes(email.toLowerCase())",
                arg=settings.email,
                timeout=remaining_ms(deadline, 45_000),
            )
            return "signed in"
        except PlaywrightTimeoutError:
            last_message = page.locator("#message").inner_text(timeout=remaining_ms(deadline, 10_000))
            if attempt < 3:
                page.wait_for_timeout(remaining_ms(deadline, 5_000 * attempt))

    auth_status = page.locator("#authStatus").inner_text(timeout=remaining_ms(deadline, 10_000))
    raise RuntimeError(f"Login did not complete. Status: {auth_status}. Message: {last_message}")


def configure_scan(page: Page, settings: Settings, deadline: float) -> list[str]:
    page.locator('[data-testid="min-rr-input"]').fill(str(settings.min_rr))
    page.locator('[data-testid="analysis-period-select"]').select_option(settings.analysis_period)
    open_tickers = read_open_position_tickers(settings.excel_path)
    watch_tickers = read_recent_watch_tickers(settings.excel_path)
    carry_forward_tickers = limited_carry_forward_tickers(open_tickers, watch_tickers)
    skipped_tickers = read_recent_skip_tickers(settings.excel_path)
    if carry_forward_tickers:
        log(
            "Carry-forward tickers added outside universe quota: "
            f"{len(carry_forward_tickers)} ({' '.join(carry_forward_tickers)})"
        )
    if skipped_tickers:
        log(
            "Recent SKIP tickers excluded from base universe: "
            f"{len(skipped_tickers)} ({' '.join(skipped_tickers[:20])}"
            f"{' ...' if len(skipped_tickers) > 20 else ''})"
        )
    if settings.tickers:
        tickers = unique_tickers(settings.tickers + carry_forward_tickers)
        set_scan_basket(page, tickers)
        return tickers

    agent_tickers = build_agent_scan_tickers(settings, carry_forward_tickers, skipped_tickers)
    if agent_tickers:
        set_scan_basket(page, agent_tickers)
        log(f"Agent selected {len(agent_tickers)} tickers: {' '.join(agent_tickers)}")
        return agent_tickers

    log("Smart Universe API selection unavailable; falling back to UI select-all flow.")
    page.locator('[data-testid="universe-select"]').select_option(settings.universe)
    page.wait_for_function(
        "() => !document.querySelector('[data-testid=\"select-all-tickers\"]')?.disabled",
        timeout=remaining_ms(deadline, 240_000),
    )
    page.locator('[data-testid="select-all-tickers"]').click()
    page.locator('[data-testid="add-selected-tickers"]').click()
    if carry_forward_tickers:
        page.locator('[data-testid="manual-ticker-input"]').fill(" ".join(carry_forward_tickers))
        page.locator('[data-testid="add-manual-ticker"]').click()
    return current_scan_basket(page)


def limited_carry_forward_tickers(open_tickers: list[str], watch_tickers: list[str]) -> list[str]:
    open_unique = unique_tickers(open_tickers)
    watch_unique = [ticker for ticker in unique_tickers(watch_tickers) if ticker not in open_unique]
    limit = int(os.getenv("MARKET_LENS_AGENT_CARRY_FORWARD_LIMIT", "30"))
    if limit <= 0:
        return unique_tickers(open_unique + watch_unique)

    remaining = max(0, limit - len(open_unique))
    selected_watch = watch_unique[:remaining]
    dropped = len(watch_unique) - len(selected_watch)
    if dropped > 0:
        log(
            "Carry-forward WATCH tickers limited to protect scan stability: "
            f"kept {len(selected_watch)} and deferred {dropped}."
        )
    return unique_tickers(open_unique + selected_watch)


def set_scan_basket(page: Page, tickers: list[str]) -> None:
    page.locator("#clearBasketButton").click()
    if not tickers:
        return
    page.locator('[data-testid="manual-ticker-input"]').fill(" ".join(tickers))
    page.locator('[data-testid="add-manual-ticker"]').click()


def current_scan_basket(page: Page) -> list[str]:
    value = page.locator("#tickers").input_value()
    return parse_tickers(value.replace(",", " "))


def build_agent_scan_tickers(
    settings: Settings,
    carry_forward_tickers: list[str],
    skipped_tickers: list[str],
) -> list[str]:
    if settings.universe != "smart-universe":
        return []
    target_count = int(os.getenv("MARKET_LENS_AGENT_UNIVERSE_TARGET", "35"))
    pool_count = int(os.getenv("MARKET_LENS_AGENT_UNIVERSE_POOL", "100"))
    hard_excluded = set(carry_forward_tickers)
    soft_excluded = set(skipped_tickers)
    max_pool_count = max(
        pool_count,
        target_count,
        int(os.getenv("MARKET_LENS_AGENT_UNIVERSE_MAX_POOL", "300")),
    )
    effective_pool_count = min(
        max_pool_count,
        max(pool_count, target_count + len(hard_excluded) + len(soft_excluded)),
    )
    if effective_pool_count > pool_count:
        log(
            "Smart agent universe pool expanded for exclusions: "
            f"{pool_count} -> {effective_pool_count} candidates."
        )

    candidates = fetch_smart_universe_tickers(settings, effective_pool_count)
    if not candidates:
        return []

    base_tickers = [ticker for ticker in candidates if ticker not in hard_excluded and ticker not in soft_excluded][
        :target_count
    ]
    fallback_tickers: list[str] = []
    if len(base_tickers) < target_count and env_bool("MARKET_LENS_AGENT_RECENT_SKIP_FALLBACK", True):
        needed = target_count - len(base_tickers)
        fallback_tickers = [
            ticker
            for ticker in candidates
            if ticker not in hard_excluded and ticker not in base_tickers
        ][:needed]
        base_tickers = unique_tickers(base_tickers + fallback_tickers)[:target_count]
        if fallback_tickers:
            log(
                "Recent SKIP fallback used to preserve scan breadth: "
                f"{len(fallback_tickers)} tickers ({' '.join(fallback_tickers[:20])}"
                f"{' ...' if len(fallback_tickers) > 20 else ''})."
            )
    final_tickers = unique_tickers(base_tickers + carry_forward_tickers)
    total_limit = int(os.getenv("MARKET_LENS_AGENT_TOTAL_SCAN_LIMIT", "0") or "0")
    if total_limit > 0 and len(final_tickers) > total_limit:
        deferred = final_tickers[total_limit:]
        final_tickers = final_tickers[:total_limit]
        log(
            "Agent total scan limit applied: "
            f"kept {len(final_tickers)} tickers and deferred {len(deferred)} "
            f"to protect production stability."
        )
    log(
        "Smart agent universe built: "
        f"{len(base_tickers)} base scan tickers + "
        f"{len([ticker for ticker in carry_forward_tickers if ticker not in base_tickers])} carry-forward tickers."
    )
    if len(base_tickers) < target_count:
        log(f"Smart agent universe warning: only {len(base_tickers)} fresh tickers available after exclusions.")
    return final_tickers


def fetch_smart_universe_tickers(settings: Settings, limit: int) -> list[str]:
    params = urlencode(
        {
            "analysis_period": settings.analysis_period,
            "limit": max(35, min(300, limit)),
            "max_per_sector": int(os.getenv("MARKET_LENS_AGENT_MAX_PER_SECTOR", "10")),
        }
    )
    url = urljoin(settings.url, f"/smart-universe?{params}")
    try:
        request = Request(url, headers={"User-Agent": "MarketLensAgent/1.0"})
        with urlopen(request, timeout=90) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        if limit > 100:
            log(
                "Smart Universe API fetch failed for expanded pool "
                f"({limit}); retrying compatibility limit 100: {exc}"
            )
            compatibility = fetch_smart_universe_tickers(settings, 100)
            if compatibility:
                return compatibility
        log(f"Smart Universe API fetch failed; using local fallback: {exc}")
        return local_smart_universe_tickers(settings, limit)

    ranked = (payload.get("companies") or []) + (payload.get("ranked") or [])
    tickers: list[str] = []
    for item in ranked:
        ticker = item.get("ticker") if isinstance(item, dict) else item
        if ticker:
            tickers.append(str(ticker).upper().strip())
    return unique_tickers(tickers)


def local_smart_universe_tickers(settings: Settings, limit: int) -> list[str]:
    safe_limit = max(35, min(300, limit))
    if env_bool("MARKET_LENS_AGENT_LOCAL_SMART_FALLBACK", False):
        max_per_sector = int(os.getenv("MARKET_LENS_AGENT_MAX_PER_SECTOR", "10"))
        try:
            payload = build_smart_universe(
                analysis_period=settings.analysis_period,
                limit=safe_limit,
                max_per_sector=max_per_sector,
            )
            tickers = tickers_from_smart_payload(payload)
            if tickers:
                log(f"Local Smart Universe fallback produced {len(tickers)} tickers.")
                return tickers
        except Exception as exc:
            log(f"Local Smart Universe fallback failed: {exc}")

    fallback = unique_tickers(list(curated_universe().keys()))[:safe_limit]
    if fallback:
        log(f"Curated universe fallback produced {len(fallback)} tickers.")
    return fallback


def tickers_from_smart_payload(payload: dict[str, Any]) -> list[str]:
    return tickers_from_smart_payload(payload)


def run_scan(page: Page, deadline: float) -> list[SetupResult]:
    last_error = ""
    for attempt in range(1, 4):
        page.evaluate("() => { window.marketLensIncludeCharts = false; }")
        page.locator('[data-testid="scan-button"]').click()
        page.wait_for_function(
            "() => !document.querySelector('[data-testid=\"scan-button\"]')?.disabled",
            timeout=remaining_ms(deadline, 600_000),
        )
        if "failed" not in page.locator("#runMeta").inner_text(timeout=remaining_ms(deadline, 10_000)).lower():
            break

        last_error = page.locator("#message").inner_text(timeout=remaining_ms(deadline, 10_000))
        if attempt < 3 and is_transient_scan_error(last_error):
            page.wait_for_timeout(remaining_ms(deadline, 15_000 * attempt))
            continue
        raise RuntimeError(last_error)
    else:
        raise RuntimeError(last_error or "Scan failed")

    page.wait_for_selector('[data-testid="result-card"]', timeout=remaining_ms(deadline, 30_000))
    cards = page.locator('[data-testid="result-card"]')
    extracted: list[dict[str, Any]] = cards.evaluate_all(
        """cards => cards.map(card => {
            const stat = label => {
              const el = Array.from(card.querySelectorAll('[data-stat-label]'))
                .find(node => node.getAttribute('data-stat-label') === label);
              return el?.querySelector('strong')?.textContent?.trim() || '';
            };
            return {
              ticker: card.querySelector('[data-testid="result-ticker"]')?.textContent?.trim() || '',
              setup_type: card.querySelector('[data-testid="result-setup-type"]')?.textContent?.trim() || '',
              score: card.querySelector('.score-box strong')?.textContent?.trim() || '',
              current_price: stat('Setup price') || stat('Price'),
              buy_zone: stat('Buy Zone'),
              stop_loss: stat('Stop'),
              targets: stat('Targets'),
              risk_reward: stat('R/R'),
              reason: card.querySelector('[data-testid="result-reason"]')?.textContent?.trim() || '',
              chart_url: (card.querySelector('img.chart-preview')?.getAttribute('data-chart')
                || card.querySelector('img.chart-preview')?.getAttribute('src')
                || '').split('?')[0],
              raw_text: card.innerText,
            };
          })"""
    )
    return [parse_result(item) for item in extracted]


def run_scan_batches(page: Page, tickers: list[str], deadline: float) -> list[SetupResult]:
    batch_size = max(1, int(os.getenv("MARKET_LENS_AGENT_SCAN_BATCH_SIZE", "20")))
    batch_pause_ms = max(0, int(os.getenv("MARKET_LENS_AGENT_BATCH_PAUSE_MS", "0") or "0"))
    if not tickers or len(tickers) <= batch_size:
        return run_scan(page, deadline)

    batches = list(chunked(tickers, batch_size))
    combined: dict[str, SetupResult] = {}
    for index, batch in enumerate(batches, start=1):
        log(f"Running UI scan batch {index}/{len(batches)} ({len(batch)} tickers): {' '.join(batch)}")
        set_scan_basket(page, batch)
        batch_results = run_scan_batch_resilient(page, batch, deadline)
        log(f"Batch {index}/{len(batches)} completed: {len(batch_results)} result cards")
        for result in batch_results:
            combined[result.ticker] = result
        if batch_pause_ms and index < len(batches):
            page.wait_for_timeout(remaining_ms(deadline, batch_pause_ms))
    return list(combined.values())


def run_scan_batch_resilient(page: Page, batch: list[str], deadline: float) -> list[SetupResult]:
    try:
        return run_scan(page, deadline)
    except RuntimeError as exc:
        message = str(exc)
        min_split_size = max(2, int(os.getenv("MARKET_LENS_AGENT_MIN_SPLIT_BATCH_SIZE", "6")))
        if len(batch) <= min_split_size or not is_transient_scan_error(message):
            raise

        midpoint = max(1, len(batch) // 2)
        parts = [batch[:midpoint], batch[midpoint:]]
        log(
            "Transient scan error; splitting batch into smaller retries: "
            f"{message}; parts={len(parts)} size={len(parts[0])}/{len(parts[1])}."
        )
        retry_pause_ms = max(0, int(os.getenv("MARKET_LENS_AGENT_RETRY_PAUSE_MS", "15000") or "0"))
        if retry_pause_ms:
            page.wait_for_timeout(remaining_ms(deadline, retry_pause_ms))
        results: dict[str, SetupResult] = {}
        for part in parts:
            set_scan_basket(page, part)
            for result in run_scan_batch_resilient(page, part, deadline):
                results[result.ticker] = result
        return list(results.values())


def chunked(items: list[str], size: int) -> list[list[str]]:
    return [items[index : index + size] for index in range(0, len(items), size)]


def apply_chart_retention_policy(
    pending: list[tuple[SetupResult, Decision]],
    *,
    base_url: str,
    run_id: str,
    analysis_period: str,
    min_rr: float,
    open_position_tickers: set[str],
) -> None:
    settings = chart_retention_settings()
    selected = select_chart_tickers(pending, settings=settings, open_position_tickers=open_position_tickers)
    deadline = time.monotonic() + 120
    for result, _decision in pending:
        if result.ticker not in selected:
            result.chart_url = ""
            continue
        if not persist_chart_image(
            result,
            base_url,
            run_id,
            deadline,
            analysis_period=analysis_period,
            min_rr=min_rr,
        ):
            result.chart_url = ""


def select_chart_tickers(
    pending: list[tuple[SetupResult, Decision]],
    *,
    settings: ChartRetentionSettings,
    open_position_tickers: set[str] | None = None,
) -> set[str]:
    open_position_tickers = open_position_tickers or set()
    selected: set[str] = set()
    rejected_candidates: list[tuple[float, str]] = []
    always_actions = {
        "BUY_SIMULATED",
        "HOLD",
        "WATCH_READY",
        "TAKE_PARTIAL_PROFIT",
        "TAKE_PROFIT",
        "EXIT_STOP",
    }

    for result, decision in pending:
        ticker = result.ticker
        action = str(decision.action or "").upper()
        decision_json = decision.decision_json or {}
        if action in always_actions or ticker in open_position_tickers:
            selected.add(ticker)
            continue
        if action not in {"WATCH", "SKIP"}:
            continue
        setup_type = str(decision_json.get("setup_type") or result.setup_type or "")
        if setup_type == "No Trade":
            continue
        setup_score = float(decision_json.get("setup_score") or result.score or 0)
        if setup_score < settings.rejected_chart_min_score:
            continue
        if str(decision_json.get("sector_regime") or "").upper() == "WEAK":
            continue
        net_rr = parse_optional_float(decision_json.get("net_rr"))
        if net_rr is not None and net_rr < 0.80:
            continue
        rank = float(decision_json.get("final_display_score") or setup_score)
        if settings.save_rejected_charts:
            selected.add(ticker)
        else:
            rejected_candidates.append((rank, ticker))

    if not settings.save_rejected_charts and settings.rejected_chart_limit:
        rejected_candidates.sort(reverse=True)
        selected.update(ticker for _rank, ticker in rejected_candidates[: settings.rejected_chart_limit])
    return selected


def persist_chart_image(
    result: SetupResult,
    base_url: str,
    run_id: str,
    deadline: float,
    *,
    analysis_period: str,
    min_rr: float,
) -> bool:
    destination = CHART_DIR / f"market_lens_agent_{run_id}_{result.ticker.lower()}.png"
    source_url = urljoin(base_url, result.chart_url)
    if result.chart_url:
        try:
            request = Request(source_url, headers={"User-Agent": "market-lens-agent/1.0"})
            with urlopen(request, timeout=max(1, remaining_ms(deadline, 60_000) // 1000)) as response:
                content_type = response.headers.get("content-type", "")
                data = response.read()
            if data and "image" in content_type.lower():
                destination.write_bytes(data)
                result.chart_url = str(destination)
                return True
        except Exception:
            pass

    return generate_local_chart_image(result, destination, analysis_period=analysis_period, min_rr=min_rr)


def generate_local_chart_image(
    result: SetupResult,
    destination: Path,
    *,
    analysis_period: str,
    min_rr: float,
) -> bool:
    try:
        detail = scan_ticker_detail(result.ticker, min_rr=min_rr, analysis_period=analysis_period)
        update: dict[str, Any] = {
            "setup_type": result.setup_type,
            "score": result.score,
            "current_price": result.current_price,
            "risk_reward": result.risk_reward,
        }
        if result.buy_zone_low is not None and result.buy_zone_high is not None:
            update["buy_zone"] = (result.buy_zone_low, result.buy_zone_high)
        if result.stop_loss is not None:
            update["stop_loss"] = result.stop_loss
        if result.target_1 is not None:
            update["target_1"] = result.target_1
        if result.target_2 is not None:
            update["target_2"] = result.target_2
        detail.result = detail.result.model_copy(update=update)
        generated = write_scan_chart(detail, CHART_DIR)
        if generated != destination:
            destination.write_bytes(generated.read_bytes())
            try:
                generated.unlink()
            except OSError:
                pass
        result.chart_url = str(destination)
        return True
    except Exception as exc:
        log(f"Chart generation skipped for {result.ticker}: {exc}")
        return False


def is_transient_scan_error(message: str) -> bool:
    text = message.lower()
    return any(marker in text for marker in ("502", "503", "504", "timeout", "network"))


def remaining_ms(deadline: float, requested_ms: int) -> int:
    remaining = int((deadline - time.monotonic()) * 1000)
    if remaining <= 0:
        raise RuntimeError("Agent run timeout reached before completion.")
    return max(1_000, min(requested_ms, remaining))


def log(message: str) -> None:
    print(f"[{datetime.now().isoformat(timespec='seconds')}] {message}", flush=True)


def parse_result(item: dict[str, Any]) -> SetupResult:
    buy_low, buy_high = parse_range(item.get("buy_zone", ""))
    target_1, target_2 = parse_range(item.get("targets", ""))
    return SetupResult(
        ticker=item.get("ticker", ""),
        setup_type=item.get("setup_type", ""),
        score=parse_float(item.get("score")),
        current_price=parse_float(item.get("current_price") or item.get("setup_price")),
        buy_zone_low=buy_low,
        buy_zone_high=buy_high,
        stop_loss=parse_optional_float(item.get("stop_loss")),
        target_1=target_1,
        target_2=target_2,
        risk_reward=parse_float(str(item.get("risk_reward", "")).replace("x", "")),
        reason=item.get("reason", ""),
        raw_text=item.get("raw_text", ""),
        chart_url=str(item.get("chart_url", "")),
    )


def update_workbook(
    *,
    settings: Settings,
    run_id: str,
    results: list[SetupResult],
    screenshot_path: Path,
    summary_path: Path,
    decision_path: Path,
    errors: list[str],
) -> dict[str, Any]:
    wb = load_workbook(settings.excel_path)
    ensure_agent_columns(wb)
    settings_values = read_settings(wb)
    currency = str(settings_values.get("budget_currency") or "USD").upper()
    currency_rate = 1.0 if currency == "USD" else float(settings_values.get("usd_ils_rate", 3.7))
    starting_capital = float(
        settings_values.get("starting_capital_usd")
        or settings_values.get("starting_capital_ils")
        or 100_000
    )
    max_position = starting_capital * float(settings_values.get("max_position_allocation_pct", 0.1))
    max_total_exposure = starting_capital * float(settings_values.get("max_total_exposure_pct", 0.4))
    max_risk = starting_capital * float(settings_values.get("max_risk_per_trade_pct", 0.01))
    min_rr = float(settings_values.get("min_risk_reward", settings.min_rr))
    sector_map = base_universe()
    run_context = build_agent_run_context(
        analysis_period=settings.analysis_period,
        starting_capital=starting_capital,
        default_max_total_exposure=max_total_exposure,
        max_position=max_position,
    )
    sector_health = run_context.sector_health or build_sector_health(settings.analysis_period)
    recent_stop_events = read_recent_stop_events(wb, run_context.config.stop_cooldown_days)
    neutral_pilot_buys_today = count_neutral_pilot_buys_today(wb)

    open_positions = read_open_positions(wb)
    cash = compute_cash(wb, starting_capital)
    exposure = sum(pos["exposure_ils"] for pos in open_positions.values())
    initial_open_position_tickers = set(open_positions)
    decisions: dict[str, Decision] = {}
    decision_records: list[dict[str, Any]] = []
    pending_decisions: list[tuple[SetupResult, Decision]] = []
    new_buy_tickers: set[str] = set()
    timestamp = datetime.now().isoformat(timespec="seconds")

    for result in results:
        decision = decide(
            result,
            open_positions=open_positions,
            cash=cash,
            exposure=exposure,
            usd_ils=currency_rate,
            max_position=max_position,
            max_total_exposure=max_total_exposure,
            max_risk=max_risk,
            min_rr=min_rr,
            sector_map=sector_map,
            sector_health=sector_health,
        )
        decision_json = evaluate_agent_candidate(
            timestamp=timestamp,
            result=result,
            initial_action=decision.action,
            initial_reason=decision.feedback,
            quantity=decision.quantity,
            cash_out=decision.cash_out_ils,
            risk_amount=decision.risk_ils,
            cash_available=cash,
            portfolio_exposure_before=exposure,
            open_positions=open_positions,
            sector_map=sector_map,
            run_context=run_context,
            recent_stop_events=recent_stop_events,
            neutral_pilot_trades_today=neutral_pilot_buys_today,
        )
        decision_json["scan_source"] = scan_source_text(settings)
        final_action = str(decision_json.get("final_action") or decision.action)
        final_reason = str(decision_json.get("reason") or decision.feedback)
        if final_action == "BUY_SIMULATED":
            decision.quantity = int(decision_json.get("position_size") or decision.quantity)
            decision.cash_out_ils = float(decision_json.get("adjusted_cash_out") or decision.cash_out_ils)
            decision.risk_ils = float(decision_json.get("adjusted_risk_amount") or decision.risk_ils)
        if final_action != decision.action:
            decision = Decision(final_action, final_reason, decision_json=decision_json)
        else:
            decision.feedback = final_reason
            decision.decision_json = decision_json
        enrich_decision_analytics(decision_json, run_id=run_id, result=result, final_action=final_action)
        decision_json["active_strategy"] = "CURRENT_AGENT_GATES"
        decision_json["shadow_strategies"] = evaluate_shadow_strategies(result, decision_json)
        result.selection_context = build_selection_context(
            result,
            decision,
            universe=settings.universe,
            explicit_tickers=bool(settings.tickers),
            sector_map=sector_map,
            sector_health=sector_health,
            decision_json=decision_json,
        )
        decision_records.append(decision_json)
        decisions[result.ticker] = decision
        pending_decisions.append((result, decision))
        if decision.action == "BUY_SIMULATED":
            open_positions[result.ticker] = position_from_buy(result, decision, currency_rate, timestamp, screenshot_path)
            new_buy_tickers.add(result.ticker)
            if decision_json.get("entry_mode") == "neutral_pilot":
                neutral_pilot_buys_today += 1
            cash -= decision.cash_out_ils
            exposure += decision.cash_out_ils
        elif decision.action in {"TAKE_PARTIAL_PROFIT", "TAKE_PROFIT", "EXIT_STOP"}:
            apply_exit_decision(open_positions, result, decision, currency_rate)
        elif result.ticker in open_positions:
            refresh_open_position(open_positions[result.ticker], result, currency_rate, decision)

    apply_chart_retention_policy(
        pending_decisions,
        base_url=settings.url,
        run_id=run_id,
        analysis_period=settings.analysis_period,
        min_rr=settings.min_rr,
        open_position_tickers=initial_open_position_tickers | set(open_positions),
    )
    for result, decision in pending_decisions:
        if result.ticker in open_positions and (result.chart_url or result.ticker in new_buy_tickers):
            open_positions[result.ticker]["chart_url"] = result.chart_url
        append_watchlist_row(wb, timestamp, result, decision)
        if decision.action == "BUY_SIMULATED":
            append_trade_log_row(wb, timestamp, result, decision, currency_rate, screenshot_path)
        elif decision.action in {"TAKE_PARTIAL_PROFIT", "TAKE_PROFIT", "EXIT_STOP"}:
            append_trade_log_row(wb, timestamp, result, decision, currency_rate, screenshot_path)

    write_open_positions(wb, open_positions)
    write_decision_jsonl(decision_path, decision_records)
    cash = compute_cash(wb, starting_capital)
    exposure = sum(pos["exposure_ils"] for pos in open_positions.values())
    open_risk = sum(pos["risk_ils"] for pos in open_positions.values())
    performance_summary_paths: dict[str, Path] = {}
    try:
        performance_summary_paths = write_performance_summaries(
            summary_dir=SUMMARY_DIR,
            decision_dir=DECISION_DIR,
            current_decision_path=decision_path,
            run_id=run_id,
            timestamp=timestamp,
            portfolio={
                "currency": currency,
                "starting_capital": starting_capital,
                "open_positions_start": len(initial_open_position_tickers),
                "open_positions_end": len(open_positions),
                "cash": cash,
                "exposure": exposure,
                "open_risk": open_risk,
                "total_portfolio_value": round(cash + exposure, 2),
                "daily_return_pct": round(((cash + exposure) - starting_capital) / starting_capital * 100, 4)
                if starting_capital
                else None,
            },
        )
    except Exception as exc:
        errors.append(f"PERFORMANCE_SUMMARY_FAILED: {exc}")
    append_update_log(
        wb,
        timestamp=timestamp,
        run_id=run_id,
        tickers=[result.ticker for result in results],
        valid_setups=sum(1 for result in results if result.setup_type != "No Trade"),
        decisions=decisions,
        cash=cash,
        exposure=exposure,
        open_risk=open_risk,
        open_positions=len(open_positions),
        summary_path=summary_path,
        screenshot_path=screenshot_path,
        decision_path=decision_path,
    )
    wb.save(settings.excel_path)
    send_new_buy_notifications(
        pending_decisions,
        open_positions=open_positions,
        settings=settings,
        run_id=run_id,
        timestamp=timestamp,
    )
    return {
        "decisions": decisions,
        "cash": cash,
        "exposure": exposure,
        "remaining_budget": max(0.0, starting_capital - exposure),
        "open_risk": open_risk,
        "open_positions": open_positions,
        "workbook": settings.excel_path,
        "currency": currency,
        "decision_path": decision_path,
        "market_regime": run_context.market_regime,
        "performance_summary_paths": performance_summary_paths,
        "daily_summary_path": performance_summary_paths.get("daily_summary_json"),
        "weekly_summary_path": performance_summary_paths.get("weekly_summary_json"),
        "excel_updated": True,
        "run_status": "OK" if not errors else "ISSUES",
    }


def send_new_buy_notifications(
    pending_decisions: list[tuple[SetupResult, Decision]],
    *,
    open_positions: dict[str, dict[str, Any]],
    settings: Settings,
    run_id: str,
    timestamp: str,
) -> None:
    dashboard_url = dashboard_url_from_app_url(settings.url)
    for result, decision in pending_decisions:
        if decision.action != "BUY_SIMULATED":
            continue
        position = open_positions.get(result.ticker)
        if not position:
            continue
        message = format_position_opened_message(
            result=result,
            decision=decision,
            position=position,
            run_id=run_id,
            timestamp=timestamp,
            dashboard_url=dashboard_url,
        )
        outcome = send_telegram_message(message)
        if outcome.sent:
            log(f"Telegram position-open notification sent for {result.ticker}.")
        elif outcome.status != "not_configured":
            log(f"Telegram position-open notification skipped for {result.ticker}: {outcome.reason}")


def workbook_snapshot(*, settings: Settings, run_status: str) -> dict[str, Any]:
    try:
        wb = load_workbook(settings.excel_path, data_only=True)
        settings_values = read_settings(wb)
        currency = str(settings_values.get("budget_currency") or "USD").upper()
        starting_capital = float(
            settings_values.get("starting_capital_usd")
            or settings_values.get("starting_capital_ils")
            or 100_000
        )
        open_positions = read_open_positions(wb)
        cash = compute_cash(wb, starting_capital)
        exposure = sum(pos["exposure_ils"] for pos in open_positions.values())
        open_risk = sum(pos["risk_ils"] for pos in open_positions.values())
        wb.close()
    except Exception:
        currency = "USD"
        starting_capital = 100_000.0
        cash = starting_capital
        exposure = 0.0
        open_risk = 0.0
        open_positions = {}
    return {
        "decisions": {},
        "cash": cash,
        "exposure": exposure,
        "remaining_budget": max(0.0, starting_capital - exposure),
        "open_risk": open_risk,
        "open_positions": open_positions,
        "workbook": settings.excel_path,
        "currency": currency,
        "decision_path": "",
        "market_regime": None,
        "excel_updated": False,
        "run_status": run_status,
    }


def decide(
    result: SetupResult,
    *,
    open_positions: dict[str, dict[str, Any]],
    cash: float,
    exposure: float,
    usd_ils: float,
    max_position: float,
    max_total_exposure: float,
    max_risk: float,
    min_rr: float,
    sector_map: dict[str, str],
    sector_health: dict[str, dict[str, Any]],
) -> Decision:
    return decide_strategy_candidate(
        normalize_strategy_candidate(result),
        open_positions=open_positions,
        cash=cash,
        exposure=exposure,
        currency_rate=usd_ils,
        max_position=max_position,
        max_total_exposure=max_total_exposure,
        max_risk=max_risk,
        min_rr=min_rr,
        sector_map=sector_map,
        sector_health=sector_health,
    )


def build_selection_context(
    result: SetupResult,
    decision: Decision,
    *,
    universe: str,
    explicit_tickers: bool,
    sector_map: dict[str, str],
    sector_health: dict[str, dict[str, Any]],
    decision_json: dict[str, Any] | None = None,
) -> str:
    sector = sector_map.get(result.ticker, "Unknown")
    health = sector_health.get(sector, {})
    if explicit_tickers:
        source = "Configured manual ticker basket"
    elif universe == "smart-universe":
        source = "Smart Universe: broad liquid US universe, diversified by sector"
    else:
        source = f"Configured universe: {universe}"

    sector_text = f"Sector: {sector}"
    if health:
        sector_text += (
            f" - {health.get('label', 'Unknown')} "
            f"({float(health.get('score', 0)):.0f}/100)"
        )
        if health.get("reason"):
            sector_text += f"; {health['reason']}"

    setup_text = (
        f"Setup: {result.setup_type}; score {result.score:.2f}; "
        f"R/R {result.risk_reward:.2f}x; price {result.current_price:.2f}"
    )
    action_text = f"Agent action: {decision.action} - {decision.feedback}"
    parts = [source, sector_text, setup_text]
    if decision_json:
        risk_bits = [
            f"Market: {decision_json.get('market_regime', 'Unknown')}",
            f"Sector regime: {decision_json.get('sector_regime', 'Unknown')}",
        ]
        if decision_json.get("net_rr") is not None:
            risk_bits.append(f"Net R/R: {float(decision_json.get('net_rr') or 0):.2f}")
        if decision_json.get("factor_tags"):
            risk_bits.append(f"Factors: {', '.join(decision_json.get('factor_tags') or [])}")
        parts.append("; ".join(risk_bits))
    parts.append(action_text)
    return " | ".join(parts)


def scan_source_text(settings: Settings) -> str:
    if settings.tickers:
        return "Manual ticker basket configured for this agent run."
    if settings.universe == "smart-universe":
        return "Smart Universe: broad liquid US universe filtered and diversified by the Market Lens UI."
    return f"Configured UI universe: {settings.universe}."


def read_settings(wb: Any) -> dict[str, Any]:
    ws = wb["Settings"]
    values = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            values[str(row[0])] = row[1]
    return values


def ensure_agent_columns(wb: Any) -> None:
    headers = {
        "Setup Watchlist": {
            16: "Chart URL",
            17: "Selection Context",
            18: "Decision JSON",
        },
        "Trade Log": {
            18: "Chart URL",
            19: "Selection Context",
            20: "Decision JSON",
            21: "Trade ID",
            22: "Setup Score Bucket",
            23: "Entry Confirmation",
            24: "MFE",
            25: "MAE",
            26: "R Multiple",
            27: "Duration",
            28: "Exit Reason",
            29: "Outcome After 1D",
            30: "Outcome After 3D",
            31: "Outcome After 5D",
            32: "Outcome After 10D",
        },
        "Open Positions": {
            16: "Chart URL",
            17: "Selection Context",
            18: "Decision JSON",
        },
        "Update Log": {
            12: "Decision JSONL",
        },
    }
    for sheet_name, sheet_headers in headers.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for col_idx, header in sheet_headers.items():
            ws.cell(1, col_idx, header)


def read_open_positions(wb: Any) -> dict[str, dict[str, Any]]:
    ws = wb["Open Positions"]
    positions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        positions[str(row[0])] = {
            "ticker": row[0],
            "entry_date": row[1],
            "entry_price": float(row[2] or 0),
            "current_price": float(row[3] or row[2] or 0),
            "quantity": int(row[4] or 0),
            "stop_loss": float(row[5] or 0),
            "target_1": float(row[6] or 0),
            "target_2": float(row[7] or 0),
            "status": row[8],
            "unrealized_usd": float(row[9] or 0),
            "unrealized_ils": float(row[10] or 0),
            "exposure_ils": float(row[11] or 0),
            "risk_ils": float(row[12] or 0),
            "notes": str(row[13] or ""),
            "screenshot": str(row[14] or ""),
            "chart_url": str(row[15] or ""),
            "selection_context": str(row[16] or ""),
            "decision_json": str(row[17] or ""),
            "partial_taken": "partial" in str(row[13] or "").lower(),
        }
    return positions


def compute_cash(wb: Any, starting_capital: float) -> float:
    ws = wb["Trade Log"]
    cash_out = 0.0
    cash_in = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        cash_out += float(row[9] or 0)
        cash_in += float(row[10] or 0)
    return round(starting_capital - cash_out + cash_in, 2)


def read_open_position_tickers(excel_path: Path) -> list[str]:
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb["Open Positions"]
    except Exception:
        return []
    try:
        tickers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                tickers.append(str(row[0]).upper())
        return tickers
    finally:
        wb.close()


def read_recent_watch_tickers(excel_path: Path, days: int | None = None) -> list[str]:
    lookback_days = days or int(os.getenv("MARKET_LENS_WATCH_CARRY_FORWARD_DAYS", "14"))
    cutoff = datetime.now() - timedelta(days=lookback_days)
    return read_recent_action_tickers(excel_path, {"WATCH", "WATCH_READY"}, cutoff)


def read_recent_skip_tickers(excel_path: Path, hours: int | None = None) -> list[str]:
    cooldown_hours = hours or int(os.getenv("MARKET_LENS_SKIP_COOLDOWN_HOURS", "8"))
    cutoff = datetime.now() - timedelta(hours=cooldown_hours)
    return read_recent_action_tickers(excel_path, "SKIP", cutoff)


def read_recent_stop_events(wb: Any, cooldown_days: int) -> dict[str, dict[str, Any]]:
    if cooldown_days <= 0 or "Trade Log" not in wb.sheetnames:
        return {}
    ws = wb["Trade Log"]
    now = datetime.now()
    events: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        timestamp = parse_agent_timestamp(row[0] if row else None)
        action = str(row[1] or "").upper().strip() if len(row) > 1 else ""
        ticker = str(row[2] or "").upper().strip() if len(row) > 2 else ""
        if action != "EXIT_STOP" or not ticker or timestamp is None:
            continue
        elapsed = business_days_between(timestamp, now)
        if elapsed > cooldown_days:
            continue
        decision_json = parse_json_cell(row[19] if len(row) > 19 else "")
        current = events.get(ticker)
        if current and timestamp <= current["timestamp"]:
            continue
        events[ticker] = {
            "timestamp": timestamp,
            "last_stop_date": timestamp.isoformat(timespec="seconds"),
            "setup_type": decision_json.get("setup_type", ""),
            "days_remaining": max(0, cooldown_days - elapsed),
        }
    return events


def count_neutral_pilot_buys_today(wb: Any) -> int:
    if "Trade Log" not in wb.sheetnames:
        return 0
    ws = wb["Trade Log"]
    today = datetime.now().date()
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        timestamp = parse_agent_timestamp(row[0] if row else None)
        action = str(row[1] or "").upper().strip() if len(row) > 1 else ""
        if timestamp is None or timestamp.date() != today or action != "BUY_SIMULATED":
            continue
        decision_json = parse_json_cell(row[19] if len(row) > 19 else "")
        if decision_json.get("entry_mode") == "neutral_pilot":
            count += 1
    return count


def business_days_between(start: datetime, end: datetime) -> int:
    if start > end:
        return 0
    day = start.date()
    end_day = end.date()
    count = 0
    while day < end_day:
        day = day + timedelta(days=1)
        if day.weekday() < 5:
            count += 1
    return count


def parse_json_cell(value: Any) -> dict[str, Any]:
    if not value:
        return {}
    try:
        parsed = json.loads(str(value))
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def read_recent_action_tickers(excel_path: Path, action_name: str | set[str], cutoff: datetime) -> list[str]:
    action_names = {action_name} if isinstance(action_name, str) else set(action_name)
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb["Setup Watchlist"]
    except Exception:
        return []
    try:
        latest: dict[str, tuple[datetime, str]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ticker = str(row[1] or "").upper().strip() if len(row) > 1 else ""
            action = str(row[12] or "").upper().strip() if len(row) > 12 else ""
            timestamp = parse_agent_timestamp(row[0] if row else None)
            if not ticker or not action or timestamp is None or timestamp < cutoff:
                continue
            current = latest.get(ticker)
            if current is None or timestamp > current[0]:
                latest[ticker] = (timestamp, action)
        matching = {
            ticker: event_time
            for ticker, (event_time, action) in latest.items()
            if action in action_names
        }
        return sorted(matching, key=lambda ticker: matching[ticker], reverse=True)
    finally:
        wb.close()


def parse_agent_timestamp(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if not value:
        return None
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text).replace(tzinfo=None)
    except ValueError:
        return None


def append_watchlist_row(wb: Any, timestamp: str, result: SetupResult, decision: Decision) -> None:
    ws = wb["Setup Watchlist"]
    row = next_row(ws)
    ws.cell(row, 1, timestamp)
    ws.cell(row, 2, result.ticker)
    ws.cell(row, 3, result.setup_type)
    ws.cell(row, 4, result.score)
    ws.cell(row, 5, result.current_price)
    ws.cell(row, 6, result.buy_zone_low)
    ws.cell(row, 7, result.buy_zone_high)
    ws.cell(row, 8, result.stop_loss)
    ws.cell(row, 9, result.target_1)
    ws.cell(row, 10, result.target_2)
    ws.cell(row, 11, result.risk_reward)
    ws.cell(row, 12, result.reason)
    ws.cell(row, 13, decision.action)
    ws.cell(row, 14, decision.feedback)
    ws.cell(row, 15, result.raw_text[:1000])
    ws.cell(row, 16, result.chart_url)
    ws.cell(row, 17, result.selection_context)
    ws.cell(row, 18, decision_json_text(decision))


def append_trade_log_row(
    wb: Any,
    timestamp: str,
    result: SetupResult,
    decision: Decision,
    usd_ils: float,
    screenshot_path: Path,
) -> None:
    ws = wb["Trade Log"]
    row = next_row(ws)
    ws.cell(row, 1, timestamp)
    ws.cell(row, 2, decision.action)
    ws.cell(row, 3, result.ticker)
    ws.cell(row, 4, result.current_price if decision.action == "BUY_SIMULATED" else None)
    ws.cell(row, 5, (decision.execution_price or result.current_price) if decision.action != "BUY_SIMULATED" else None)
    ws.cell(row, 6, decision.quantity)
    ws.cell(row, 7, usd_ils)
    ws.cell(row, 8, decision.quantity * result.current_price * usd_ils if decision.action == "BUY_SIMULATED" else 0)
    ws.cell(row, 9, decision.cash_in_ils)
    ws.cell(row, 10, decision.cash_out_ils)
    ws.cell(row, 11, decision.cash_in_ils)
    ws.cell(row, 12, result.stop_loss)
    ws.cell(row, 13, result.target_1)
    ws.cell(row, 14, result.target_2)
    ws.cell(row, 15, decision.risk_ils)
    ws.cell(row, 16, decision.feedback)
    ws.cell(row, 17, str(screenshot_path))
    ws.cell(row, 18, result.chart_url)
    ws.cell(row, 19, result.selection_context)
    ws.cell(row, 20, decision_json_text(decision))
    write_trade_analytics_columns(ws, row, decision.decision_json)


def write_trade_analytics_columns(ws: Any, row: int, decision_json: dict[str, Any]) -> None:
    values = [
        decision_json.get("trade_id", ""),
        decision_json.get("setup_score_bucket", ""),
        decision_json.get("entry_confirmation_status") or decision_json.get("confirmation_status", ""),
        decision_json.get("mfe"),
        decision_json.get("mae"),
        decision_json.get("r_multiple"),
        decision_json.get("duration"),
        decision_json.get("exit_reason"),
        decision_json.get("outcome_after_1d"),
        decision_json.get("outcome_after_3d"),
        decision_json.get("outcome_after_5d"),
        decision_json.get("outcome_after_10d"),
    ]
    for offset, value in enumerate(values, start=21):
        ws.cell(row, offset, value)


def position_from_buy(
    result: SetupResult,
    decision: Decision,
    usd_ils: float,
    timestamp: str,
    screenshot_path: Path,
) -> dict[str, Any]:
    exposure = decision.quantity * result.current_price * usd_ils
    risk = decision.quantity * (result.current_price - float(result.stop_loss or 0)) * usd_ils
    return {
        "ticker": result.ticker,
        "entry_date": timestamp,
        "entry_price": result.current_price,
        "current_price": result.current_price,
        "quantity": decision.quantity,
        "stop_loss": result.stop_loss,
        "target_1": result.target_1,
        "target_2": result.target_2,
        "status": "OPEN",
        "unrealized_usd": 0,
        "unrealized_ils": 0,
        "exposure_ils": round(exposure, 2),
        "risk_ils": round(risk, 2),
        "notes": decision.feedback,
        "screenshot": str(screenshot_path),
        "chart_url": result.chart_url,
        "selection_context": result.selection_context,
        "decision_json": decision_json_text(decision),
    }


def apply_exit_decision(
    open_positions: dict[str, dict[str, Any]],
    result: SetupResult,
    decision: Decision,
    usd_ils: float,
) -> None:
    pos = open_positions.get(result.ticker)
    if not pos:
        return
    if decision.action == "TAKE_PARTIAL_PROFIT":
        closed_qty = min(int(pos["quantity"]), max(1, decision.quantity))
        remaining_qty = int(pos["quantity"]) - closed_qty
        if remaining_qty <= 0:
            open_positions.pop(result.ticker, None)
            return
        pos["quantity"] = remaining_qty
        pos["stop_loss"] = pos["entry_price"]
        pos["notes"] = "Partial profit taken; stop moved to breakeven."
        refresh_open_position(pos, result, usd_ils, decision)
        return
    open_positions.pop(result.ticker, None)


def refresh_open_position(
    pos: dict[str, Any],
    result: SetupResult,
    usd_ils: float,
    decision: Decision | None = None,
) -> None:
    qty = int(pos.get("quantity") or 0)
    entry = float(pos.get("entry_price") or 0)
    current = result.current_price
    stop = float(pos.get("stop_loss") or 0)
    pos["current_price"] = current
    pos["unrealized_usd"] = round((current - entry) * qty, 2)
    pos["unrealized_ils"] = round((current - entry) * qty * usd_ils, 2)
    pos["exposure_ils"] = round(current * qty * usd_ils, 2)
    pos["risk_ils"] = round(max(0.0, current - stop) * qty * usd_ils, 2)
    if result.chart_url:
        pos["chart_url"] = result.chart_url
    if result.selection_context:
        pos["selection_context"] = result.selection_context
    if decision and decision.decision_json:
        # Preserve the latest structured explanation for dashboard drill-down.
        pos["decision_json"] = decision_json_text(decision)


def write_open_positions(wb: Any, positions: dict[str, dict[str, Any]]) -> None:
    ws = wb["Open Positions"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row_idx, pos in enumerate(positions.values(), start=2):
        values = [
            pos["ticker"], pos["entry_date"], pos["entry_price"], pos["current_price"], pos["quantity"],
            pos["stop_loss"], pos["target_1"], pos["target_2"], pos["status"], pos["unrealized_usd"],
            pos["unrealized_ils"], pos["exposure_ils"], pos["risk_ils"], pos.get("notes", ""), pos.get("screenshot", ""),
            pos.get("chart_url", ""), pos.get("selection_context", ""), pos.get("decision_json", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)


def decision_json_text(decision: Decision) -> str:
    if not decision.decision_json:
        return ""
    return json.dumps(decision.decision_json, ensure_ascii=False, sort_keys=True, default=str)


def enrich_decision_analytics(
    decision_json: dict[str, Any],
    *,
    run_id: str,
    result: SetupResult,
    final_action: str,
) -> None:
    setup_score = float(decision_json.get("setup_score") or result.score or 0)
    if final_action == "BUY_SIMULATED":
        decision_json.setdefault("trade_id", f"{run_id}-{result.ticker}")
    else:
        decision_json.setdefault("trade_id", "")
    decision_json["setup_score_bucket"] = setup_score_bucket(setup_score)
    decision_json["weighted_net_rr"] = decision_json.get("net_rr")
    decision_json["entry_confirmation_status"] = decision_json.get("confirmation_status")
    decision_json.setdefault("mfe", None)
    decision_json.setdefault("mae", None)
    decision_json.setdefault("r_multiple", None)
    decision_json.setdefault("duration", None)
    decision_json.setdefault("exit_reason", None)
    decision_json.setdefault("outcome_after_1d", None)
    decision_json.setdefault("outcome_after_3d", None)
    decision_json.setdefault("outcome_after_5d", None)
    decision_json.setdefault("outcome_after_10d", None)


def setup_score_bucket(score: float) -> str:
    if score < 0.40:
        return "<0.40"
    if score < 0.50:
        return "0.40-0.49"
    if score < 0.60:
        return "0.50-0.59"
    if score < 0.70:
        return "0.60-0.69"
    return "0.70+"


def write_decision_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = "\n".join(json.dumps(record, ensure_ascii=False, sort_keys=True, default=str) for record in records)
    path.write_text(payload + ("\n" if payload else ""), encoding="utf-8")


def append_update_log(
    wb: Any,
    *,
    timestamp: str,
    run_id: str,
    tickers: list[str],
    valid_setups: int,
    decisions: dict[str, Decision],
    cash: float,
    exposure: float,
    open_risk: float,
    open_positions: int,
    summary_path: Path,
    screenshot_path: Path,
    decision_path: Path,
) -> None:
    ws = wb["Update Log"]
    row = next_row(ws)
    actions = ", ".join(f"{ticker}:{decision.action}" for ticker, decision in decisions.items())
    ws.cell(row, 1, timestamp)
    ws.cell(row, 2, run_id)
    ws.cell(row, 3, " ".join(tickers))
    ws.cell(row, 4, valid_setups)
    ws.cell(row, 5, actions)
    ws.cell(row, 6, cash)
    ws.cell(row, 7, exposure)
    ws.cell(row, 8, open_risk)
    ws.cell(row, 9, open_positions)
    ws.cell(row, 10, str(summary_path))
    ws.cell(row, 11, str(screenshot_path))
    ws.cell(row, 12, str(decision_path))


def write_summary(
    *,
    settings: Settings,
    run_id: str,
    login_status: str,
    scan_status: str,
    results: list[SetupResult],
    decisions: dict[str, Decision],
    screenshot_path: Path,
    summary_path: Path,
    workbook_context: dict[str, Any],
    errors: list[str],
) -> None:
    valid = [result for result in results if result.setup_type != "No Trade"]
    buys = [ticker for ticker, decision in decisions.items() if decision.action == "BUY_SIMULATED"]
    watch_ready = [ticker for ticker, decision in decisions.items() if decision.action == "WATCH_READY"]
    watch = [ticker for ticker, decision in decisions.items() if decision.action == "WATCH"]
    closed = [ticker for ticker, decision in decisions.items() if decision.action in {"TAKE_PROFIT", "EXIT_STOP", "TAKE_PARTIAL_PROFIT"}]
    market_regime = workbook_context.get("market_regime")
    market_text = (
        f"{market_regime.label} ({market_regime.score:.2f}) - {market_regime.reason}"
        if market_regime
        else "Not available"
    )
    run_status = str(workbook_context.get("run_status") or ("OK" if not errors else "ISSUES"))
    excel_line = (
        f"Excel updated: {settings.excel_path}"
        if workbook_context.get("excel_updated", True)
        else "Excel updated: skipped"
    )
    decision_line = (
        f"Decision JSONL saved: {workbook_context.get('decision_path', '')}"
        if workbook_context.get("decision_path")
        else "Decision JSONL saved: skipped"
    )
    lines = [
        "Market Lens Agent Update",
        "",
        f"Date: {datetime.now().isoformat(timespec='seconds')}",
        f"Run status: {run_status}",
        f"Login status: {login_status}",
        f"Scan status: {scan_status}",
        f"Tickers scanned: {' '.join(result.ticker for result in results)}",
        f"Valid setups found: {len(valid)}",
        f"Market regime: {market_text}",
        f"Actions taken: {', '.join(f'{ticker}:{decision.action}' for ticker, decision in decisions.items())}",
        f"New simulated buys: {', '.join(buys) or 'None'}",
        f"Watch ready setups: {', '.join(watch_ready) or 'None'}",
        f"Positions on watch: {', '.join(watch) or 'None'}",
        f"Positions closed: {', '.join(closed) or 'None'}",
        f"Cash remaining: {workbook_context['cash']:.2f} {workbook_context['currency']}",
        f"Current exposure: {workbook_context['exposure']:.2f} {workbook_context['currency']}",
        f"Remaining available budget: {workbook_context['remaining_budget']:.2f} {workbook_context['currency']}",
        f"Total open risk: {workbook_context['open_risk']:.2f} {workbook_context['currency']}",
        excel_line,
        f"Screenshot saved: {screenshot_path}",
        decision_line,
        f"Daily summary saved: {workbook_context.get('daily_summary_path') or 'skipped'}",
        f"Weekly summary saved: {workbook_context.get('weekly_summary_path') or 'skipped'}",
        f"Errors: {'; '.join(errors) if errors else 'None'}",
        "Agent feedback:",
    ]
    for result in results:
        decision = decisions.get(result.ticker)
        if decision:
            lines.append(f"- {result.ticker}: {decision.action} - {decision.feedback}")
            if decision.decision_json:
                warnings = decision.decision_json.get("warnings") or []
                if warnings:
                    lines.append(f"  Warnings: {'; '.join(str(item) for item in warnings[:4])}")
            if result.selection_context:
                lines.append(f"  Context: {result.selection_context}")
    summary_path.write_text("\n".join(lines), encoding="utf-8")


def next_row(ws: Any) -> int:
    for row in range(2, ws.max_row + 2):
        if ws.cell(row, 1).value is None:
            return row
    return ws.max_row + 1


def parse_tickers(value: str) -> list[str]:
    return [item.upper() for item in re.split(r"[\s,]+", value.strip()) if item]


def unique_tickers(tickers: list[str]) -> list[str]:
    seen = set()
    result = []
    for ticker in tickers:
        normalized = ticker.upper()
        if normalized and normalized not in seen:
            seen.add(normalized)
            result.append(normalized)
    return result


def parse_float(value: Any) -> float:
    parsed = parse_optional_float(value)
    return parsed if parsed is not None else 0.0


def parse_optional_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).replace(",", "").strip()
    if text in {"", "-"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def parse_range(value: Any) -> tuple[float | None, float | None]:
    text = str(value or "").replace(",", "")
    nums = [float(item) for item in re.findall(r"-?\d+(?:\.\d+)?", text)]
    if len(nums) >= 2:
        return nums[0], nums[1]
    if len(nums) == 1:
        return nums[0], None
    return None, None


if __name__ == "__main__":
    main()
