from __future__ import annotations

import os
from pathlib import Path
from typing import Any


DEFAULT_WATCHLIST_MAX_ROWS = 80_000
DEFAULT_TRACKER_REWRITE_BYTES = 75_000_000
DEFAULT_TRACKER_HARD_LIMIT_BYTES = 90_000_000


def compact_setup_watchlist(
    workbook: Any,
    *,
    max_rows: int | None = None,
) -> dict[str, int | bool]:
    """Keep the operational Excel tracker below GitHub's file-size ceiling.

    Full scan history remains available in decision JSONL and performance summaries.
    The workbook only needs a rolling window large enough for carry-forward selection,
    dashboard diagnostics, and recent manual review.
    """
    configured_max = max_rows
    if configured_max is None:
        configured_max = _env_int(
            "MARKET_LENS_WORKBOOK_WATCHLIST_MAX_ROWS",
            DEFAULT_WATCHLIST_MAX_ROWS,
        )

    if configured_max <= 0 or "Setup Watchlist" not in workbook.sheetnames:
        return {
            "enabled": False,
            "rows_before": 0,
            "rows_after": 0,
            "rows_removed": 0,
            "max_rows": configured_max,
        }

    worksheet = workbook["Setup Watchlist"]
    rows_before = max(0, worksheet.max_row - 1)
    rows_to_remove = max(0, rows_before - configured_max)
    if rows_to_remove:
        worksheet.delete_rows(2, rows_to_remove)

    return {
        "enabled": True,
        "rows_before": rows_before,
        "rows_after": rows_before - rows_to_remove,
        "rows_removed": rows_to_remove,
        "max_rows": configured_max,
    }


def enforce_tracker_size(
    tracker_path: Path,
    *,
    rewrite_bytes: int | None = None,
    hard_limit_bytes: int | None = None,
    max_rows: int | None = None,
) -> dict[str, Any]:
    """Rewrite an oversized tracker and fail before GitHub rejects the push."""
    rewrite_at = rewrite_bytes or _env_int(
        "MARKET_LENS_WORKBOOK_REWRITE_BYTES",
        DEFAULT_TRACKER_REWRITE_BYTES,
    )
    hard_limit = hard_limit_bytes or _env_int(
        "MARKET_LENS_WORKBOOK_HARD_LIMIT_BYTES",
        DEFAULT_TRACKER_HARD_LIMIT_BYTES,
    )
    size_before = tracker_path.stat().st_size
    result: dict[str, Any] = {
        "path": str(tracker_path),
        "size_before": size_before,
        "size_after": size_before,
        "rewritten": False,
        "hard_limit": hard_limit,
    }
    if size_before < rewrite_at:
        return result

    from openpyxl import load_workbook

    workbook = load_workbook(tracker_path)
    try:
        result["retention"] = compact_setup_watchlist(workbook, max_rows=max_rows)
        temporary_path = tracker_path.with_suffix(".compacting.xlsx")
        workbook.save(temporary_path)
    finally:
        workbook.close()

    os.replace(temporary_path, tracker_path)
    size_after = tracker_path.stat().st_size
    result.update({"size_after": size_after, "rewritten": True})
    if size_after >= hard_limit:
        raise RuntimeError(
            f"Tracker remains too large after compaction: {size_after} bytes "
            f"(hard limit {hard_limit})."
        )
    return result


def _env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)))
    except (TypeError, ValueError):
        return default
