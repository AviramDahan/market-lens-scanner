from __future__ import annotations

import json
import math
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from statistics import median
from typing import Any, Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class CapitalScenario:
    name: str
    description: str
    preserve_original_size: bool = False
    max_position_pct: float = 0.10
    base_risk_pct: float = 0.005
    high_quality_risk_pct: float = 0.0075
    total_heat_pct: float = 0.025
    max_exposure_pct: float = 0.60
    neutral_exposure_pct: float = 0.45
    bull_exposure_pct: float = 0.60
    sector_risk_pct: float = 0.0075
    factor_risk_pct: float = 0.01


DEFAULT_SCENARIOS = (
    CapitalScenario(
        name="ACTUAL_BASELINE",
        description="Original quantities and original trade lifecycle; no sizing changes.",
        preserve_original_size=True,
    ),
    CapitalScenario(
        name="CONSERVATIVE_DYNAMIC",
        description="30% neutral / 50% bull exposure, 0.40%-0.60% trade risk, 2% heat.",
        base_risk_pct=0.004,
        high_quality_risk_pct=0.006,
        total_heat_pct=0.02,
        max_exposure_pct=0.50,
        neutral_exposure_pct=0.30,
        bull_exposure_pct=0.50,
        sector_risk_pct=0.006,
        factor_risk_pct=0.008,
    ),
    CapitalScenario(
        name="BALANCED_DYNAMIC",
        description="Continuous regime exposure up to 60%, 0.50%-0.75% risk, 2.5% heat.",
    ),
    CapitalScenario(
        name="GROWTH_DIAGNOSTIC",
        description="Diagnostic only: up to 70% exposure, 0.65%-0.90% risk, 3% heat.",
        base_risk_pct=0.0065,
        high_quality_risk_pct=0.009,
        total_heat_pct=0.03,
        max_exposure_pct=0.70,
        neutral_exposure_pct=0.55,
        bull_exposure_pct=0.70,
        sector_risk_pct=0.01,
        factor_risk_pct=0.0125,
    ),
)


def parse_json(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if not value:
        return {}
    try:
        parsed = json.loads(str(value))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def parse_timestamp(value: Any) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value or "").strip().replace("Z", "+00:00")
        parsed = datetime.fromisoformat(text)
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def to_float(value: Any, default: float = 0.0) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return default
    return result if math.isfinite(result) else default


def load_closed_trades(workbook_path: Path) -> list[dict[str, Any]]:
    """Read completed paper trades without modifying the workbook."""
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if "Trade Log" not in workbook.sheetnames:
        return []
    sheet = workbook["Trade Log"]
    open_trades: dict[str, dict[str, Any]] = {}
    closed: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        action = str(row[1] or "").upper()
        ticker = str(row[2] or "").upper().strip()
        quantity = max(0, int(to_float(row[5])))
        if not ticker or quantity <= 0:
            continue
        if action == "BUY_SIMULATED":
            decision = parse_json(row[19])
            open_trades[ticker] = {
                "trade_id": str(row[20] or decision.get("trade_id") or f"{ticker}|{row[0]}"),
                "ticker": ticker,
                "entry_timestamp": parse_timestamp(row[0]),
                "entry_price": to_float(row[3]),
                "stop_loss": to_float(row[11]),
                "target_1": to_float(row[12]),
                "target_2": to_float(row[13]),
                "initial_quantity": quantity,
                "remaining_quantity": quantity,
                "decision": decision,
                "exits": [],
            }
            continue
        if action not in {"TAKE_PARTIAL_PROFIT", "TAKE_PROFIT", "EXIT_STOP"}:
            continue
        trade = open_trades.get(ticker)
        if not trade:
            continue
        used = min(quantity, int(trade["remaining_quantity"]))
        if used <= 0:
            continue
        trade["exits"].append(
            {
                "timestamp": parse_timestamp(row[0]),
                "action": action,
                "price": to_float(row[4]),
                "quantity": used,
                "currency_rate": to_float(row[6], 1.0) or 1.0,
            }
        )
        trade["remaining_quantity"] -= used
        if action in {"TAKE_PROFIT", "EXIT_STOP"} or trade["remaining_quantity"] <= 0:
            trade["exit_timestamp"] = trade["exits"][-1]["timestamp"]
            trade["exit_action"] = action
            closed.append(dict(trade))
            del open_trades[ticker]
    workbook.close()
    return sorted(closed, key=lambda trade: trade["entry_timestamp"])


def market_risk_points(decision: dict[str, Any]) -> float:
    indicators = decision.get("market_regime_indicators") or {}
    if not indicators:
        regime = str(decision.get("market_regime") or "NEUTRAL").upper()
        return {"BULL": 4.0, "NEUTRAL": 2.0, "BEAR": -2.0}.get(regime, 2.0)
    points = 0.0
    for ticker in ("SPY", "QQQ"):
        trend = str((indicators.get(ticker) or {}).get("trend") or "").lower()
        points += 2.0 if trend == "bullish" else (-2.0 if trend == "bearish" else 0.0)
    iwm = str((indicators.get("IWM") or {}).get("trend") or "").lower()
    points += -1.0 if iwm == "bearish" else 1.0
    vix = indicators.get("VIX") or {}
    vix_price = to_float(vix.get("price"))
    vix_trend = str(vix.get("trend") or "").lower()
    if vix_price >= 25 or vix_trend in {"stressed", "bullish"}:
        points -= 2.0
    elif vix_price and vix_price < 20:
        points += 1.0
    us10y = str((indicators.get("US10Y") or {}).get("trend") or "").lower()
    points += -0.5 if us10y == "bullish" else (0.25 if us10y == "bearish" else 0.0)
    dxy = str((indicators.get("DXY") or {}).get("trend") or "").lower()
    points += -0.25 if dxy == "bullish" else (0.25 if dxy == "bearish" else 0.0)
    return round(points, 2)


def exposure_limit_pct(scenario: CapitalScenario, decision: dict[str, Any]) -> float:
    if scenario.preserve_original_size:
        return 1.0
    points = market_risk_points(decision)
    if points <= -2:
        return 0.0
    if scenario.name != "BALANCED_DYNAMIC":
        regime = str(decision.get("market_regime") or "NEUTRAL").upper()
        if regime == "BEAR":
            return 0.0
        return scenario.bull_exposure_pct if regime == "BULL" else scenario.neutral_exposure_pct
    if points <= 0:
        return 0.10 + ((points + 2.0) / 2.0) * 0.10
    if points <= 2:
        return 0.20 + (points / 2.0) * 0.10
    if points <= 4:
        return 0.30 + ((points - 2.0) / 2.0) * 0.15
    return min(scenario.max_exposure_pct, 0.45 + ((points - 4.0) / 2.0) * 0.15)


def quality_multiplier(decision: dict[str, Any]) -> float:
    score = to_float(decision.get("setup_score"))
    confirmed = bool(decision.get("entry_confirmation_passed"))
    sector = str(decision.get("sector_regime") or "").upper()
    if score >= 0.55 and confirmed and sector == "STRONG":
        return 1.0
    if score >= 0.50 and confirmed:
        return 0.75
    if score >= 0.45 and confirmed:
        return 0.50
    return 0.35


def _scenario_quantity(
    scenario: CapitalScenario,
    trade: dict[str, Any],
    *,
    equity: float,
    cash: float,
    exposure: float,
    open_risk: float,
    sector_risk: float,
    factor_risks: dict[str, float],
) -> tuple[int, dict[str, Any]]:
    original_quantity = int(trade["initial_quantity"])
    if scenario.preserve_original_size:
        return original_quantity, {"binding_constraint": "original_quantity"}
    decision = trade["decision"]
    entry = to_float(trade["entry_price"])
    risk_per_share = max(0.0, entry - to_float(trade["stop_loss"]))
    if entry <= 0 or risk_per_share <= 0 or equity <= 0:
        return 0, {"binding_constraint": "invalid_entry_or_stop"}

    multiplier = quality_multiplier(decision)
    risk_pct = (
        scenario.high_quality_risk_pct
        if multiplier >= 1.0
        else scenario.base_risk_pct * multiplier
    )
    exposure_pct = exposure_limit_pct(scenario, decision)
    factors = [str(value) for value in decision.get("factor_tags") or [] if value]
    constraints = {
        "risk_budget": equity * risk_pct / risk_per_share,
        "position_cap": equity * scenario.max_position_pct / entry,
        "market_exposure": max(0.0, equity * exposure_pct - exposure) / entry,
        "cash": max(0.0, cash) / entry,
        "portfolio_heat": max(0.0, equity * scenario.total_heat_pct - open_risk) / risk_per_share,
        "sector_risk": max(0.0, equity * scenario.sector_risk_pct - sector_risk) / risk_per_share,
    }
    if factors:
        constraints["factor_risk"] = min(
            max(0.0, equity * scenario.factor_risk_pct - factor_risks.get(factor, 0.0))
            / risk_per_share
            for factor in factors
        )
    binding = min(constraints, key=constraints.get)
    quantity = max(0, int(math.floor(max(0.0, constraints[binding]))))
    return quantity, {
        "binding_constraint": binding,
        "quality_multiplier": multiplier,
        "risk_pct": round(risk_pct, 6),
        "market_risk_points": market_risk_points(decision),
        "exposure_limit_pct": round(exposure_pct, 4),
    }


def replay_scenario(
    trades: list[dict[str, Any]],
    scenario: CapitalScenario,
    *,
    starting_capital: float = 100_000.0,
    idle_cash_annual_yield: float = 0.05,
) -> dict[str, Any]:
    events: list[tuple[datetime, int, str, dict[str, Any], dict[str, Any] | None]] = []
    for trade in trades:
        events.append((trade["entry_timestamp"], 1, "ENTRY", trade, None))
        for exit_event in trade["exits"]:
            events.append((exit_event["timestamp"], 0, "EXIT", trade, exit_event))
    events.sort(key=lambda item: (item[0], item[1]))
    if not events:
        return {"scenario": scenario.name, "completed_trades": 0, "ending_equity": starting_capital}

    cash = starting_capital
    realized_pnl = 0.0
    hypothetical_cash_yield = 0.0
    exposure = 0.0
    open_risk = 0.0
    sector_risk: dict[str, float] = defaultdict(float)
    factor_risk: dict[str, float] = defaultdict(float)
    positions: dict[str, dict[str, Any]] = {}
    completed: list[dict[str, Any]] = []
    skipped = 0
    downsized = 0
    upsized = 0
    binding_constraints: Counter[str] = Counter()
    max_exposure_pct_seen = 0.0
    weighted_utilization_seconds = 0.0
    elapsed_seconds = 0.0
    equity_peak = starting_capital
    max_drawdown = 0.0
    previous_timestamp = events[0][0]

    for timestamp, _, event_type, trade, exit_event in events:
        delta_seconds = max(0.0, (timestamp - previous_timestamp).total_seconds())
        equity_before = starting_capital + realized_pnl + hypothetical_cash_yield
        if delta_seconds:
            hypothetical_cash_yield += cash * idle_cash_annual_yield * delta_seconds / (365.25 * 86400)
            utilization = exposure / equity_before if equity_before > 0 else 0.0
            weighted_utilization_seconds += utilization * delta_seconds
            elapsed_seconds += delta_seconds
        previous_timestamp = timestamp

        trade_id = str(trade["trade_id"])
        if event_type == "ENTRY":
            equity = starting_capital + realized_pnl
            decision = trade["decision"]
            sector = str(decision.get("sector") or "Unknown")
            factors = [str(value) for value in decision.get("factor_tags") or [] if value]
            quantity, sizing = _scenario_quantity(
                scenario,
                trade,
                equity=equity,
                cash=cash,
                exposure=exposure,
                open_risk=open_risk,
                sector_risk=sector_risk[sector],
                factor_risks=factor_risk,
            )
            binding_constraints[sizing["binding_constraint"]] += 1
            original_quantity = int(trade["initial_quantity"])
            if quantity <= 0:
                skipped += 1
                positions[trade_id] = {"skipped": True}
                continue
            downsized += quantity < original_quantity
            upsized += quantity > original_quantity
            entry = to_float(trade["entry_price"])
            per_share_risk = max(0.0, entry - to_float(trade["stop_loss"]))
            position_risk = per_share_risk * quantity
            cost = entry * quantity
            cash -= cost
            exposure += cost
            open_risk += position_risk
            sector_risk[sector] += position_risk
            for factor in factors:
                factor_risk[factor] += position_risk
            positions[trade_id] = {
                "skipped": False,
                "quantity": quantity,
                "remaining": quantity,
                "entry": entry,
                "cost": cost,
                "risk": position_risk,
                "remaining_risk": position_risk,
                "sector": sector,
                "factors": factors,
                "pnl": 0.0,
                "sizing": sizing,
                "original_quantity": original_quantity,
                "ticker": trade["ticker"],
            }
            equity_now = starting_capital + realized_pnl
            max_exposure_pct_seen = max(
                max_exposure_pct_seen, exposure / equity_now if equity_now > 0 else 0.0
            )
            continue

        position = positions.get(trade_id)
        if not position or position.get("skipped") or not exit_event:
            continue
        original_exit_quantity = int(exit_event["quantity"])
        original_total = int(trade["initial_quantity"])
        is_final = exit_event is trade["exits"][-1]
        if is_final:
            exit_quantity = int(position["remaining"])
        else:
            exit_quantity = min(
                int(position["remaining"]),
                max(1, int(round(position["quantity"] * original_exit_quantity / original_total))),
            )
        if exit_quantity <= 0:
            continue
        exit_price = to_float(exit_event["price"])
        rate = to_float(exit_event.get("currency_rate"), 1.0) or 1.0
        pnl = (exit_price - position["entry"]) * exit_quantity * rate
        released_cost = position["entry"] * exit_quantity
        released_risk = position["risk"] * (exit_quantity / position["quantity"])
        cash += released_cost + pnl
        exposure = max(0.0, exposure - released_cost)
        open_risk = max(0.0, open_risk - released_risk)
        sector_risk[position["sector"]] = max(
            0.0, sector_risk[position["sector"]] - released_risk
        )
        for factor in position["factors"]:
            factor_risk[factor] = max(0.0, factor_risk[factor] - released_risk)
        position["remaining"] -= exit_quantity
        position["pnl"] += pnl
        realized_pnl += pnl
        if position["remaining"] <= 0:
            completed.append(
                {
                    "trade_id": trade_id,
                    "ticker": position["ticker"],
                    "quantity": position["quantity"],
                    "original_quantity": position["original_quantity"],
                    "pnl": round(position["pnl"], 2),
                    "sizing": position["sizing"],
                }
            )
        equity_now = starting_capital + realized_pnl
        equity_peak = max(equity_peak, equity_now)
        max_drawdown = min(max_drawdown, equity_now - equity_peak)

    pnl_values = [trade["pnl"] for trade in completed]
    gross_profit = sum(value for value in pnl_values if value > 0)
    gross_loss = abs(sum(value for value in pnl_values if value < 0))
    ending_equity = starting_capital + realized_pnl
    return {
        "scenario": scenario.name,
        "description": scenario.description,
        "starting_capital": round(starting_capital, 2),
        "ending_equity": round(ending_equity, 2),
        "trading_pnl": round(realized_pnl, 2),
        "trading_return_pct": round(realized_pnl / starting_capital * 100, 4),
        "hypothetical_idle_cash_yield": round(hypothetical_cash_yield, 2),
        "ending_equity_with_idle_cash_yield": round(ending_equity + hypothetical_cash_yield, 2),
        "total_return_with_idle_cash_yield_pct": round(
            (realized_pnl + hypothetical_cash_yield) / starting_capital * 100, 4
        ),
        "completed_trades": len(completed),
        "skipped_original_trades": skipped,
        "downsized_trades": downsized,
        "upsized_trades": upsized,
        "wins": sum(value > 0 for value in pnl_values),
        "losses": sum(value < 0 for value in pnl_values),
        "win_rate_pct": round(sum(value > 0 for value in pnl_values) / len(pnl_values) * 100, 2)
        if pnl_values
        else None,
        "profit_factor": round(gross_profit / gross_loss, 4) if gross_loss else None,
        "average_trade_pnl": round(sum(pnl_values) / len(pnl_values), 2) if pnl_values else None,
        "max_realized_drawdown": round(max_drawdown, 2),
        "average_time_weighted_exposure_pct": round(
            weighted_utilization_seconds / elapsed_seconds * 100, 2
        )
        if elapsed_seconds
        else 0.0,
        "max_exposure_pct": round(max_exposure_pct_seen * 100, 2),
        "binding_constraints": dict(binding_constraints.most_common()),
        "trades": completed,
        "configuration": asdict(scenario),
    }


def _capital_blocked(record: dict[str, Any]) -> bool:
    if str(record.get("initial_action") or "").upper() != "BUY_SIMULATED":
        return False
    if str(record.get("final_action") or "").upper() == "BUY_SIMULATED":
        return False
    if not record.get("entry_confirmation_passed"):
        return False
    if record.get("earnings_blackout") or record.get("cooldown_active"):
        return False
    if record.get("correlation_warning"):
        return False
    if not record.get("market_session_can_open_new_buy", True):
        return False
    setup_score = to_float(record.get("setup_score"))
    minimum_score = to_float(record.get("minimum_setup_score_required"), 0.45)
    net_rr = to_float(record.get("net_rr"))
    minimum_net_rr = to_float(record.get("minimum_net_rr_required"), 2.0)
    primary_rr = to_float(record.get("net_rr_1"))
    pilot_eligible = bool(record.get("neutral_pilot_eligible"))
    if not pilot_eligible and (setup_score < minimum_score or net_rr < minimum_net_rr):
        return False
    if primary_rr and primary_rr < 0.80:
        return False
    target_status = str(record.get("target_feasibility_status") or "UNKNOWN").upper()
    if target_status not in {"OK", "UNKNOWN"}:
        return False
    reason = str(record.get("reason") or "").lower()
    return bool(
        record.get("sector_exposure_limit_exceeded")
        or record.get("factor_exposure_limit_exceeded")
        or "exposure limit" in reason
        or "available cash" in reason
        or "portfolio exposure" in reason
        or "position size" in reason
    )


def analyze_capital_blocked_candidates(decision_dir: Path) -> dict[str, Any]:
    """Measure capital-blocked signals using later recorded scan prices as a proxy."""
    total_records = 0
    initial_buy_signals = 0
    final_buy_signals = 0
    malformed = 0
    candidates: dict[tuple[str, str, str], dict[str, Any]] = {}
    daily_prices: dict[str, dict[str, tuple[datetime, float]]] = defaultdict(dict)
    for path in sorted(decision_dir.glob("*.jsonl")):
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                try:
                    record = json.loads(line)
                except (ValueError, json.JSONDecodeError):
                    malformed += 1
                    continue
                if not isinstance(record, dict):
                    continue
                total_records += 1
                initial_buy_signals += str(record.get("initial_action") or "").upper() == "BUY_SIMULATED"
                final_buy_signals += str(record.get("final_action") or "").upper() == "BUY_SIMULATED"
                ticker = str(record.get("ticker") or "").upper().strip()
                price = to_float(record.get("price"))
                try:
                    timestamp = parse_timestamp(record.get("timestamp"))
                except (TypeError, ValueError):
                    continue
                if ticker and price > 0:
                    day = timestamp.date().isoformat()
                    prior = daily_prices[ticker].get(day)
                    if prior is None or timestamp > prior[0]:
                        daily_prices[ticker][day] = (timestamp, price)
                if not ticker or not _capital_blocked(record):
                    continue
                key = (timestamp.date().isoformat(), ticker, str(record.get("setup_type") or ""))
                existing = candidates.get(key)
                if existing is None or to_float(record.get("setup_score")) > to_float(
                    existing.get("setup_score")
                ):
                    candidates[key] = {
                        "timestamp": timestamp.isoformat(),
                        "date": timestamp.date().isoformat(),
                        "ticker": ticker,
                        "setup_type": record.get("setup_type"),
                        "setup_score": to_float(record.get("setup_score")),
                        "entry_price": to_float(record.get("executable_entry")) or price,
                        "reason": str(record.get("reason") or ""),
                        "market_regime": record.get("market_regime"),
                        "sector": record.get("sector"),
                    }

    returns: dict[int, list[float]] = defaultdict(list)
    enriched: list[dict[str, Any]] = []
    reason_counts: Counter[str] = Counter()
    for candidate in candidates.values():
        prices = sorted(daily_prices.get(candidate["ticker"], {}).items())
        later = [(day, value[1]) for day, value in prices if day > candidate["date"]]
        entry = to_float(candidate["entry_price"])
        outcomes: dict[str, float | None] = {}
        for horizon in (1, 3, 5, 10):
            outcome = (
                (later[horizon - 1][1] / entry - 1.0) * 100
                if entry > 0 and len(later) >= horizon
                else None
            )
            outcomes[f"return_after_{horizon}_scan_days_pct"] = (
                round(outcome, 4) if outcome is not None else None
            )
            if outcome is not None:
                returns[horizon].append(outcome)
        reason = candidate["reason"].lower()
        category = (
            "sector_exposure"
            if "sector exposure" in reason
            else "factor_exposure"
            if "factor" in reason
            else "portfolio_exposure"
            if "exposure" in reason
            else "cash_or_position_size"
        )
        reason_counts[category] += 1
        enriched.append({**candidate, **outcomes, "blocker_category": category})

    def outcome_summary(values: list[float]) -> dict[str, Any]:
        return {
            "sample_size": len(values),
            "average_pct": round(sum(values) / len(values), 4) if values else None,
            "median_pct": round(median(values), 4) if values else None,
            "positive_rate_pct": round(sum(value > 0 for value in values) / len(values) * 100, 2)
            if values
            else None,
        }

    return {
        "method": "Deduplicated ticker/setup/day signals; outcomes use later recorded scan prices, not OHLC fills.",
        "limitations": [
            "This is a directional opportunity proxy, not an executable portfolio backtest.",
            "Intraday target/stop order is unknown and repeated scan prices may miss touches.",
            "Only signals blocked primarily by capital/exposure after active non-capital gates are included.",
        ],
        "total_decision_records": total_records,
        "malformed_records": malformed,
        "initial_buy_signals": initial_buy_signals,
        "final_buy_signals": final_buy_signals,
        "unique_capital_blocked_candidates": len(enriched),
        "blocker_categories": dict(reason_counts.most_common()),
        "outcomes": {f"{horizon}_scan_days": outcome_summary(values) for horizon, values in returns.items()},
        "top_candidates_by_5d_return": sorted(
            [item for item in enriched if item.get("return_after_5_scan_days_pct") is not None],
            key=lambda item: to_float(item.get("return_after_5_scan_days_pct")),
            reverse=True,
        )[:20],
        "bottom_candidates_by_5d_return": sorted(
            [item for item in enriched if item.get("return_after_5_scan_days_pct") is not None],
            key=lambda item: to_float(item.get("return_after_5_scan_days_pct")),
        )[:20],
    }


def build_capital_replay_report(
    *,
    workbook_path: Path,
    decision_dir: Path,
    starting_capital: float = 100_000.0,
    idle_cash_annual_yield: float = 0.05,
    include_candidate_analysis: bool = True,
    scenarios: Iterable[CapitalScenario] = DEFAULT_SCENARIOS,
) -> dict[str, Any]:
    trades = load_closed_trades(workbook_path)
    scenario_results = [
        replay_scenario(
            trades,
            scenario,
            starting_capital=starting_capital,
            idle_cash_annual_yield=idle_cash_annual_yield,
        )
        for scenario in scenarios
    ]
    first_entry = trades[0]["entry_timestamp"].isoformat() if trades else None
    last_exit = max(trade["exit_timestamp"] for trade in trades).isoformat() if trades else None
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "mode": "READ_ONLY_CAPITAL_REPLAY",
        "active_trading_logic_changed": False,
        "workbook": str(workbook_path),
        "decision_dir": str(decision_dir),
        "sample": {
            "closed_trades": len(trades),
            "first_entry": first_entry,
            "last_exit": last_exit,
            "starting_capital": starting_capital,
            "idle_cash_annual_yield_assumption": idle_cash_annual_yield,
        },
        "methodology": {
            "exact_layer": "Replays original entries and exits with alternative quantities and portfolio constraints.",
            "counterfactual_limit": "Exact layer cannot prove outcomes for trades the active agent never opened.",
            "cash_yield": "Hypothetical simple annual yield accrued on uninvested cash between trade events.",
            "drawdown": "Realized-equity drawdown; intratrade mark-to-market drawdown is not reconstructed.",
        },
        "scenarios": scenario_results,
    }
    if include_candidate_analysis:
        report["capital_blocked_candidate_analysis"] = analyze_capital_blocked_candidates(decision_dir)
    return report


def report_markdown(report: dict[str, Any]) -> str:
    sample = report["sample"]
    lines = [
        "# Market Lens Capital Replay",
        "",
        "This report is read-only. It did not change active trading rules, the workbook, or positions.",
        "",
        "## Sample",
        "",
        f"- Closed trades: {sample['closed_trades']}",
        f"- Period: {sample['first_entry']} to {sample['last_exit']}",
        f"- Starting capital: ${sample['starting_capital']:,.2f}",
        f"- Idle-cash yield assumption: {sample['idle_cash_annual_yield_assumption'] * 100:.2f}% annual",
        "",
        "## Scenario Comparison",
        "",
        "| Scenario | Trading PnL | Return | With cash yield | Avg exposure | Max exposure | PF | Drawdown |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for scenario in report["scenarios"]:
        lines.append(
            "| {scenario} | ${pnl:,.2f} | {ret:.2f}% | {total:.2f}% | {avg:.2f}% | "
            "{max_exp:.2f}% | {pf} | -${dd:,.2f} |".format(
                scenario=scenario["scenario"],
                pnl=scenario["trading_pnl"],
                ret=scenario["trading_return_pct"],
                total=scenario["total_return_with_idle_cash_yield_pct"],
                avg=scenario["average_time_weighted_exposure_pct"],
                max_exp=scenario["max_exposure_pct"],
                pf=scenario["profit_factor"] if scenario["profit_factor"] is not None else "n/a",
                dd=abs(scenario["max_realized_drawdown"]),
            )
        )
    blocked = report.get("capital_blocked_candidate_analysis")
    if blocked:
        lines.extend(
            [
                "",
                "## Capital-Blocked Opportunity Proxy",
                "",
                f"- Unique candidates: {blocked['unique_capital_blocked_candidates']}",
                f"- Initial BUY signals: {blocked['initial_buy_signals']}",
                f"- Final BUY signals: {blocked['final_buy_signals']}",
                f"- Blockers: {json.dumps(blocked['blocker_categories'], sort_keys=True)}",
                "",
                "These outcomes use later recorded scan prices. They are diagnostic and must not be read as executable backtest returns.",
            ]
        )
        for horizon, values in blocked["outcomes"].items():
            lines.append(
                f"- {horizon}: n={values['sample_size']}, avg={values['average_pct']}%, "
                f"median={values['median_pct']}%, positive={values['positive_rate_pct']}%"
            )
    lines.extend(
        [
            "",
            "## Limitations",
            "",
            "- The exact replay changes size only for trades that were actually opened.",
            "- It reuses recorded exits and does not model price impact or altered monitor timing.",
            "- Drawdown is based on realized equity, not intraday mark-to-market equity.",
            "- No scenario should be activated from this sample alone; use shadow validation first.",
            "",
        ]
    )
    return "\n".join(lines)


def write_capital_replay_report(report: dict[str, Any], output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = output_dir / f"capital_replay_{stamp}.json"
    md_path = output_dir / f"capital_replay_{stamp}.md"
    json_path.write_text(json.dumps(report, indent=2, sort_keys=True, default=str), encoding="utf-8")
    md_path.write_text(report_markdown(report), encoding="utf-8")
    return {"json": json_path, "markdown": md_path}
