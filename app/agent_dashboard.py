from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
from collections import defaultdict, deque
from datetime import datetime, time
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from app.smart_universe import base_universe, company_name_for


TRACKER_NAME = "market_lens_agent_portfolio_budget_100k.xlsx"
_SECTOR_MAP: dict[str, str] | None = None


def build_agent_dashboard(project_root: Path, selected_date: str | None = None) -> dict[str, Any]:
    tracker_path = project_root / "agent_tracker" / TRACKER_NAME
    results_dir = project_root / "agent_results"
    screenshot_dir = results_dir / "screenshots"
    summary_dir = results_dir / "summaries"
    position_monitor_dir = results_dir / "position_monitor"
    decision_dir = results_dir / "decisions"

    if not tracker_path.exists():
        return {
            "status": "missing_tracker",
            "error": f"{tracker_path} was not found.",
            "tracker_url": "/agent/tracker",
        }

    workbook_path = historical_tracker_copy(tracker_path) if selected_date else tracker_path
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    settings = read_settings(wb)
    currency = str(settings.get("budget_currency") or "USD").upper()
    starting_capital = to_float(
        settings.get("starting_capital_usd") or settings.get("starting_capital_ils"),
        100_000.0,
    )

    updates = read_updates(wb)
    trades = read_trades(wb)
    latest_update = select_update(updates, selected_date)
    latest_run_timestamp = latest_update.get("timestamp")
    latest_scan_update = select_latest_scan_update(updates, latest_update)
    latest_scan_timestamp = latest_scan_update.get("timestamp")
    setup_rows = (
        read_setup_rows(wb, cutoff=latest_run_timestamp)
        if selected_date
        else read_setup_rows(wb, run_date=latest_scan_timestamp)
    )
    if selected_date and not setup_rows:
        decision_path = resolve_record_file(
            latest_scan_update.get("decision_jsonl") or latest_update.get("decision_jsonl"),
            decision_dir,
            ".jsonl",
        )
        setup_rows = read_decision_setup_rows(decision_path)
    if not setup_rows and not selected_date:
        setup_rows = read_setup_rows(wb, recent_limit=25)
    scoped_updates = filter_records_until(updates, latest_run_timestamp)
    scoped_trades = filter_records_until(trades, latest_run_timestamp)
    current_snapshot = not selected_date and latest_update == (updates[-1] if updates else {})
    open_positions = (
        read_open_positions(wb)
        if current_snapshot
        else reconstruct_open_positions(scoped_trades, setup_rows, latest_run_timestamp)
    )
    realized = compute_realized_pnl(scoped_trades)
    full_trade_performance = compute_full_trade_performance(scoped_trades)
    annotated_trades = realized.get("trades", scoped_trades)
    latest_monitor_update = select_latest_monitor_update(updates, latest_update)

    cash = to_float(latest_update.get("cash_ils"), compute_cash(scoped_trades, starting_capital))
    exposure = to_float(
        latest_update.get("exposure_ils"),
        sum(to_float(position.get("exposure_ils")) for position in open_positions),
    )
    open_risk = to_float(
        latest_update.get("open_risk_ils"),
        sum(to_float(position.get("open_risk_ils")) for position in open_positions),
    )
    equity = round(cash + exposure, 2)
    total_pnl = round(equity - starting_capital, 2)
    total_pnl_pct = round(total_pnl / starting_capital * 100, 2) if starting_capital else 0

    latest_summary_path = resolve_record_file(latest_update.get("summary"), summary_dir, ".md")
    if not latest_summary_path:
        latest_summary_path = resolve_record_file(latest_update.get("summary"), position_monitor_dir, ".md")
    if not selected_date and not latest_summary_path:
        latest_summary_path = resolve_latest_file(summary_dir, ".md")
    latest_summary = latest_summary_path.read_text(encoding="utf-8") if latest_summary_path else ""
    screenshot_source = latest_update.get("screenshot") or latest_scan_update.get("screenshot")
    if not selected_date and not screenshot_source:
        screenshot_source = resolve_latest_file(screenshot_dir, ".png")
    latest_screenshot = resolve_asset_url(screenshot_source)

    latest_setups = [
        row for row in setup_rows if latest_scan_timestamp and row.get("run_date") == latest_scan_timestamp
    ]
    if not latest_setups and not selected_date:
        latest_setups = setup_rows[-25:]

    action_counts: dict[str, int] = defaultdict(int)
    for setup in latest_setups:
        action_counts[str(setup.get("action") or "UNKNOWN")] += 1
    latest_decisions = [setup["decision_json"] for setup in latest_setups if setup.get("decision_json")]
    latest_dt = parse_timestamp(latest_update.get("timestamp"))
    daily_summary = load_period_summary(summary_dir, "daily", latest_dt)
    weekly_summary = load_period_summary(summary_dir, "weekly", latest_dt)
    summary = {
        "starting_capital_ils": starting_capital,
        "currency": currency,
        "cash_ils": cash,
        "exposure_ils": exposure,
        "equity_ils": equity,
        "total_pnl_ils": total_pnl,
        "total_pnl_pct": total_pnl_pct,
        "realized_pnl_ils": round(realized["total"], 2),
        "unrealized_pnl_ils": round(
            sum(to_float(position.get("unrealized_pnl_ils")) for position in open_positions),
            2,
        ),
        "open_risk_ils": round(open_risk, 2),
        "open_positions": len(open_positions),
        "closed_trades": len(realized["closed"]),
        "wins": realized["wins"],
        "losses": realized["losses"],
        "win_rate": round(realized["wins"] / len(realized["closed"]) * 100, 2)
        if realized["closed"]
        else 0,
        "exit_event_wins": realized["wins"],
        "exit_event_losses": realized["losses"],
        "exit_event_breakeven": realized.get("breakeven", 0),
        "exit_event_win_rate": round(realized["wins"] / len(realized["closed"]) * 100, 2)
        if realized["closed"]
        else 0,
        "full_trades": full_trade_performance["closed_count"],
        "full_trade_wins": full_trade_performance["wins"],
        "full_trade_losses": full_trade_performance["losses"],
        "full_trade_breakeven": full_trade_performance["breakeven"],
        "full_trade_win_rate": full_trade_performance["win_rate"],
        "full_trade_realized_pnl_ils": full_trade_performance["total_pnl_ils"],
        "open_full_trades": full_trade_performance["open_count"],
    }

    dashboard = {
        "status": "ok",
        "tracker_url": "/agent/tracker",
        "github_actions_url": "https://github.com/AviramDahan/market-lens-scanner/actions/workflows/market-lens-agent.yml",
        "snapshot": {
            "selected_date": selected_date or "",
            "resolved_timestamp": latest_run_timestamp,
            "is_historical": bool(selected_date),
            "available_dates": available_dates(updates),
        },
        "summary": summary,
        "latest_run": {
            "timestamp": latest_update.get("timestamp"),
            "run_id": latest_update.get("run_id"),
            "tickers": latest_update.get("tickers", []),
            "valid_setups": latest_update.get("valid_setups", 0),
            "trade_ready_setups": count_trade_ready(latest_setups),
            "actions_summary": latest_update.get("actions_summary", ""),
            "screenshot_url": latest_screenshot,
            "summary_url": resolve_asset_url(latest_summary_path),
            "decision_jsonl_url": resolve_asset_url(
                latest_scan_update.get("decision_jsonl") or latest_update.get("decision_jsonl") or resolve_latest_file(decision_dir, ".jsonl")
            ),
            "summary_text": latest_summary,
            "action_counts": dict(sorted(action_counts.items())),
            "market_regime": latest_decisions[0].get("market_regime", "") if latest_decisions else "",
            "latest_scan_run_id": latest_scan_update.get("run_id"),
            "latest_scan_timestamp": latest_scan_update.get("timestamp"),
        },
        # Keep current-run assets early in the snapshot. The deployed app syncs
        # referenced files lazily, so latest charts should be discovered before
        # older historical trade media.
        "latest_setups": latest_setups,
        "latest_decisions": latest_decisions,
        "decision_diagnostics": build_decision_diagnostics(latest_setups),
        "daily_summary": daily_summary,
        "weekly_summary": weekly_summary,
        "system_health": build_system_health(
            tracker_path=tracker_path,
            updates=scoped_updates,
            latest_update=latest_update,
            latest_scan_update=latest_scan_update,
            latest_monitor_update=latest_monitor_update,
        ),
        "open_positions": open_positions,
        "position_attention": build_position_attention(open_positions),
        "risk_dashboard": build_risk_dashboard(open_positions, summary, latest_decisions),
        "position_timeline": build_position_timeline(open_positions),
        "equity_curve": build_equity_curve(scoped_updates, starting_capital),
        "recent_trades": annotated_trades,
        "closed_trades": realized["closed"],
        "full_trade_performance": full_trade_performance,
        "score_calibration": build_score_calibration(realized["closed"]),
        "recent_runs": scoped_updates[-20:],
        "pagination": {
            "actions": {"total": len(latest_setups), "returned": len(latest_setups), "offset": 0},
            "trades": {
                "total": len(annotated_trades),
                "closed_total": len(realized["closed"]),
                "returned": len(annotated_trades),
                "offset": 0,
            },
        },
    }
    return sanitize_dashboard_media_urls(dashboard, project_root)


def historical_tracker_copy(tracker_path: Path) -> Path:
    """Cache a workbook without the oversized watchlist sheet for historical views."""
    try:
        stat = tracker_path.stat()
        cache_dir = Path(tempfile.gettempdir()) / "market-lens-dashboard"
        cache_dir.mkdir(parents=True, exist_ok=True)
        source_key = hashlib.sha1(str(tracker_path.resolve()).encode("utf-8")).hexdigest()[:12]
        cached = cache_dir / f"tracker-{source_key}-{stat.st_size}-{stat.st_mtime_ns}.xlsx"
        if cached.exists():
            return cached

        pending = cached.with_suffix(".tmp")
        empty_sheet = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b"<sheetData/></worksheet>"
        )
        with ZipFile(tracker_path) as source, ZipFile(pending, "w", ZIP_DEFLATED) as target:
            for item in source.infolist():
                content = empty_sheet if item.filename == "xl/worksheets/sheet2.xml" else source.read(item.filename)
                target.writestr(item, content)
        os.replace(pending, cached)
        return cached
    except (OSError, ValueError):
        return tracker_path


def compact_agent_dashboard_payload(
    dashboard: dict[str, Any],
    *,
    action_limit: int = 10,
    trade_limit: int = 10,
) -> dict[str, Any]:
    """Return the same dashboard shape with heavy expandable collections trimmed."""
    if dashboard.get("status") != "ok":
        return dashboard

    actions = dashboard.get("latest_setups") if isinstance(dashboard.get("latest_setups"), list) else []
    trades = dashboard.get("recent_trades") if isinstance(dashboard.get("recent_trades"), list) else []
    closed = dashboard.get("closed_trades") if isinstance(dashboard.get("closed_trades"), list) else []
    compact = dict(dashboard)
    compact["latest_setups"] = actions[: max(0, action_limit)]
    compact["latest_decisions"] = []
    compact["recent_trades"] = list(reversed(trades))[: max(0, trade_limit)]
    compact["closed_trades"] = []
    compact["recent_runs"] = []
    compact["full_trade_performance"] = {}
    compact["equity_curve"] = downsample_dashboard_series(dashboard.get("equity_curve"), max_points=240)
    compact["decision_diagnostics"] = compact_decision_diagnostics(dashboard.get("decision_diagnostics"))
    compact["daily_summary"] = compact_performance_summary(dashboard.get("daily_summary"))
    compact["weekly_summary"] = compact_performance_summary(dashboard.get("weekly_summary"))
    compact["payload"] = {
        "mode": "compact",
        "generated_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
    }
    compact["pagination"] = {
        **(dashboard.get("pagination") if isinstance(dashboard.get("pagination"), dict) else {}),
        "actions": {
            "total": len(actions),
            "returned": len(compact["latest_setups"]),
            "offset": 0,
            "has_more": len(actions) > len(compact["latest_setups"]),
        },
        "trades": {
            "total": len(trades),
            "closed_total": len(closed),
            "returned": len(compact["recent_trades"]),
            "offset": 0,
            "has_more": len(trades) > len(compact["recent_trades"]),
        },
    }
    latest_run = dict(compact.get("latest_run")) if isinstance(compact.get("latest_run"), dict) else None
    if isinstance(latest_run, dict):
        latest_run["summary_text"] = trim_text(str(latest_run.get("summary_text") or ""), 24_000)
        compact["latest_run"] = latest_run
    return compact


def downsample_dashboard_series(value: Any, *, max_points: int) -> list[Any]:
    rows = list(value) if isinstance(value, list) else []
    if max_points <= 0:
        return []
    if len(rows) <= max_points:
        return rows
    if max_points == 1:
        return [rows[-1]]
    last_index = len(rows) - 1
    indexes = {round(index * last_index / (max_points - 1)) for index in range(max_points)}
    return [rows[index] for index in sorted(indexes)]


def compact_decision_diagnostics(value: Any) -> dict[str, Any]:
    diagnostics = dict(value) if isinstance(value, dict) else {}
    drilldowns = diagnostics.get("drilldowns") if isinstance(diagnostics.get("drilldowns"), dict) else {}
    watch_ready = drilldowns.get("WATCH_READY") if isinstance(drilldowns.get("WATCH_READY"), list) else []
    diagnostics["drilldowns"] = {"WATCH_READY": watch_ready[:4]}
    return diagnostics


def compact_performance_summary(value: Any) -> dict[str, Any]:
    summary = value if isinstance(value, dict) else {}
    keys = {
        "BUY_SIMULATED_count",
        "WATCH_READY_count",
        "WATCH_READY_unique_count",
        "WATCH_READY_conversion",
        "WATCH_READY_session_breakdown",
        "recommendations_for_next_week",
    }
    return {key: summary[key] for key in keys if key in summary}


def write_diagnostic_snapshot(project_root: Path, dashboard: dict[str, Any]) -> Path | None:
    if dashboard.get("status") != "ok":
        return None
    latest_run = dashboard.get("latest_run") if isinstance(dashboard.get("latest_run"), dict) else {}
    run_id = sanitize_snapshot_id(str(latest_run.get("run_id") or latest_run.get("timestamp") or "latest"))
    payload = {
        "status": "ok",
        "generated_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "run_id": latest_run.get("run_id", ""),
        "run_timestamp": latest_run.get("timestamp", ""),
        "snapshot": dashboard.get("snapshot", {}),
        "latest_run": {
            "timestamp": latest_run.get("timestamp", ""),
            "run_id": latest_run.get("run_id", ""),
            "tickers": latest_run.get("tickers", []),
            "valid_setups": latest_run.get("valid_setups", 0),
            "trade_ready_setups": latest_run.get("trade_ready_setups", 0),
            "action_counts": latest_run.get("action_counts", {}),
            "market_regime": latest_run.get("market_regime", ""),
        },
        "decision_diagnostics": dashboard.get("decision_diagnostics", {}),
    }
    diagnostics_dir = project_root / "agent_results" / "diagnostics"
    diagnostics_dir.mkdir(parents=True, exist_ok=True)
    path = diagnostics_dir / f"diagnostics_{run_id}.json"
    path.write_text(json.dumps(payload, default=str, separators=(",", ":")), encoding="utf-8")
    return path


def sanitize_snapshot_id(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_\-]+", "_", value).strip("_")
    return cleaned[:80] or "latest"


def dashboard_section_payload(
    dashboard: dict[str, Any],
    *,
    section: str,
    offset: int = 0,
    limit: int = 10,
    diagnostic_key: str | None = None,
    sector: str | None = None,
    setup_type: str | None = None,
    chart_filter: str = "all",
    confirmation: str = "all",
    sort: str = "closest",
) -> dict[str, Any]:
    offset = max(0, offset)
    limit = max(1, min(limit, 100))
    if dashboard.get("status") != "ok":
        return dashboard
    if section == "actions":
        items = dashboard.get("latest_setups") if isinstance(dashboard.get("latest_setups"), list) else []
    elif section == "trades":
        items = list(reversed(dashboard.get("recent_trades") if isinstance(dashboard.get("recent_trades"), list) else []))
    elif section == "diagnostics":
        return dashboard_diagnostic_payload(
            dashboard,
            diagnostic_key=diagnostic_key or "WATCH_READY",
            offset=offset,
            limit=limit,
            sector=sector,
            setup_type=setup_type,
            chart_filter=chart_filter,
            confirmation=confirmation,
            sort=sort,
        )
    else:
        return {"status": "error", "error": f"Unknown dashboard section: {section}"}
    page = items[offset : offset + limit]
    if section == "trades":
        page = list(page)
    return {
        "status": "ok",
        "section": section,
        "offset": offset,
        "limit": limit,
        "total": len(items),
        "has_more": offset + len(page) < len(items),
        "items": page,
    }


def dashboard_diagnostic_payload(
    dashboard: dict[str, Any],
    *,
    diagnostic_key: str,
    offset: int = 0,
    limit: int = 60,
    sector: str | None = None,
    setup_type: str | None = None,
    chart_filter: str = "all",
    confirmation: str = "all",
    sort: str = "closest",
) -> dict[str, Any]:
    diagnostics = dashboard.get("decision_diagnostics") if isinstance(dashboard.get("decision_diagnostics"), dict) else {}
    drilldowns = diagnostics.get("drilldowns") if isinstance(diagnostics.get("drilldowns"), dict) else {}
    items = drilldowns.get(diagnostic_key) if isinstance(drilldowns.get(diagnostic_key), list) else []
    facets = diagnostic_facets(items)
    filtered = filter_diagnostic_items(
        items,
        sector=sector,
        setup_type=setup_type,
        chart_filter=chart_filter,
        confirmation=confirmation,
    )
    sorted_items = sort_diagnostic_items(filtered, sort=sort)
    page = sorted_items[offset : offset + limit]
    return {
        "status": "ok",
        "section": "diagnostics",
        "diagnostic_key": diagnostic_key,
        "offset": offset,
        "limit": limit,
        "total": len(sorted_items),
        "unfiltered_total": len(items),
        "has_more": offset + len(page) < len(sorted_items),
        "facets": facets,
        "filters": {
            "sector": sector or "",
            "setup_type": setup_type or "",
            "chart_filter": chart_filter,
            "confirmation": confirmation,
            "sort": sort,
        },
        "items": page,
    }


def diagnostic_facets(items: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    sectors: dict[str, int] = defaultdict(int)
    setup_types: dict[str, int] = defaultdict(int)
    chart_counts = {"with_chart": 0, "missing_chart": 0}
    confirmation_counts = {"passed": 0, "missing": 0}
    for item in items:
        sector = str(item.get("sector") or "Unknown")
        setup = str(item.get("setup_type") or "Setup")
        sectors[sector] += 1
        setup_types[setup] += 1
        chart_counts["with_chart" if item.get("chart_url") else "missing_chart"] += 1
        confirmation_counts["passed" if item.get("entry_confirmation_passed") else "missing"] += 1
    return {
        "sectors": count_options(sectors),
        "setup_types": count_options(setup_types),
        "charts": count_options(chart_counts),
        "confirmation": count_options(confirmation_counts),
    }


def count_options(counts: dict[str, int]) -> list[dict[str, Any]]:
    return [
        {"value": key, "count": value}
        for key, value in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
        if value
    ]


def filter_diagnostic_items(
    items: list[dict[str, Any]],
    *,
    sector: str | None,
    setup_type: str | None,
    chart_filter: str,
    confirmation: str,
) -> list[dict[str, Any]]:
    filtered = list(items)
    if sector:
        filtered = [item for item in filtered if str(item.get("sector") or "Unknown") == sector]
    if setup_type:
        filtered = [item for item in filtered if str(item.get("setup_type") or "Setup") == setup_type]
    if chart_filter == "with_chart":
        filtered = [item for item in filtered if item.get("chart_url")]
    elif chart_filter == "missing_chart":
        filtered = [item for item in filtered if not item.get("chart_url")]
    if confirmation == "passed":
        filtered = [item for item in filtered if item.get("entry_confirmation_passed")]
    elif confirmation == "missing":
        filtered = [item for item in filtered if not item.get("entry_confirmation_passed")]
    return filtered


def sort_diagnostic_items(items: list[dict[str, Any]], *, sort: str) -> list[dict[str, Any]]:
    if sort == "score":
        key = lambda item: (to_float(item.get("setup_score")), to_float(item.get("weighted_net_rr") or item.get("net_rr")))
        return sorted(items, key=key, reverse=True)
    if sort == "rr":
        key = lambda item: (to_float(item.get("weighted_net_rr") or item.get("net_rr")), to_float(item.get("setup_score")))
        return sorted(items, key=key, reverse=True)
    if sort == "ticker":
        return sorted(items, key=lambda item: str(item.get("ticker") or ""))
    key = lambda item: (
        item.get("entry_confirmation_passed") is True,
        -diagnostic_buy_zone_distance(item),
        to_float(item.get("setup_score")),
        to_float(item.get("weighted_net_rr") or item.get("net_rr")),
    )
    return sorted(items, key=key, reverse=True)


def diagnostic_buy_zone_distance(item: dict[str, Any]) -> float:
    price = to_float(item.get("current_price_usd"))
    low = to_float(item.get("buy_zone_low"))
    high = to_float(item.get("buy_zone_high"))
    if not price or not low or not high:
        return 999.0
    lower, upper = sorted((low, high))
    if lower <= price <= upper:
        return 0.0
    nearest = lower if price < lower else upper
    return abs(price - nearest) / price * 100 if price else 999.0


def load_period_summary(summary_dir: Path, period: str, timestamp: datetime) -> dict[str, Any]:
    if timestamp == datetime.min:
        return {}
    if period == "daily":
        path = summary_dir / f"daily_summary_{timestamp.date().isoformat()}.json"
    elif period == "weekly":
        year, week, _weekday = timestamp.date().isocalendar()
        path = summary_dir / f"weekly_summary_{year}-W{week:02d}.json"
    else:
        return {}
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def build_decision_diagnostics(setups: list[dict[str, Any]]) -> dict[str, Any]:
    blockers: dict[str, int] = defaultdict(int)
    entry_blockers: dict[str, dict[str, Any]] = {}
    action_counts: dict[str, int] = defaultdict(int)
    drilldowns: dict[str, list[dict[str, Any]]] = defaultdict(list)
    near_misses: list[dict[str, Any]] = []
    closest_to_entry: list[dict[str, Any]] = []

    for setup in setups:
        action = str(setup.get("action") or "UNKNOWN").upper()
        action_counts[action] += 1
        decision = setup.get("decision_json") or {}
        setup_type = str(setup.get("setup_type") or decision.get("setup_type") or "")
        reason = str(setup.get("reason") or decision.get("reason") or "")
        warnings = " ".join(str(warning) for warning in decision.get("warnings") or [])
        text = f"{reason} {warnings}".lower()
        drilldown_item = diagnostic_drilldown_item(setup, decision, action, setup_type, reason)
        missing_conditions = entry_missing_conditions(setup, decision, action, setup_type, text)
        drilldown_item["missing_conditions"] = missing_conditions
        drilldown_item["entry_readiness_score"] = entry_readiness_score(setup, decision, action, setup_type, missing_conditions)

        if action == "BUY_SIMULATED":
            drilldowns["BUY"].append(drilldown_item)
        if action == "WATCH_READY" or any(
            str(warning).upper().startswith("WATCH_READY:") for warning in decision.get("warnings") or []
        ):
            drilldowns["WATCH_READY"].append(drilldown_item)
        if bool(decision.get("capital_blocked_only")) or str(
            decision.get("entry_eligibility_status") or ""
        ).upper() == "QUALIFIED_CAPITAL_BLOCKED":
            blockers["Qualified but capital blocked"] += 1
            drilldowns["CAPITAL_BLOCKED"].append(drilldown_item)

        if setup_type.lower() == "no trade" or "no trade result" in text:
            blockers["No Trade"] += 1
        if "risk/reward" in text or "net r/r" in text or "weighted" in text:
            blockers["R/R below gate"] += 1
            drilldowns["RR_BLOCKED"].append(drilldown_item)
        if "setup score" in text:
            blockers["Setup score below gate"] += 1
            drilldowns["SCORE_BLOCKED"].append(drilldown_item)
        if "entry confirmation" in text or "confirmed entry" in text or "confirmation failed" in text:
            blockers["Entry confirmation missing"] += 1
            drilldowns["CONFIRM_BLOCKED"].append(drilldown_item)
        if "earnings blackout" in text:
            blockers["Earnings blackout"] += 1
            drilldowns["WEAK_EARNINGS"].append(drilldown_item)
        if "bear market regime blocks" in text:
            blockers["BEAR blocks new buys"] += 1
        if "sector regime is weak" in text or "weak sector" in text:
            blockers["Weak sector"] += 1
            drilldowns["WEAK_EARNINGS"].append(drilldown_item)

        for condition in missing_conditions:
            register_entry_blocker(entry_blockers, condition, drilldown_item)

        if action in {"WATCH", "WATCH_READY", "SKIP"} and setup_type and setup_type.lower() != "no trade":
            near_misses.append(drilldown_item)
            closest_to_entry.append(drilldown_item)

    near_misses.sort(
        key=lambda item: (
            item.get("action") == "WATCH_READY",
            to_float(item.get("setup_score")),
            to_float(item.get("weighted_net_rr") or item.get("net_rr")),
        ),
        reverse=True,
    )
    closest_to_entry.sort(
        key=lambda item: (
            to_float(item.get("entry_readiness_score")),
            item.get("action") == "WATCH_READY",
            to_float(item.get("setup_score")),
            to_float(item.get("weighted_net_rr") or item.get("net_rr")),
        ),
        reverse=True,
    )
    return {
        "total_results": len(setups),
        "action_counts": dict(sorted(action_counts.items())),
        "blockers": dict(sorted(blockers.items(), key=lambda item: item[1], reverse=True)),
        "entry_blockers_summary": sorted_entry_blockers(entry_blockers),
        "why_no_buys": build_why_no_buys(blockers, action_counts, len(setups)),
        "watch_ready_funnel": build_watch_ready_funnel(setups),
        "near_misses": near_misses[:10],
        "closest_to_entry": closest_to_entry[:10],
        "drilldowns": {
            key: sorted_unique_diagnostic_items(items)[:60] for key, items in sorted(drilldowns.items())
        },
        "watch_ready_count": sum(
            1
            for setup in setups
            if str(setup.get("action") or "").upper() == "WATCH_READY"
            or any(str(warning).upper().startswith("WATCH_READY:") for warning in (setup.get("decision_json") or {}).get("warnings") or [])
        ),
    }


def register_entry_blocker(
    entry_blockers: dict[str, dict[str, Any]],
    condition: dict[str, Any],
    item: dict[str, Any],
) -> None:
    key = str(condition.get("key") or condition.get("label") or "unknown")
    current = entry_blockers.setdefault(
        key,
        {
            "key": key,
            "label": condition.get("label") or key,
            "count": 0,
            "severity": condition.get("severity") or "warn",
            "detail": condition.get("detail") or "",
            "examples": [],
        },
    )
    current["count"] += 1
    if len(current["examples"]) < 5:
        current["examples"].append(
            {
                "ticker": item.get("ticker"),
                "action": item.get("action"),
                "setup_type": item.get("setup_type"),
                "readiness": item.get("entry_readiness_score"),
                "reason": item.get("reason"),
            }
        )


def sorted_entry_blockers(entry_blockers: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    severity_rank = {"fail": 3, "warn": 2, "need": 1}
    return sorted(
        entry_blockers.values(),
        key=lambda item: (to_int(item.get("count")), severity_rank.get(str(item.get("severity")), 0)),
        reverse=True,
    )


def entry_missing_conditions(
    setup: dict[str, Any],
    decision: dict[str, Any],
    action: str,
    setup_type: str,
    text: str,
) -> list[dict[str, Any]]:
    if action in {"BUY_SIMULATED", "HOLD", "TAKE_PARTIAL_PROFIT", "TAKE_PROFIT", "EXIT_STOP"}:
        return []

    conditions: list[dict[str, Any]] = []
    normalized_setup = str(setup_type or "").lower()
    setup_score = to_float(decision.get("setup_score") or setup.get("score"))
    min_setup_score = to_float(decision.get("minimum_setup_score_required"))
    net_rr = to_float(decision.get("weighted_net_rr") or decision.get("net_rr"))
    min_rr = minimum_net_rr_for_decision(decision)
    net_rr_1 = to_float(decision.get("net_rr_1") or decision.get("gross_rr_1"))
    target_status = str(decision.get("target_feasibility_status") or "").upper()
    session_phase = str(decision.get("market_session_phase") or decision.get("market_session") or "").upper()

    if normalized_setup == "no trade" or "no trade result" in text:
        conditions.append(blocker_condition("no_trade", "No technical setup", "No actionable setup structure was detected.", "warn"))
    if str(decision.get("market_regime") or "").upper() == "BEAR":
        conditions.append(blocker_condition("market_bear", "BEAR market", "Bear market regime blocks new paper buys.", "fail"))
    if str(decision.get("sector_regime") or "").upper() == "WEAK" or "weak sector" in text:
        conditions.append(blocker_condition("weak_sector", "Weak sector", "Sector regime is weak, so auto-buy eligibility is blocked.", "fail"))
    if min_setup_score and setup_score < min_setup_score:
        conditions.append(
            blocker_condition(
                "setup_score",
                "Setup score below gate",
                f"Score {setup_score:.2f}, needs {min_setup_score:.2f}.",
                "warn",
            )
        )
    if min_rr and net_rr < min_rr and min_rr < 900:
        conditions.append(
            blocker_condition(
                "net_rr",
                "Weighted/net R/R below gate",
                f"Weighted/net R/R {net_rr:.2f}x, needs {min_rr:.2f}x.",
                "warn",
            )
        )
    if net_rr_1 and net_rr_1 < 0.8:
        conditions.append(
            blocker_condition(
                "primary_rr",
                "TP1 R/R too low",
                f"Primary R/R {net_rr_1:.2f}x is below the minimum TP1 gate.",
                "warn",
            )
        )
    if bool(decision.get("earnings_blackout")) or "earnings blackout" in text:
        conditions.append(blocker_condition("earnings", "Earnings blackout", "Earnings timing blocks a fresh entry.", "fail"))
    if bool(decision.get("sector_exposure_limit_exceeded")):
        conditions.append(blocker_condition("sector_exposure", "Sector exposure cap", "The trade would exceed sector exposure limits.", "fail"))
    if bool(decision.get("factor_exposure_limit_exceeded")):
        conditions.append(blocker_condition("factor_exposure", "Factor exposure cap", "The trade would exceed factor/theme exposure limits.", "fail"))
    if bool(decision.get("capital_blocked_only")):
        conditions.append(
            blocker_condition(
                "capital_capacity",
                "Qualified, capital blocked",
                "All entry-quality gates passed, but exposure, heat, or concentration left no executable size.",
                "need",
            )
        )
    if bool(decision.get("cooldown_active")) or "stop-loss cooldown" in text:
        conditions.append(blocker_condition("cooldown", "Stop cooldown", "Recent stop-loss cooldown blocks re-entry.", "warn"))
    if bool(decision.get("correlation_warning")):
        ticker = decision.get("highest_correlation_ticker") or "an open position"
        value = to_float(decision.get("highest_correlation_value"))
        conditions.append(
            blocker_condition("correlation", "Correlation warning", f"High correlation with {ticker} ({value:.2f}).", "warn")
        )
    if not bool(decision.get("entry_confirmation_passed")) and (
        "entry confirmation" in text
        or "confirmed entry" in text
        or action in {"WATCH", "WATCH_READY"}
        or bool(decision.get("regular_session_confirmation_required"))
    ):
        conditions.append(
            blocker_condition(
                "entry_confirmation",
                "Entry confirmation missing",
                "Needs completed-candle confirmation before BUY_SIMULATED.",
                "warn",
            )
        )
    if bool(decision.get("regular_session_confirmation_required")) or "outside regular market hours" in text or session_phase in {"PRE_MARKET", "AFTER_MARKET", "OVERNIGHT", "WEEKEND"}:
        conditions.append(blocker_condition("session", "Regular-session confirmation", "Off-hours candidate must be reviewed during regular market hours.", "need"))
    if target_status and target_status not in {"OK", "UNKNOWN"}:
        conditions.append(blocker_condition("target_quality", "Target quality", f"Target feasibility status: {target_status}.", "warn"))
    if "target 1 is too close" in text:
        conditions.append(blocker_condition("target1_distance", "TP1 too close", "Target 1 is too close versus daily ATR.", "warn"))

    deduped: dict[str, dict[str, Any]] = {}
    for condition in conditions:
        deduped[str(condition["key"])] = condition
    return list(deduped.values())


def blocker_condition(key: str, label: str, detail: str, severity: str) -> dict[str, Any]:
    return {"key": key, "label": label, "detail": detail, "severity": severity}


def entry_readiness_score(
    setup: dict[str, Any],
    decision: dict[str, Any],
    action: str,
    setup_type: str,
    missing_conditions: list[dict[str, Any]],
) -> int:
    if action == "BUY_SIMULATED":
        return 100
    if str(setup_type or "").lower() == "no trade":
        return 0

    setup_score = max(0.0, min(1.0, to_float(decision.get("setup_score") or setup.get("score"))))
    min_score = to_float(decision.get("minimum_setup_score_required")) or 0.45
    rr = to_float(decision.get("weighted_net_rr") or decision.get("net_rr"))
    min_rr = minimum_net_rr_for_decision(decision)
    rr_component = 0.0 if min_rr >= 900 else max(0.0, min(1.0, rr / min_rr if min_rr else 0.0))
    score_component = max(0.0, min(1.0, setup_score / min_score if min_score else setup_score))
    confirmation_component = 1.0 if decision.get("entry_confirmation_passed") else 0.0
    sector_component = 0.0 if str(decision.get("sector_regime") or "").upper() == "WEAK" else 1.0
    target_component = 1.0 if str(decision.get("target_feasibility_status") or "").upper() in {"", "OK", "UNKNOWN"} else 0.45
    base = (
        setup_score * 30
        + score_component * 15
        + rr_component * 25
        + confirmation_component * 15
        + sector_component * 10
        + target_component * 5
    )
    for condition in missing_conditions:
        key = condition.get("key")
        if key in {"earnings", "market_bear"}:
            base -= 35
        elif key in {"sector_exposure", "factor_exposure", "weak_sector"}:
            base -= 18
        elif key in {"cooldown", "correlation"}:
            base -= 10
        elif key == "session":
            base -= 5
    return int(max(0, min(100, round(base))))


def diagnostic_drilldown_item(
    setup: dict[str, Any],
    decision: dict[str, Any],
    action: str,
    setup_type: str,
    reason: str,
) -> dict[str, Any]:
    return {
        "ticker": setup.get("ticker"),
        "company_name": setup.get("company_name", ""),
        "sector": setup.get("sector", "") or decision.get("sector", ""),
        "action": action,
        "setup_type": setup_type,
        "setup_score": round(to_float(setup.get("setup_score") or decision.get("setup_score") or setup.get("score")), 3),
        "net_rr": round(to_float(decision.get("net_rr") or setup.get("net_rr") or setup.get("risk_reward")), 3),
        "net_rr_1": round(to_float(decision.get("net_rr_1")), 3),
        "net_rr_2": round(to_float(decision.get("net_rr_2")), 3),
        "weighted_net_rr": round(to_float(decision.get("weighted_net_rr")), 3),
        "entry_confirmation_passed": bool(decision.get("entry_confirmation_passed")),
        "market_regime": decision.get("market_regime", ""),
        "sector_regime": decision.get("sector_regime", ""),
        "current_price_usd": setup.get("current_price_usd") or decision.get("price"),
        "buy_zone_low": setup.get("buy_zone_low") or decision.get("buy_zone_low"),
        "buy_zone_high": setup.get("buy_zone_high") or decision.get("buy_zone_high"),
        "stop_loss": setup.get("stop_loss") or decision.get("stop_loss"),
        "target_1": setup.get("target_1") or decision.get("target_1"),
        "target_2": setup.get("target_2") or decision.get("target_2"),
        "chart_url": setup.get("chart_url", ""),
        "selection_context": setup.get("selection_context", ""),
        "reason": reason,
    }


def build_why_no_buys(
    blockers: dict[str, int],
    action_counts: dict[str, int],
    total_results: int,
) -> list[dict[str, Any]]:
    buys = to_int(action_counts.get("BUY_SIMULATED"))
    if buys > 0:
        return [
            {
                "label": "BUY_SIMULATED found",
                "count": buys,
                "detail": "The latest scan opened paper positions; blockers below are only for rejected candidates.",
                "tone": "good",
            }
        ]

    if not total_results:
        return [
            {
                "label": "No scan results",
                "count": 0,
                "detail": "No setup rows were available for the selected snapshot.",
                "tone": "warn",
            }
        ]

    ranked = sorted(blockers.items(), key=lambda item: item[1], reverse=True)
    if not ranked:
        return [
            {
                "label": "No explicit blocker",
                "count": total_results,
                "detail": "The scan produced results, but no dominant rejection reason was recorded.",
                "tone": "warn",
            }
        ]

    details = {
        "Entry confirmation missing": "Candidates need a completed-candle trigger before BUY_SIMULATED.",
        "R/R below gate": "Weighted/net reward did not clear the active market-regime threshold.",
        "Setup score below gate": "Technical quality score was below the active regime floor.",
        "No Trade": "Scanner found no actionable setup structure for these tickers.",
        "Weak sector": "Sector regime reduced eligibility for new paper buys.",
        "Earnings blackout": "Earnings timing blocked fresh entries.",
        "BEAR blocks new buys": "Market regime blocked all new BUY_SIMULATED entries.",
    }
    return [
        {
            "label": label,
            "count": count,
            "detail": details.get(label, "Recorded by the active risk/decision layer."),
            "tone": "bad" if label in {"BEAR blocks new buys", "Earnings blackout"} else "warn",
        }
        for label, count in ranked[:5]
    ]


def build_watch_ready_funnel(setups: list[dict[str, Any]]) -> dict[str, Any]:
    detected_records = []
    unique_tickers = set()
    regular_reviewed = set()
    confirmation_passed = set()
    rr_passed = set()
    bought = set()

    for setup in setups:
        ticker = str(setup.get("ticker") or "").upper()
        if not ticker:
            continue
        action = str(setup.get("action") or "").upper()
        decision = setup.get("decision_json") or {}
        warnings = [str(warning).upper() for warning in decision.get("warnings") or []]
        watch_ready = action == "WATCH_READY" or any(warning.startswith("WATCH_READY:") for warning in warnings)
        if action == "BUY_SIMULATED":
            bought.add(ticker)
        if not watch_ready and action != "BUY_SIMULATED":
            continue
        detected_records.append(setup)
        unique_tickers.add(ticker)

        session = str(decision.get("market_session") or decision.get("confirmation_session") or "").lower()
        if "regular" in session or not session:
            regular_reviewed.add(ticker)
        if bool(decision.get("entry_confirmation_passed")):
            confirmation_passed.add(ticker)
        weighted_rr = to_float(decision.get("weighted_net_rr") or decision.get("decision_rr"))
        net_rr = to_float(decision.get("net_rr"))
        required = minimum_net_rr_for_decision(decision)
        if max(weighted_rr, net_rr) >= required:
            rr_passed.add(ticker)

    return {
        "detected_records": len(detected_records),
        "unique_detected": len(unique_tickers),
        "regular_reviewed_unique": len(regular_reviewed),
        "confirmation_passed_unique": len(confirmation_passed),
        "rr_passed_unique": len(rr_passed),
        "buy_simulated_unique": len(bought),
        "steps": [
            {
                "key": "detected",
                "label": "Detected",
                "value": len(unique_tickers),
                "detail": f"{len(detected_records)} WATCH_READY/BUY records",
            },
            {
                "key": "regular",
                "label": "Regular reviewed",
                "value": len(regular_reviewed),
                "detail": "Reviewed in or eligible for regular-session confirmation",
            },
            {
                "key": "confirmation",
                "label": "Confirmation passed",
                "value": len(confirmation_passed),
                "detail": "Completed-candle entry confirmation passed",
            },
            {
                "key": "rr",
                "label": "R/R passed",
                "value": len(rr_passed),
                "detail": "Weighted/net R/R passed the active threshold",
            },
            {
                "key": "buy",
                "label": "BUY",
                "value": len(bought),
                "detail": "Opened as paper trades by active gates",
            },
        ],
    }


def minimum_net_rr_for_decision(decision: dict[str, Any]) -> float:
    explicit = to_float(decision.get("minimum_net_rr_required") or decision.get("required_net_rr"))
    if explicit:
        return explicit
    regime = str(decision.get("market_regime") or "").upper()
    if regime == "NEUTRAL":
        return 2.5
    if regime == "BEAR":
        return 999.0
    return 2.0


def sorted_unique_diagnostic_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen = set()
    unique = []
    for item in sorted(
        items,
        key=lambda candidate: (
            candidate.get("action") == "WATCH_READY",
            to_float(candidate.get("setup_score")),
            to_float(candidate.get("weighted_net_rr") or candidate.get("net_rr")),
        ),
        reverse=True,
    ):
        key = (
            item.get("ticker"),
            item.get("action"),
            item.get("setup_type"),
            item.get("reason"),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(item)
    return unique


def count_trade_ready(setups: list[dict[str, Any]]) -> int:
    total = 0
    for setup in setups:
        action = str(setup.get("action") or (setup.get("decision_json") or {}).get("final_action") or "").upper()
        if action in {"BUY_SIMULATED", "WATCH_READY"}:
            total += 1
    return total


def build_system_health(
    *,
    tracker_path: Path,
    updates: list[dict[str, Any]],
    latest_update: dict[str, Any],
    latest_scan_update: dict[str, Any],
    latest_monitor_update: dict[str, Any],
) -> dict[str, Any]:
    now = datetime.utcnow().replace(microsecond=0)
    latest_scan_at = parse_timestamp(latest_scan_update.get("timestamp"))
    latest_monitor_at = parse_timestamp(latest_monitor_update.get("timestamp"))
    latest_update_at = parse_timestamp(latest_update.get("timestamp"))
    scan_age = age_minutes(now, latest_scan_at)
    monitor_age = age_minutes(now, latest_monitor_at)
    notes = []
    if not latest_scan_update:
        notes.append("No scanner update found in tracker.")
    elif scan_age is not None and scan_age > 36 * 60:
        notes.append("Latest scanner update is older than 36 hours.")
    try:
        tracker_mtime = datetime.utcfromtimestamp(tracker_path.stat().st_mtime).replace(microsecond=0)
    except OSError:
        tracker_mtime = datetime.min
        notes.append("Tracker file timestamp is unavailable.")

    return {
        "status": "ok" if not notes else "attention",
        "generated_at": now.isoformat() + "Z",
        "tracker_updated_at": tracker_mtime.isoformat() + "Z" if tracker_mtime != datetime.min else "",
        "latest_update_at": latest_update_at.isoformat() if latest_update_at != datetime.min else "",
        "latest_scan_at": latest_scan_at.isoformat() if latest_scan_at != datetime.min else "",
        "latest_scan_run_id": latest_scan_update.get("run_id", ""),
        "latest_scan_age_minutes": scan_age,
        "latest_monitor_at": latest_monitor_at.isoformat() if latest_monitor_at != datetime.min else "",
        "latest_monitor_run_id": latest_monitor_update.get("run_id", ""),
        "latest_monitor_age_minutes": monitor_age,
        "latest_monitor_policy": "Monitor action timestamp only changes after TP/SL portfolio updates.",
        "recent_runs_loaded": len(updates),
        "notes": notes,
    }


def age_minutes(now: datetime, timestamp: datetime) -> int | None:
    if timestamp == datetime.min:
        return None
    return max(0, int((now - timestamp).total_seconds() // 60))


def sanitize_dashboard_media_urls(value: Any, project_root: Path) -> Any:
    if isinstance(value, dict):
        return {
            key: sanitize_dashboard_media_urls(item, project_root)
            for key, item in value.items()
        }
    if isinstance(value, list):
        return [sanitize_dashboard_media_urls(item, project_root) for item in value]
    if not isinstance(value, str) or not value.startswith("/agent-results/"):
        return value
    relative = value.split("/agent-results/", 1)[1].split("?", 1)[0].split("#", 1)[0]
    return value if (project_root / "agent_results" / relative).exists() else ""


def read_settings(wb: Any) -> dict[str, Any]:
    ws = wb["Settings"]
    values = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            values[str(row[0])] = row[1]
    return values


def read_updates(wb: Any) -> list[dict[str, Any]]:
    rows = []
    ws = wb["Update Log"]
    for row in data_rows(ws):
        rows.append(
            {
                "timestamp": row[0],
                "run_id": row[1],
                "tickers": split_tickers(row[2]),
                "valid_setups": to_int(row[3]),
                "actions_summary": row[4] or "",
                "cash_ils": round(to_float(row[5]), 2),
                "exposure_ils": round(to_float(row[6]), 2),
                "open_risk_ils": round(to_float(row[7]), 2),
                "open_positions": to_int(row[8]),
                "summary": row[9] or "",
                "screenshot": row[10] or "",
                "decision_jsonl": cell(row, 11),
            }
        )
    return rows


def read_open_positions(wb: Any) -> list[dict[str, Any]]:
    rows = []
    ws = wb["Open Positions"]
    for row in data_rows(ws):
        entry = to_float(row[2])
        current = to_float(row[3], entry)
        target_1 = to_float(row[6])
        progress = 0.0
        if target_1 > entry:
            progress = max(0.0, min(100.0, (current - entry) / (target_1 - entry) * 100))
        rows.append(
            with_position_calculations(
                with_ticker_meta(
                    {
                        "ticker": row[0],
                        "entry_date": row[1],
                        "entry_price_usd": round(entry, 2),
                        "current_price_usd": round(current, 2),
                        "quantity": to_int(row[4]),
                        "stop_loss": round(to_float(row[5]), 2),
                        "target_1": round(target_1, 2),
                        "target_2": round(to_float(row[7]), 2),
                        "status": row[8] or "OPEN",
                        "unrealized_pnl_usd": round(to_float(row[9]), 2),
                        "unrealized_pnl_ils": round(to_float(row[10]), 2),
                        "exposure_ils": round(to_float(row[11]), 2),
                        "open_risk_ils": round(to_float(row[12]), 2),
                        "notes": row[13] or "",
                        "screenshot_url": resolve_asset_url(row[14]),
                        "chart_url": resolve_asset_url(cell(row, 15)),
                        "selection_context": cell(row, 16),
                        "decision_json": parse_json(cell(row, 17), {}),
                        "progress_to_target_1": round(progress, 2),
                    }
                )
            )
        )
    return rows


def read_trades(wb: Any) -> list[dict[str, Any]]:
    rows = []
    ws = wb["Trade Log"]
    for row in data_rows(ws):
        action = row[1]
        price = row[3] if action == "BUY_SIMULATED" else row[4]
        trade = with_ticker_meta(
            {
                "timestamp": row[0],
                "action": action,
                "ticker": row[2],
                "price_usd": round(to_float(price), 2) if price is not None else None,
                "entry_price_usd": round(to_float(row[3]), 2) if row[3] is not None else None,
                "exit_price_usd": round(to_float(row[4]), 2) if row[4] is not None else None,
                "quantity": to_int(row[5]),
                "usd_ils": to_float(row[6], 1.0),
                "buy_value_ils": round(to_float(row[7]), 2),
                "sell_value_ils": round(to_float(row[8]), 2),
                "cash_out_ils": round(to_float(row[9]), 2),
                "cash_in_ils": round(to_float(row[10]), 2),
                "stop_loss": round(to_float(row[11]), 2),
                "target_1": round(to_float(row[12]), 2),
                "target_2": round(to_float(row[13]), 2),
                "risk_ils": round(to_float(row[14]), 2),
                "reason": row[15] or "",
                "screenshot_url": resolve_asset_url(row[16]),
                "chart_url": resolve_asset_url(cell(row, 17)),
                "selection_context": cell(row, 18),
                "decision_json": parse_json(cell(row, 19), {}),
                "trade_id": cell(row, 20),
                "setup_score_bucket": cell(row, 21),
                "entry_confirmation_status": cell(row, 22),
                "mfe": cell(row, 23, None),
                "mae": cell(row, 24, None),
                "r_multiple": cell(row, 25, None),
                "duration": cell(row, 26),
                "exit_reason": cell(row, 27),
                "outcome_after_1d": cell(row, 28, None),
                "outcome_after_3d": cell(row, 29, None),
                "outcome_after_5d": cell(row, 30, None),
                "outcome_after_10d": cell(row, 31, None),
            }
        )
        rows.append(with_trade_potential(trade))
    return rows


def read_setup_rows(
    wb: Any,
    *,
    run_date: Any = None,
    cutoff: Any = None,
    recent_limit: int | None = None,
) -> list[dict[str, Any]]:
    raw_rows = []
    ws = wb["Setup Watchlist"]
    cutoff_time = parse_timestamp(cutoff) if cutoff else None
    for row in data_rows(ws):
        if run_date and row[0] != run_date:
            continue
        if cutoff_time and parse_timestamp(row[0]) > cutoff_time:
            continue
        raw_rows.append(row)
    if recent_limit and len(raw_rows) > recent_limit:
        raw_rows = raw_rows[-recent_limit:]

    rows = []
    for row in raw_rows:
        setup = with_ticker_meta(
            {
                "run_date": row[0],
                "ticker": row[1],
                "setup_type": row[2],
                "score": round(to_float(row[3]), 2),
                "current_price_usd": round(to_float(row[4]), 2),
                "buy_zone_low": round(to_float(row[5]), 2) if row[5] is not None else None,
                "buy_zone_high": round(to_float(row[6]), 2) if row[6] is not None else None,
                "stop_loss": round(to_float(row[7]), 2) if row[7] is not None else None,
                "target_1": round(to_float(row[8]), 2) if row[8] is not None else None,
                "target_2": round(to_float(row[9]), 2) if row[9] is not None else None,
                "risk_reward": round(to_float(row[10]), 2),
                "reason": row[11] or "",
                "action": row[12] or "",
                "feedback": row[13] or "",
                "chart_url": resolve_asset_url(cell(row, 15)),
                "selection_context": cell(row, 16),
                "decision_json": parse_json(cell(row, 17), {}),
            }
        )
        rows.append(with_setup_potential(setup))
    return rows


def read_decision_setup_rows(path: Path | None) -> list[dict[str, Any]]:
    if not path or not path.exists():
        return []

    rows = []
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except OSError:
        return []
    for line in lines:
        try:
            decision = json.loads(line)
        except (TypeError, json.JSONDecodeError):
            continue
        if not isinstance(decision, dict):
            continue
        setup = with_ticker_meta(
            {
                "run_date": decision.get("timestamp"),
                "ticker": decision.get("ticker"),
                "setup_type": decision.get("setup_type") or "No Trade",
                "score": round(to_float(decision.get("setup_score")), 2),
                "current_price_usd": round(to_float(decision.get("price")), 2),
                "buy_zone_low": decision.get("buy_zone_low"),
                "buy_zone_high": decision.get("buy_zone_high"),
                "stop_loss": decision.get("stop_loss"),
                "target_1": decision.get("target_1"),
                "target_2": decision.get("target_2"),
                "risk_reward": round(to_float(decision.get("net_rr")), 2),
                "reason": decision.get("reason") or "",
                "action": decision.get("final_action") or "",
                "feedback": decision.get("reason") or "",
                "chart_url": resolve_asset_url(decision.get("chart_url")),
                "selection_context": decision.get("scan_source") or "",
                "decision_json": decision,
            }
        )
        rows.append(with_setup_potential(setup))
    return rows


def select_update(updates: list[dict[str, Any]], selected_date: str | None) -> dict[str, Any]:
    if not updates:
        return {}
    if not selected_date:
        return updates[-1]

    end_of_day = parse_selected_date_end(selected_date)
    if not end_of_day:
        return updates[-1]

    selected = [update for update in updates if parse_timestamp(update.get("timestamp")) <= end_of_day]
    return selected[-1] if selected else {}


def select_latest_scan_update(updates: list[dict[str, Any]], latest_update: dict[str, Any]) -> dict[str, Any]:
    if not updates:
        return {}
    end_index = len(updates) - 1
    if latest_update in updates:
        end_index = updates.index(latest_update)
    for update in reversed(updates[: end_index + 1]):
        if not is_monitor_update(update):
            return update
    return {}


def select_latest_monitor_update(updates: list[dict[str, Any]], latest_update: dict[str, Any]) -> dict[str, Any]:
    if not updates:
        return {}
    end_index = len(updates) - 1
    if latest_update in updates:
        end_index = updates.index(latest_update)
    for update in reversed(updates[: end_index + 1]):
        if is_monitor_update(update):
            return update
    return {}


def is_monitor_update(update: dict[str, Any]) -> bool:
    return str(update.get("run_id") or "").startswith("monitor_")


def filter_records_until(records: list[dict[str, Any]], cutoff: Any, key: str = "timestamp") -> list[dict[str, Any]]:
    if not cutoff:
        return []
    cutoff_time = parse_timestamp(cutoff)
    return [record for record in records if parse_timestamp(record.get(key)) <= cutoff_time]


def available_dates(updates: list[dict[str, Any]]) -> list[str]:
    dates = []
    seen = set()
    for update in updates:
        parsed = parse_timestamp(update.get("timestamp"))
        text = parsed.date().isoformat()
        if text not in seen:
            seen.add(text)
            dates.append(text)
    return dates


def reconstruct_open_positions(
    trades: list[dict[str, Any]],
    setup_rows: list[dict[str, Any]],
    cutoff: Any,
) -> list[dict[str, Any]]:
    positions: dict[str, dict[str, Any]] = {}
    latest_setup = latest_setup_by_ticker(setup_rows, cutoff)

    for trade in trades:
        ticker = str(trade.get("ticker") or "")
        if not ticker:
            continue
        action = str(trade.get("action") or "")
        quantity = to_int(trade.get("quantity"))
        if quantity <= 0:
            continue

        if action == "BUY_SIMULATED":
            entry = to_float(trade.get("entry_price_usd"))
            current = to_float(latest_setup.get(ticker, {}).get("current_price_usd"), entry)
            positions[ticker] = with_ticker_meta(
                {
                    "ticker": ticker,
                    "entry_date": trade.get("timestamp"),
                    "entry_price_usd": round(entry, 2),
                    "current_price_usd": round(current, 2),
                    "quantity": quantity,
                    "stop_loss": round(to_float(trade.get("stop_loss")), 2),
                    "target_1": round(to_float(trade.get("target_1")), 2),
                    "target_2": round(to_float(trade.get("target_2")), 2),
                    "status": "OPEN",
                    "notes": trade.get("reason") or "",
                    "screenshot_url": trade.get("screenshot_url") or "",
                    "chart_url": trade.get("chart_url") or latest_setup.get(ticker, {}).get("chart_url") or "",
                    "selection_context": (
                        trade.get("selection_context")
                        or latest_setup.get(ticker, {}).get("selection_context")
                        or ""
                    ),
                    "decision_json": trade.get("decision_json") or latest_setup.get(ticker, {}).get("decision_json") or {},
                }
            )
            continue

        if action not in {"TAKE_PARTIAL_PROFIT", "TAKE_PROFIT", "EXIT_STOP"} or ticker not in positions:
            continue

        if action == "TAKE_PARTIAL_PROFIT":
            positions[ticker]["quantity"] = max(0, to_int(positions[ticker].get("quantity")) - quantity)
            positions[ticker]["stop_loss"] = positions[ticker]["entry_price_usd"]
            positions[ticker]["notes"] = "Partial profit taken; stop moved to breakeven."
            if positions[ticker]["quantity"] <= 0:
                positions.pop(ticker, None)
            continue

        positions.pop(ticker, None)

    rebuilt = []
    for ticker, position in positions.items():
        setup = latest_setup.get(ticker, {})
        if setup:
            position["current_price_usd"] = setup.get("current_price_usd") or position["current_price_usd"]
            position["chart_url"] = position.get("chart_url") or setup.get("chart_url") or ""
            position["selection_context"] = (
                position.get("selection_context") or setup.get("selection_context") or ""
            )
            position["decision_json"] = position.get("decision_json") or setup.get("decision_json") or {}
        rebuilt.append(with_position_calculations(position))
    return rebuilt


def latest_setup_by_ticker(setup_rows: list[dict[str, Any]], cutoff: Any) -> dict[str, dict[str, Any]]:
    if not cutoff:
        return {}
    cutoff_time = parse_timestamp(cutoff)
    latest: dict[str, dict[str, Any]] = {}
    for row in setup_rows:
        if parse_timestamp(row.get("run_date")) <= cutoff_time:
            latest[str(row.get("ticker") or "")] = row
    return latest


def with_ticker_meta(record: dict[str, Any]) -> dict[str, Any]:
    ticker = str(record.get("ticker") or "").upper()
    record["ticker"] = ticker
    record["sector"] = sector_map().get(ticker, "Unknown")
    record["company_name"] = company_name_for(ticker)
    return record


def sector_map() -> dict[str, str]:
    global _SECTOR_MAP
    if _SECTOR_MAP is None:
        _SECTOR_MAP = base_universe()
    return _SECTOR_MAP


def with_position_calculations(position: dict[str, Any]) -> dict[str, Any]:
    quantity = to_int(position.get("quantity"))
    current = to_float(position.get("current_price_usd"), position.get("entry_price_usd"))
    entry = to_float(position.get("entry_price_usd"), current)
    stop = to_float(position.get("stop_loss"))
    target_1 = to_float(position.get("target_1"))
    target_2 = to_float(position.get("target_2"))

    position["unrealized_pnl_usd"] = round((current - entry) * quantity, 2)
    position["unrealized_pnl_ils"] = position["unrealized_pnl_usd"]
    position["exposure_ils"] = round(current * quantity, 2)
    position["open_risk_ils"] = round(max(0.0, current - stop) * quantity, 2)
    position["potential_profit_t1_ils"] = round(max(0.0, target_1 - current) * quantity, 2)
    position["potential_profit_t2_ils"] = round(max(0.0, target_2 - current) * quantity, 2)
    position["potential_profit_plan_ils"] = weighted_target_profit(
        position["potential_profit_t1_ils"],
        position["potential_profit_t2_ils"],
        bool(target_1),
        bool(target_2),
    )
    position["reward_to_risk_plan"] = (
        round(position["potential_profit_plan_ils"] / position["open_risk_ils"], 2)
        if position["open_risk_ils"] > 0
        else 0
    )
    position["progress_to_target_1"] = progress_to_target(current, entry, target_1)
    position["partial_taken"] = "partial profit taken" in str(position.get("notes") or "").lower()
    position["position_attention"] = position_attention_status(position)
    return position


def with_trade_potential(trade: dict[str, Any]) -> dict[str, Any]:
    if trade.get("action") != "BUY_SIMULATED":
        trade["potential_profit_plan_ils"] = 0.0
        return trade

    quantity = to_int(trade.get("quantity"))
    entry = to_float(trade.get("entry_price_usd"))
    target_1 = to_float(trade.get("target_1"))
    target_2 = to_float(trade.get("target_2"))
    t1 = round(max(0.0, target_1 - entry) * quantity, 2)
    t2 = round(max(0.0, target_2 - entry) * quantity, 2)
    trade["potential_profit_t1_ils"] = t1
    trade["potential_profit_t2_ils"] = t2
    trade["potential_profit_plan_ils"] = weighted_target_profit(t1, t2, bool(target_1), bool(target_2))
    trade["reward_to_risk_plan"] = (
        round(trade["potential_profit_plan_ils"] / trade["risk_ils"], 2) if trade["risk_ils"] > 0 else 0
    )
    return trade


def with_setup_potential(setup: dict[str, Any]) -> dict[str, Any]:
    current = to_float(setup.get("current_price_usd"))
    target_1 = to_float(setup.get("target_1"))
    target_2 = to_float(setup.get("target_2"))
    t1 = round(max(0.0, target_1 - current), 2)
    t2 = round(max(0.0, target_2 - current), 2)
    setup["potential_profit_t1_per_share"] = t1
    setup["potential_profit_t2_per_share"] = t2
    setup["potential_profit_plan_per_share"] = weighted_target_profit(t1, t2, bool(target_1), bool(target_2))
    return setup


def weighted_target_profit(t1: float, t2: float, has_t1: bool, has_t2: bool) -> float:
    if has_t1 and has_t2:
        return round((t1 * 0.5) + (t2 * 0.5), 2)
    if has_t1:
        return round(t1, 2)
    if has_t2:
        return round(t2, 2)
    return 0.0


def progress_to_target(current: float, entry: float, target_1: float) -> float:
    if target_1 <= entry:
        return 0.0
    return round(max(0.0, min(100.0, (current - entry) / (target_1 - entry) * 100)), 2)


def build_position_attention(positions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates = []
    rank = {"immediate": 0, "high": 1, "medium": 2, "low": 3}
    for position in positions:
        attention = position.get("position_attention") or position_attention_status(position)
        if attention.get("level") == "low":
            continue
        candidates.append(
            {
                "ticker": position.get("ticker", ""),
                "company_name": position.get("company_name", ""),
                "sector": position.get("sector", ""),
                "current_price_usd": position.get("current_price_usd"),
                "entry_price_usd": position.get("entry_price_usd"),
                "chart_url": position.get("chart_url", ""),
                "attention": attention,
            }
        )
    candidates.sort(
        key=lambda item: (
            rank.get(str((item.get("attention") or {}).get("level") or "low"), 9),
            to_float((item.get("attention") or {}).get("distance_pct"), 999),
            str(item.get("ticker") or ""),
        )
    )
    return candidates[:8]


def build_risk_dashboard(
    positions: list[dict[str, Any]],
    summary: dict[str, Any],
    latest_decisions: list[dict[str, Any]],
) -> dict[str, Any]:
    exposure = to_float(summary.get("exposure_ils"))
    cash = to_float(summary.get("cash_ils"))
    open_risk = to_float(summary.get("open_risk_ils"))
    starting_capital = to_float(summary.get("starting_capital_ils"), 100_000.0)
    latest_decision = latest_decisions[0] if latest_decisions else {}
    market_regime = str(latest_decision.get("market_regime") or "UNKNOWN").upper()
    max_total_exposure = (
        to_float(latest_decision.get("dynamic_exposure_limit"))
        if "dynamic_exposure_limit" in latest_decision
        else market_exposure_limit(market_regime)
    )
    remaining_capacity = max(0.0, max_total_exposure - exposure)
    portfolio_heat_cap = to_float(
        latest_decision.get("portfolio_heat_cap"), starting_capital * 0.025
    )
    remaining_heat_capacity = max(0.0, portfolio_heat_cap - open_risk)

    sector_totals: dict[str, dict[str, Any]] = defaultdict(lambda: {"name": "", "exposure": 0.0, "count": 0})
    factor_totals: dict[str, dict[str, Any]] = defaultdict(lambda: {"name": "", "exposure": 0.0, "count": 0})
    for position in positions:
        position_exposure = to_float(position.get("exposure_ils"))
        sector = str(position.get("sector") or "Unknown")
        sector_totals[sector]["name"] = sector
        sector_totals[sector]["exposure"] += position_exposure
        sector_totals[sector]["count"] += 1

        tags = factor_tags_for_position(position)
        for tag in tags:
            factor_totals[tag]["name"] = tag
            factor_totals[tag]["exposure"] += position_exposure
            factor_totals[tag]["count"] += 1

    return {
        "market_regime": market_regime,
        "cash": round(cash, 2),
        "total_exposure": round(exposure, 2),
        "max_total_exposure": round(max_total_exposure, 2),
        "remaining_exposure_capacity": round(remaining_capacity, 2),
        "remaining_new_trade_budget": round(max(0.0, min(cash, remaining_capacity)), 2),
        "open_risk": round(open_risk, 2),
        "open_risk_pct": round(open_risk / starting_capital * 100, 2) if starting_capital else 0.0,
        "portfolio_heat_cap": round(portfolio_heat_cap, 2),
        "remaining_heat_capacity": round(remaining_heat_capacity, 2),
        "dynamic_exposure_enabled": bool(latest_decision.get("dynamic_exposure_enabled")),
        "market_regime_risk_points": latest_decision.get("market_regime_risk_points"),
        "sector_exposure": exposure_rows(sector_totals, exposure),
        "factor_exposure": exposure_rows(factor_totals, exposure),
    }


def market_exposure_limit(market_regime: str) -> float:
    if market_regime == "BULL":
        return 40_000.0
    if market_regime == "NEUTRAL":
        return 20_000.0
    if market_regime == "BEAR":
        return 0.0
    return 20_000.0


def factor_tags_for_position(position: dict[str, Any]) -> list[str]:
    decision = position.get("decision_json") if isinstance(position.get("decision_json"), dict) else {}
    raw_tags = decision.get("factor_tags") if isinstance(decision, dict) else []
    tags = [str(tag).strip() for tag in raw_tags or [] if str(tag).strip()]
    if tags:
        return tags
    sector = str(position.get("sector") or "").strip()
    return [sector] if sector else ["Unclassified"]


def exposure_rows(groups: dict[str, dict[str, Any]], total_exposure: float) -> list[dict[str, Any]]:
    rows = []
    for item in groups.values():
        exposure = to_float(item.get("exposure"))
        rows.append(
            {
                "name": item.get("name") or "Unknown",
                "exposure": round(exposure, 2),
                "pct_of_exposure": round(exposure / total_exposure * 100, 2) if total_exposure else 0.0,
                "count": to_int(item.get("count")),
            }
        )
    rows.sort(key=lambda row: (to_float(row.get("exposure")), str(row.get("name") or "")), reverse=True)
    return rows


def build_position_timeline(positions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    timelines = []
    for position in positions:
        current = to_float(position.get("current_price_usd"), position.get("entry_price_usd"))
        entry = to_float(position.get("entry_price_usd"), current)
        stop = to_float(position.get("stop_loss"))
        target_1 = to_float(position.get("target_1"))
        target_2 = to_float(position.get("target_2"))
        partial_taken = bool(position.get("partial_taken")) or "partial profit taken" in str(position.get("notes") or "").lower()
        breakeven_stop = entry > 0 and stop > 0 and abs(stop - entry) / entry <= 0.001
        next_attention = position.get("position_attention") or position_attention_status(position)

        steps = [
            {
                "label": "Entry",
                "level": round(entry, 2),
                "status": "complete",
                "detail": f"Opened {position.get('entry_date') or ''}".strip(),
            },
            {
                "label": "TP1 partial",
                "level": round(target_1, 2) if target_1 else None,
                "status": "complete" if partial_taken else "active" if next_attention.get("event") == "TAKE_PARTIAL_PROFIT" else "pending",
                "detail": "Take partial profit on 50% and reduce risk.",
            },
            {
                "label": "Stop to entry",
                "level": round(entry, 2) if entry else None,
                "status": "complete" if breakeven_stop else "active" if partial_taken else "pending",
                "detail": "After TP1, remaining stop should be protected at entry.",
            },
            {
                "label": "Current",
                "level": round(current, 2) if current else None,
                "status": "active",
                "detail": (next_attention.get("reason") or "Live/current tracker price."),
            },
            {
                "label": "TP2 / SL",
                "level": round(target_2, 2) if target_2 else None,
                "status": "pending",
                "detail": f"Stop {round(stop, 2) if stop else 'N/A'} / Target 2 {round(target_2, 2) if target_2 else 'N/A'}",
            },
        ]
        timelines.append(
            {
                "ticker": position.get("ticker", ""),
                "company_name": position.get("company_name", ""),
                "sector": position.get("sector", ""),
                "entry_price_usd": round(entry, 2),
                "current_price_usd": round(current, 2),
                "stop_loss": round(stop, 2) if stop else None,
                "target_1": round(target_1, 2) if target_1 else None,
                "target_2": round(target_2, 2) if target_2 else None,
                "partial_taken": partial_taken,
                "breakeven_stop": breakeven_stop,
                "chart_url": position.get("chart_url", ""),
                "steps": steps,
            }
        )
    return timelines


def position_attention_status(position: dict[str, Any]) -> dict[str, Any]:
    current = to_float(position.get("current_price_usd"), position.get("entry_price_usd"))
    entry = to_float(position.get("entry_price_usd"), current)
    stop = to_float(position.get("stop_loss"))
    target_1 = to_float(position.get("target_1"))
    target_2 = to_float(position.get("target_2"))
    partial_taken = "partial profit taken" in str(position.get("notes") or "").lower()

    if current <= 0:
        return {
            "level": "low",
            "event": "NO_PRICE",
            "label": "No live price",
            "distance_pct": None,
            "threshold": None,
            "reason": "Current price is unavailable.",
        }

    candidates: list[dict[str, Any]] = []
    if stop > 0:
        stop_label = "Breakeven stop" if entry > 0 and abs(stop - entry) / entry <= 0.001 else "Stop loss"
        if current <= stop:
            return attention_payload("immediate", "EXIT_STOP", stop_label, current, stop, "below")
        candidates.append(attention_payload("", "EXIT_STOP", stop_label, current, stop, "below"))

    if target_2 > 0:
        if current >= target_2:
            return attention_payload("immediate", "TAKE_PROFIT", "Target 2", current, target_2, "above")
        candidates.append(attention_payload("", "TAKE_PROFIT", "Target 2", current, target_2, "above"))

    if target_1 > 0 and not partial_taken:
        if current >= target_1:
            return attention_payload("immediate", "TAKE_PARTIAL_PROFIT", "Target 1", current, target_1, "above")
        candidates.append(attention_payload("", "TAKE_PARTIAL_PROFIT", "Target 1", current, target_1, "above"))

    if not candidates:
        return {
            "level": "low",
            "event": "NO_TARGETS",
            "label": "No active target",
            "distance_pct": None,
            "threshold": None,
            "reason": "No stop/target levels are available.",
        }

    nearest = min(candidates, key=lambda item: to_float(item.get("distance_pct"), 999))
    distance = to_float(nearest.get("distance_pct"), 999)
    if distance <= 0.5:
        nearest["level"] = "high"
    elif distance <= 2.0:
        nearest["level"] = "medium"
    else:
        nearest["level"] = "low"
    return nearest


def attention_payload(
    level: str,
    event: str,
    label: str,
    current: float,
    threshold: float,
    direction: str,
) -> dict[str, Any]:
    distance_pct = abs(current - threshold) / current * 100 if current else None
    if level == "immediate":
        reason = f"{label} is already touched or crossed."
    elif direction == "above":
        reason = f"{label} is {round(distance_pct or 0, 2)}% above current price."
    else:
        reason = f"{label} is {round(distance_pct or 0, 2)}% below current price."
    return {
        "level": level,
        "event": event,
        "label": label,
        "distance_pct": round(distance_pct, 2) if distance_pct is not None else None,
        "threshold": round(threshold, 2),
        "reason": reason,
    }


def compute_cash(trades: list[dict[str, Any]], starting_capital: float) -> float:
    cash = starting_capital
    for trade in trades:
        cash -= to_float(trade.get("cash_out_ils"))
        cash += to_float(trade.get("cash_in_ils"))
    return round(cash, 2)


def compute_realized_pnl(trades: list[dict[str, Any]]) -> dict[str, Any]:
    lots: dict[str, deque[dict[str, float]]] = defaultdict(deque)
    closed = []
    annotated = []
    total = 0.0
    wins = 0
    losses = 0

    for trade in trades:
        annotated_trade = dict(trade)
        ticker = str(trade.get("ticker") or "")
        quantity = to_int(trade.get("quantity"))
        action = str(trade.get("action") or "")
        if quantity <= 0 or not ticker:
            annotated.append(annotated_trade)
            continue

        if action == "BUY_SIMULATED":
            cost = to_float(trade.get("cash_out_ils"), trade.get("buy_value_ils"))
            lots[ticker].append(
                {
                    "quantity": quantity,
                    "unit_cost": cost / quantity if quantity else 0,
                    "entry_price": to_float(trade.get("entry_price_usd"), trade.get("price_usd")),
                    "stop_loss": to_float(trade.get("stop_loss")),
                }
            )
            annotated.append(annotated_trade)
            continue

        if action not in {"TAKE_PARTIAL_PROFIT", "TAKE_PROFIT", "EXIT_STOP"}:
            annotated.append(annotated_trade)
            continue

        remaining = quantity
        cost_basis = 0.0
        entry_value = 0.0
        stop_value = 0.0
        while remaining > 0 and lots[ticker]:
            lot = lots[ticker][0]
            used = min(remaining, int(lot["quantity"]))
            cost_basis += used * lot["unit_cost"]
            entry_value += used * lot.get("entry_price", 0)
            stop_value += used * lot.get("stop_loss", 0)
            lot["quantity"] -= used
            remaining -= used
            if lot["quantity"] <= 0:
                lots[ticker].popleft()

        matched_quantity = quantity - remaining
        cash_in = to_float(trade.get("cash_in_ils"), trade.get("sell_value_ils"))
        exit_price = to_float(trade.get("exit_price_usd"), trade.get("price_usd"))
        avg_entry = entry_value / matched_quantity if matched_quantity else to_float(trade.get("entry_price_usd"))
        avg_stop = stop_value / matched_quantity if matched_quantity else to_float(trade.get("stop_loss"))
        if cost_basis <= 0 and avg_entry and exit_price:
            cost_basis = avg_entry * quantity * to_float(trade.get("usd_ils"), 1.0)
        pnl = round(cash_in - cost_basis, 2)
        pnl_pct = round(pnl / cost_basis * 100, 2) if cost_basis else 0
        r_multiple = to_float(trade.get("r_multiple"))
        if not r_multiple and avg_entry > avg_stop > 0 and exit_price:
            r_multiple = round((exit_price - avg_entry) / (avg_entry - avg_stop), 4)
        annotated_trade.update(
            {
                "entry_price_usd": round(avg_entry, 2) if avg_entry else trade.get("entry_price_usd"),
                "price_usd": round(exit_price, 2) if exit_price else trade.get("price_usd"),
                "cost_basis_ils": round(cost_basis, 2),
                "pnl_ils": pnl,
                "pnl_pct": pnl_pct,
                "r_multiple": r_multiple,
            }
        )
        total += pnl
        wins += 1 if pnl > 0 else 0
        losses += 1 if pnl < 0 else 0
        closed.append(annotated_trade)
        annotated.append(annotated_trade)

    breakeven = len(closed) - wins - losses
    return {"total": total, "closed": closed, "wins": wins, "losses": losses, "breakeven": breakeven, "trades": annotated}


def compute_full_trade_performance(trades: list[dict[str, Any]]) -> dict[str, Any]:
    lots: dict[str, deque[dict[str, Any]]] = defaultdict(deque)
    closed: list[dict[str, Any]] = []
    wins = 0
    losses = 0
    breakeven = 0
    total_pnl = 0.0

    for trade in trades:
        ticker = str(trade.get("ticker") or "").upper().strip()
        action = str(trade.get("action") or "")
        quantity = to_int(trade.get("quantity"))
        if not ticker or quantity <= 0:
            continue

        if action == "BUY_SIMULATED":
            cash_out = to_float(trade.get("cash_out_ils"), trade.get("buy_value_ils"))
            entry_price = to_float(trade.get("entry_price_usd"), trade.get("price_usd"))
            usd_ils = to_float(trade.get("usd_ils"), 1.0)
            stop = to_float(trade.get("stop_loss"))
            decision = trade.get("decision_json") if isinstance(trade.get("decision_json"), dict) else {}
            risk_per_share = max(0.0, entry_price - stop) * usd_ils
            lot = {
                "trade_id": trade_identity(trade),
                "ticker": ticker,
                "entry_timestamp": trade.get("timestamp"),
                "entry_price_usd": entry_price,
                "stop_loss": stop,
                "initial_quantity": quantity,
                "remaining_quantity": quantity,
                "unit_cost_ils": cash_out / quantity if quantity else 0.0,
                "cost_basis_ils": cash_out,
                "initial_risk_ils": risk_per_share * quantity,
                "realized_pnl_ils": 0.0,
                "cash_in_ils": 0.0,
                "exit_events": [],
                "setup_type": trade.get("setup_type") or decision.get("setup_type", ""),
                "setup_score_bucket": trade.get("setup_score_bucket") or decision.get("setup_score_bucket", ""),
                "setup_score": decision.get("setup_score"),
                "market_regime": decision.get("market_regime", ""),
                "sector_regime": decision.get("sector_regime", ""),
                "sector": trade.get("sector") or decision.get("sector", ""),
                "net_rr_1": decision.get("net_rr_1"),
                "net_rr_2": decision.get("net_rr_2"),
                "weighted_net_rr": decision.get("weighted_net_rr") or decision.get("net_rr"),
                "entry_confirmation_status": decision.get("entry_confirmation_status")
                or decision.get("confirmation_status"),
                "mfe": decision.get("mfe"),
                "mae": decision.get("mae"),
                "mfe_r": decision.get("mfe_r"),
                "mae_r": decision.get("mae_r"),
            }
            lots[ticker].append(lot)
            continue

        if action not in {"TAKE_PARTIAL_PROFIT", "TAKE_PROFIT", "EXIT_STOP"}:
            continue

        remaining = quantity
        cash_in = to_float(trade.get("cash_in_ils"), trade.get("sell_value_ils"))
        exit_price = to_float(trade.get("exit_price_usd"), trade.get("price_usd"))
        timestamp = trade.get("timestamp")
        while remaining > 0 and lots[ticker]:
            lot = lots[ticker][0]
            used = min(remaining, to_int(lot.get("remaining_quantity")))
            if used <= 0:
                lots[ticker].popleft()
                continue
            ratio = used / quantity if quantity else 0.0
            allocated_cash_in = cash_in * ratio if cash_in else exit_price * used * to_float(trade.get("usd_ils"), 1.0)
            cost_basis = used * to_float(lot.get("unit_cost_ils"))
            pnl = round(allocated_cash_in - cost_basis, 2)
            lot["realized_pnl_ils"] = round(to_float(lot.get("realized_pnl_ils")) + pnl, 2)
            lot["cash_in_ils"] = round(to_float(lot.get("cash_in_ils")) + allocated_cash_in, 2)
            lot["remaining_quantity"] = to_int(lot.get("remaining_quantity")) - used
            lot["exit_events"].append(
                {
                    "timestamp": timestamp,
                    "action": action,
                    "quantity": used,
                    "price_usd": round(exit_price, 2) if exit_price else None,
                    "pnl_ils": pnl,
                }
            )
            exit_decision = trade.get("decision_json") if isinstance(trade.get("decision_json"), dict) else {}
            for metric in ("mfe", "mae", "mfe_r", "mae_r"):
                if exit_decision.get(metric) not in (None, ""):
                    lot[metric] = exit_decision.get(metric)
            remaining -= used

            if to_int(lot.get("remaining_quantity")) <= 0:
                completed = completed_full_trade(lot, timestamp)
                closed.append(completed)
                total_pnl = round(total_pnl + to_float(completed.get("pnl_ils")), 2)
                if to_float(completed.get("pnl_ils")) > 0:
                    wins += 1
                elif to_float(completed.get("pnl_ils")) < 0:
                    losses += 1
                else:
                    breakeven += 1
                lots[ticker].popleft()

    open_lots = [lot for ticker_lots in lots.values() for lot in ticker_lots if to_int(lot.get("remaining_quantity")) > 0]
    closed_count = len(closed)
    return {
        "closed_count": closed_count,
        "open_count": len(open_lots),
        "wins": wins,
        "losses": losses,
        "breakeven": breakeven,
        "win_rate": round(wins / closed_count * 100, 2) if closed_count else 0,
        "total_pnl_ils": total_pnl,
        "closed": closed,
    }


def trade_identity(trade: dict[str, Any]) -> str:
    explicit = str(trade.get("trade_id") or "").strip()
    if explicit:
        return explicit
    return "|".join(
        [
            str(trade.get("ticker") or "").upper(),
            str(trade.get("timestamp") or ""),
            str(trade.get("entry_price_usd") or trade.get("price_usd") or ""),
            str(trade.get("quantity") or ""),
        ]
    )


def completed_full_trade(lot: dict[str, Any], exit_timestamp: Any) -> dict[str, Any]:
    pnl = round(to_float(lot.get("realized_pnl_ils")), 2)
    cost_basis = to_float(lot.get("cost_basis_ils"))
    initial_risk = to_float(lot.get("initial_risk_ils"))
    exit_events = list(lot.get("exit_events") or [])
    return {
        "trade_id": lot.get("trade_id", ""),
        "ticker": lot.get("ticker", ""),
        "entry_timestamp": lot.get("entry_timestamp"),
        "exit_timestamp": exit_timestamp,
        "entry_price_usd": round(to_float(lot.get("entry_price_usd")), 2),
        "stop_loss": round(to_float(lot.get("stop_loss")), 2),
        "initial_quantity": to_int(lot.get("initial_quantity")),
        "pnl_ils": pnl,
        "pnl_pct": round(pnl / cost_basis * 100, 2) if cost_basis else 0,
        "r_multiple": round(pnl / initial_risk, 4) if initial_risk else None,
        "result": "WIN" if pnl > 0 else "LOSS" if pnl < 0 else "BREAKEVEN",
        "exit_count": len(exit_events),
        "exit_actions": [str(event.get("action") or "") for event in exit_events],
        "exit_events": exit_events,
        "setup_type": lot.get("setup_type", ""),
        "setup_score": lot.get("setup_score"),
        "setup_score_bucket": lot.get("setup_score_bucket", ""),
        "sector": lot.get("sector", ""),
        "market_regime": lot.get("market_regime", ""),
        "sector_regime": lot.get("sector_regime", ""),
        "net_rr_1": lot.get("net_rr_1"),
        "net_rr_2": lot.get("net_rr_2"),
        "weighted_net_rr": lot.get("weighted_net_rr"),
        "entry_confirmation_status": lot.get("entry_confirmation_status"),
        "mfe": lot.get("mfe"),
        "mae": lot.get("mae"),
        "mfe_r": lot.get("mfe_r"),
        "mae_r": lot.get("mae_r"),
    }


def build_score_calibration(closed_trades: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets = [
        ("0.30-0.40", 0.30, 0.40),
        ("0.40-0.50", 0.40, 0.50),
        ("0.50-0.60", 0.50, 0.60),
        ("0.60-0.70", 0.60, 0.70),
        ("0.70+", 0.70, 9.99),
    ]
    stats: dict[str, dict[str, Any]] = {
        label: {"bucket": label, "trades": 0, "wins": 0, "losses": 0, "pnl_ils": 0.0}
        for label, _, _ in buckets
    }
    for trade in closed_trades:
        decision = trade.get("decision_json") or {}
        score = to_float(decision.get("setup_score"), trade.get("score"))
        label = next((name for name, low, high in buckets if low <= score < high), "0.30-0.40")
        pnl = to_float(trade.get("pnl_ils"))
        stats[label]["trades"] += 1
        stats[label]["wins"] += 1 if pnl > 0 else 0
        stats[label]["losses"] += 1 if pnl < 0 else 0
        stats[label]["pnl_ils"] += pnl

    rows = []
    for label, _, _ in buckets:
        item = stats[label]
        trades = item["trades"]
        rows.append(
            {
                **item,
                "pnl_ils": round(item["pnl_ils"], 2),
                "win_rate": round(item["wins"] / trades * 100, 2) if trades else 0,
                "avg_pnl_ils": round(item["pnl_ils"] / trades, 2) if trades else 0,
            }
        )
    return rows


def build_equity_curve(updates: list[dict[str, Any]], starting_capital: float) -> list[dict[str, Any]]:
    curve = [{"timestamp": "Start", "equity_ils": round(starting_capital, 2), "pnl_ils": 0.0}]
    for update in updates:
        equity = to_float(update.get("cash_ils")) + to_float(update.get("exposure_ils"))
        curve.append(
            {
                "timestamp": update.get("timestamp"),
                "equity_ils": round(equity, 2),
                "pnl_ils": round(equity - starting_capital, 2),
            }
        )
    return curve


def cell(row: tuple[Any, ...], index: int, default: Any = "") -> Any:
    if index >= len(row):
        return default
    value = row[index]
    return default if value is None else value


def parse_json(value: Any, default: Any) -> Any:
    if not value:
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except (TypeError, ValueError, json.JSONDecodeError):
        return default


def trim_text(value: str, max_chars: int) -> str:
    if len(value) <= max_chars:
        return value
    return value[:max_chars].rstrip() + "\n\n[Summary trimmed for dashboard load speed.]"


def data_rows(ws: Any) -> list[tuple[Any, ...]]:
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(value is not None for value in row):
            rows.append(row)
    return rows


def parse_selected_date_end(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.strptime(value[:10], "%Y-%m-%d").date()
    except ValueError:
        return None
    return datetime.combine(parsed, time.max)


def parse_timestamp(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if not value:
        return datetime.min
    text = str(value).strip()
    if not text:
        return datetime.min
    normalized = text.replace("Z", "").replace(" ", "T")
    try:
        return datetime.fromisoformat(normalized).replace(tzinfo=None)
    except ValueError:
        try:
            return datetime.strptime(text[:19], "%Y-%m-%dT%H:%M:%S")
        except ValueError:
            return datetime.min


def resolve_latest_file(directory: Path, suffix: str) -> Path | None:
    if not directory.exists():
        return None
    files = sorted(
        [path for path in directory.iterdir() if path.is_file() and path.suffix.lower() == suffix],
        key=lambda path: path.stat().st_mtime,
    )
    return files[-1] if files else None


def resolve_record_file(value: Any, directory: Path, suffix: str) -> Path | None:
    if not value:
        return None
    text = str(value).replace("\\", "/")
    path = Path(text)
    if path.exists() and path.suffix.lower() == suffix:
        return path
    candidate = directory / path.name
    if candidate.exists() and candidate.suffix.lower() == suffix:
        return candidate
    return None


def resolve_asset_url(value: Any) -> str:
    if not value:
        return ""
    text = str(value).replace("\\", "/")
    marker = "agent_results/"
    if marker in text:
        return "/agent-results/" + text.split(marker, 1)[1]
    return text


def split_tickers(value: Any) -> list[str]:
    return [ticker for ticker in str(value or "").replace(",", " ").split() if ticker]


def to_float(value: Any, default: Any = 0.0) -> float:
    try:
        if value is None or value == "":
            return float(default or 0.0)
        return float(value)
    except (TypeError, ValueError):
        return float(default or 0.0)


def to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default
