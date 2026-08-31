from openpyxl import Workbook, load_workbook

from app.workbook_retention import compact_setup_watchlist, enforce_tracker_size


def test_compact_setup_watchlist_keeps_header_and_latest_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Setup Watchlist"
    worksheet.append(["timestamp", "ticker"])
    for index in range(1, 8):
        worksheet.append([f"2026-08-{index:02d}", f"T{index}"])

    result = compact_setup_watchlist(workbook, max_rows=3)

    assert result == {
        "enabled": True,
        "rows_before": 7,
        "rows_after": 3,
        "rows_removed": 4,
        "max_rows": 3,
    }
    assert [row[1].value for row in worksheet.iter_rows(min_row=2)] == ["T5", "T6", "T7"]


def test_compact_setup_watchlist_can_be_disabled() -> None:
    workbook = Workbook()
    workbook.active.title = "Setup Watchlist"
    workbook["Setup Watchlist"].append(["timestamp"])
    workbook["Setup Watchlist"].append(["2026-08-01"])

    result = compact_setup_watchlist(workbook, max_rows=0)

    assert result["enabled"] is False
    assert workbook["Setup Watchlist"].max_row == 2


def test_enforce_tracker_size_rewrites_and_keeps_latest_rows(tmp_path) -> None:
    tracker = tmp_path / "tracker.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Setup Watchlist"
    worksheet.append(["timestamp", "ticker"])
    for index in range(1, 8):
        worksheet.append([f"2026-08-{index:02d}", f"T{index}"])
    workbook.save(tracker)
    workbook.close()

    result = enforce_tracker_size(
        tracker,
        rewrite_bytes=1,
        hard_limit_bytes=1_000_000,
        max_rows=3,
    )

    assert result["rewritten"] is True
    loaded = load_workbook(tracker, read_only=True)
    try:
        assert [row[1].value for row in loaded["Setup Watchlist"].iter_rows(min_row=2)] == [
            "T5",
            "T6",
            "T7",
        ]
    finally:
        loaded.close()
