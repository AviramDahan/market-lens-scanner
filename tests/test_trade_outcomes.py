from __future__ import annotations

import json
from datetime import date

import pandas as pd
from openpyxl import Workbook

from app.trade_outcomes import backfill_trade_outcomes


def trade_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Trade Log"
    sheet.append([f"Column {index}" for index in range(1, 33)])
    entry_decision = {
        "trade_id": "AAA-1",
        "setup_type": "Breakout + Retest",
        "setup_score_bucket": "0.50-0.59",
        "stop_loss": 95,
    }
    sheet.append(
        [
            "2026-08-03T14:00:00+00:00",
            "BUY_SIMULATED",
            "AAA",
            100,
            None,
            10,
            1,
            1000,
            0,
            1000,
            0,
            95,
            110,
            120,
            50,
            "buy",
            "",
            "",
            "",
            json.dumps(entry_decision),
            "AAA-1",
            "0.50-0.59",
            "PASSED",
        ]
    )
    sheet.append(
        [
            "2026-08-04T14:00:00+00:00",
            "TAKE_PARTIAL_PROFIT",
            "AAA",
            None,
            110,
            5,
            1,
            0,
            550,
            0,
            550,
            95,
            110,
            120,
            0,
            "partial",
            "",
            "",
            "",
            json.dumps(entry_decision),
        ]
    )
    sheet.append(
        [
            "2026-08-05T14:00:00+00:00",
            "EXIT_STOP",
            "AAA",
            None,
            100,
            5,
            1,
            0,
            500,
            0,
            500,
            100,
            110,
            120,
            0,
            "breakeven exit",
            "",
            "",
            "",
            json.dumps(entry_decision),
        ]
    )
    return workbook


def daily_frame() -> pd.DataFrame:
    index = pd.to_datetime(
        [
            "2026-08-03T00:00:00Z",
            "2026-08-04T00:00:00Z",
            "2026-08-05T00:00:00Z",
            "2026-08-06T00:00:00Z",
            "2026-08-07T00:00:00Z",
            "2026-08-10T00:00:00Z",
            "2026-08-11T00:00:00Z",
            "2026-08-12T00:00:00Z",
            "2026-08-13T00:00:00Z",
            "2026-08-14T00:00:00Z",
            "2026-08-17T00:00:00Z",
            "2026-08-18T00:00:00Z",
            "2026-08-19T00:00:00Z",
        ]
    )
    return pd.DataFrame(
        {
            "High": [102, 112, 108, 104, 106, 107, 108, 109, 110, 111, 112, 113, 114],
            "Low": [98, 99, 97, 98, 99, 100, 101, 102, 103, 104, 105, 106, 107],
            "Close": [101, 110, 100, 102, 103, 104, 105, 106, 107, 108, 109, 110, 111],
        },
        index=index,
    )


def test_backfill_reconstructs_full_trade_and_future_outcomes() -> None:
    workbook = trade_workbook()
    calls: list[str] = []

    def fetcher(ticker: str, *, period: str) -> pd.DataFrame:
        calls.append(f"{ticker}:{period}")
        return daily_frame()

    result = backfill_trade_outcomes(
        workbook,
        max_tickers=1,
        frame_fetcher=fetcher,
        as_of_date=date(2026, 8, 31),
    )
    sheet = workbook["Trade Log"]
    exit_decision = json.loads(sheet.cell(4, 20).value)

    assert calls == ["AAA:6mo"]
    assert result["errors"] == []
    assert result["checked_today"] == 1
    assert [sheet.cell(row, 21).value for row in (2, 3, 4)] == ["AAA-1", "AAA-1", "AAA-1"]
    assert sheet.cell(4, 24).value == 120
    assert sheet.cell(4, 25).value == 30
    assert sheet.cell(4, 26).value == 1
    assert sheet.cell(4, 29).value == 2
    assert sheet.cell(4, 31).value == 6
    assert exit_decision["full_trade_pnl"] == 50
    assert exit_decision["full_trade_r_multiple"] == 1
    assert exit_decision["analytics_source"] == "daily_bar_approximation"


def test_backfill_does_not_overwrite_live_excursion_metrics() -> None:
    workbook = trade_workbook()
    sheet = workbook["Trade Log"]
    existing = json.loads(sheet.cell(4, 20).value)
    existing.update({"mfe": 125, "mae": 35, "mfe_r": 2.5, "mae_r": 0.7, "excursion_source": "intraday_bars"})
    sheet.cell(4, 20, json.dumps(existing))
    sheet.cell(4, 24, 125)
    sheet.cell(4, 25, 35)

    backfill_trade_outcomes(
        workbook,
        max_tickers=1,
        frame_fetcher=lambda *_args, **_kwargs: daily_frame(),
        as_of_date=date(2026, 8, 31),
    )
    updated = json.loads(sheet.cell(4, 20).value)

    assert sheet.cell(4, 24).value == 125
    assert sheet.cell(4, 25).value == 35
    assert updated["mfe"] == 125
    assert updated["mae"] == 35
    assert updated["mfe_r"] == 2.5
    assert updated["excursion_source"] == "intraday_bars"


def test_zero_ticker_limit_skips_network_but_backfills_trade_identity() -> None:
    workbook = trade_workbook()
    sheet = workbook["Trade Log"]
    sheet.cell(3, 21, "")
    sheet.cell(4, 21, "")

    result = backfill_trade_outcomes(
        workbook,
        max_tickers=0,
        frame_fetcher=lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("must not fetch")),
        as_of_date=date(2026, 8, 31),
    )

    assert result["fetched_tickers"] == []
    assert sheet.cell(3, 21).value == "AAA-1"
    assert sheet.cell(4, 21).value == "AAA-1"


def test_backfill_checks_incomplete_trade_at_most_once_per_day() -> None:
    workbook = trade_workbook()
    short_frame = daily_frame().iloc[:4]
    calls = 0

    def fetcher(*_args, **_kwargs):
        nonlocal calls
        calls += 1
        return short_frame

    first = backfill_trade_outcomes(
        workbook,
        max_tickers=8,
        frame_fetcher=fetcher,
        as_of_date=date(2026, 8, 31),
    )
    second = backfill_trade_outcomes(
        workbook,
        max_tickers=8,
        frame_fetcher=fetcher,
        as_of_date=date(2026, 8, 31),
    )

    assert calls == 1
    assert first["checked_today"] == 1
    assert second["checked_today"] == 1
    assert second["fetched_tickers"] == []
    assert second["pending_trades"] == 0
