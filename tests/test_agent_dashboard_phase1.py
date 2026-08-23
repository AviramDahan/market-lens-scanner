from __future__ import annotations

from app.agent_dashboard import (
    build_position_attention,
    build_decision_diagnostics,
    build_position_timeline,
    build_risk_dashboard,
    compact_agent_dashboard_payload,
    compute_full_trade_performance,
    compute_realized_pnl,
    dashboard_section_payload,
    historical_tracker_copy,
    read_decision_setup_rows,
    with_position_calculations,
    write_diagnostic_snapshot,
)


def test_compute_realized_pnl_annotates_exit_trade_with_entry_and_r() -> None:
    trades = [
        {
            "timestamp": "2026-08-05T14:00:00",
            "action": "BUY_SIMULATED",
            "ticker": "AAA",
            "entry_price_usd": 100,
            "price_usd": 100,
            "quantity": 10,
            "cash_out_ils": 1000,
            "buy_value_ils": 1000,
            "cash_in_ils": 0,
            "stop_loss": 95,
        },
        {
            "timestamp": "2026-08-06T15:00:00",
            "action": "TAKE_PROFIT",
            "ticker": "AAA",
            "exit_price_usd": 112,
            "price_usd": 112,
            "quantity": 10,
            "cash_out_ils": 0,
            "cash_in_ils": 1120,
            "sell_value_ils": 1120,
            "stop_loss": 95,
        },
    ]

    result = compute_realized_pnl(trades)

    assert result["total"] == 120
    assert result["wins"] == 1
    assert result["losses"] == 0
    assert result["closed"][0]["entry_price_usd"] == 100
    assert result["closed"][0]["pnl_ils"] == 120
    assert result["closed"][0]["pnl_pct"] == 12
    assert result["closed"][0]["r_multiple"] == 2.4
    assert result["trades"][1]["pnl_ils"] == 120


def test_full_trade_winrate_counts_partial_then_breakeven_as_one_win() -> None:
    trades = [
        {
            "timestamp": "2026-08-05T14:00:00",
            "action": "BUY_SIMULATED",
            "ticker": "AAA",
            "entry_price_usd": 100,
            "price_usd": 100,
            "quantity": 10,
            "cash_out_ils": 1000,
            "buy_value_ils": 1000,
            "cash_in_ils": 0,
            "stop_loss": 95,
        },
        {
            "timestamp": "2026-08-05T15:00:00",
            "action": "TAKE_PARTIAL_PROFIT",
            "ticker": "AAA",
            "exit_price_usd": 110,
            "price_usd": 110,
            "quantity": 5,
            "cash_out_ils": 0,
            "cash_in_ils": 550,
            "sell_value_ils": 550,
            "stop_loss": 95,
        },
        {
            "timestamp": "2026-08-06T15:00:00",
            "action": "EXIT_STOP",
            "ticker": "AAA",
            "exit_price_usd": 100,
            "price_usd": 100,
            "quantity": 5,
            "cash_out_ils": 0,
            "cash_in_ils": 500,
            "sell_value_ils": 500,
            "stop_loss": 100,
        },
    ]

    exit_events = compute_realized_pnl(trades)
    full_trades = compute_full_trade_performance(trades)

    assert exit_events["wins"] == 1
    assert exit_events["losses"] == 0
    assert exit_events["breakeven"] == 1
    assert len(exit_events["closed"]) == 2
    assert full_trades["closed_count"] == 1
    assert full_trades["wins"] == 1
    assert full_trades["losses"] == 0
    assert full_trades["breakeven"] == 0
    assert full_trades["closed"][0]["pnl_ils"] == 50
    assert full_trades["closed"][0]["exit_actions"] == ["TAKE_PARTIAL_PROFIT", "EXIT_STOP"]


def test_full_trade_winrate_counts_stop_loss_as_money_loss() -> None:
    trades = [
        {
            "timestamp": "2026-08-05T14:00:00",
            "action": "BUY_SIMULATED",
            "ticker": "AAA",
            "entry_price_usd": 100,
            "price_usd": 100,
            "quantity": 10,
            "cash_out_ils": 1000,
            "buy_value_ils": 1000,
            "cash_in_ils": 0,
            "stop_loss": 95,
        },
        {
            "timestamp": "2026-08-06T15:00:00",
            "action": "EXIT_STOP",
            "ticker": "AAA",
            "exit_price_usd": 95,
            "price_usd": 95,
            "quantity": 10,
            "cash_out_ils": 0,
            "cash_in_ils": 950,
            "sell_value_ils": 950,
            "stop_loss": 95,
        },
    ]

    full_trades = compute_full_trade_performance(trades)

    assert full_trades["closed_count"] == 1
    assert full_trades["wins"] == 0
    assert full_trades["losses"] == 1
    assert full_trades["closed"][0]["result"] == "LOSS"
    assert full_trades["closed"][0]["pnl_ils"] == -50


def test_full_trade_winrate_ignores_still_open_partial_position() -> None:
    trades = [
        {
            "timestamp": "2026-08-05T14:00:00",
            "action": "BUY_SIMULATED",
            "ticker": "AAA",
            "entry_price_usd": 100,
            "price_usd": 100,
            "quantity": 10,
            "cash_out_ils": 1000,
            "buy_value_ils": 1000,
            "cash_in_ils": 0,
            "stop_loss": 95,
        },
        {
            "timestamp": "2026-08-05T15:00:00",
            "action": "TAKE_PARTIAL_PROFIT",
            "ticker": "AAA",
            "exit_price_usd": 110,
            "price_usd": 110,
            "quantity": 5,
            "cash_out_ils": 0,
            "cash_in_ils": 550,
            "sell_value_ils": 550,
            "stop_loss": 95,
        },
    ]

    full_trades = compute_full_trade_performance(trades)

    assert full_trades["closed_count"] == 0
    assert full_trades["open_count"] == 1
    assert full_trades["wins"] == 0
    assert full_trades["losses"] == 0


def test_compact_dashboard_payload_trims_heavy_collections_and_keeps_totals() -> None:
    dashboard = {
        "status": "ok",
        "latest_run": {"summary_text": "x" * 30_000},
        "latest_setups": [{"ticker": f"A{index}"} for index in range(25)],
        "latest_decisions": [{"ticker": "heavy"}],
        "recent_trades": [{"ticker": f"T{index}"} for index in range(18)],
        "closed_trades": [{"ticker": f"C{index}"} for index in range(7)],
        "recent_runs": [{"run": index} for index in range(20)],
        "equity_curve": [{"run": index} for index in range(500)],
        "full_trade_performance": {"closed": [1, 2, 3]},
        "decision_diagnostics": {
            "watch_ready_count": 6,
            "drilldowns": {
                "WATCH_READY": [{"ticker": f"W{index}"} for index in range(6)],
                "RR_BLOCKED": [{"ticker": "HEAVY"}],
            },
        },
        "daily_summary": {
            "BUY_SIMULATED_count": 1,
            "WATCH_READY_count": 6,
            "heavy_records": list(range(1_000)),
        },
    }

    compact = compact_agent_dashboard_payload(dashboard, action_limit=10, trade_limit=5)

    assert compact["payload"]["mode"] == "compact"
    assert len(compact["latest_setups"]) == 10
    assert compact["latest_decisions"] == []
    assert len(compact["recent_trades"]) == 5
    assert compact["recent_trades"][0]["ticker"] == "T17"
    assert compact["closed_trades"] == []
    assert compact["recent_runs"] == []
    assert compact["full_trade_performance"] == {}
    assert len(compact["equity_curve"]) == 240
    assert compact["equity_curve"][0] == {"run": 0}
    assert compact["equity_curve"][-1] == {"run": 499}
    assert len(compact["decision_diagnostics"]["drilldowns"]["WATCH_READY"]) == 4
    assert "RR_BLOCKED" not in compact["decision_diagnostics"]["drilldowns"]
    assert compact["daily_summary"] == {"BUY_SIMULATED_count": 1, "WATCH_READY_count": 6}
    assert compact["pagination"]["actions"]["total"] == 25
    assert compact["pagination"]["actions"]["has_more"] is True
    assert compact["pagination"]["trades"]["closed_total"] == 7
    assert "trimmed" in compact["latest_run"]["summary_text"].lower()
    assert len(dashboard["latest_run"]["summary_text"]) == 30_000


def test_historical_tracker_copy_removes_watchlist_rows_and_keeps_other_sheets(tmp_path) -> None:
    from openpyxl import Workbook, load_workbook

    tracker = tmp_path / "tracker.xlsx"
    workbook = Workbook()
    workbook.active.title = "Dashboard"
    workbook.create_sheet("Setup Watchlist").append(["large", "row"])
    workbook.create_sheet("Trade Log").append(["trade"])
    workbook.create_sheet("Open Positions")
    workbook.create_sheet("Update Log").append(["update"])
    workbook.create_sheet("Settings")
    workbook.create_sheet("Sources")
    workbook.create_sheet("Agent Control")
    workbook.create_sheet("Position Events")
    workbook.save(tracker)

    compact_tracker = historical_tracker_copy(tracker)
    loaded = load_workbook(compact_tracker, data_only=True, read_only=True)

    assert compact_tracker != tracker
    assert loaded["Setup Watchlist"]["A1"].value is None
    assert loaded["Update Log"]["A1"].value == "update"


def test_read_decision_setup_rows_recovers_historical_scan(tmp_path) -> None:
    path = tmp_path / "decisions.jsonl"
    path.write_text(
        "\n".join(
            [
                "not-json",
                '{"timestamp":"2026-08-21T15:00:00","ticker":"MSFT","setup_type":"Breakout + Retest",'
                '"setup_score":0.58,"price":421.5,"buy_zone_low":418,"buy_zone_high":422,'
                '"stop_loss":412,"target_1":435,"target_2":448,"net_rr":2.3,'
                '"final_action":"WATCH_READY","reason":"Waiting for confirmation"}',
            ]
        ),
        encoding="utf-8",
    )

    rows = read_decision_setup_rows(path)

    assert len(rows) == 1
    assert rows[0]["ticker"] == "MSFT"
    assert rows[0]["action"] == "WATCH_READY"
    assert rows[0]["decision_json"]["net_rr"] == 2.3


def test_dashboard_section_payload_paginates_actions_and_trades() -> None:
    dashboard = {
        "status": "ok",
        "latest_setups": [{"ticker": f"A{index}"} for index in range(15)],
        "recent_trades": [{"ticker": f"T{index}"} for index in range(12)],
    }

    actions = dashboard_section_payload(dashboard, section="actions", offset=10, limit=10)
    trades = dashboard_section_payload(dashboard, section="trades", offset=0, limit=3)

    assert [item["ticker"] for item in actions["items"]] == ["A10", "A11", "A12", "A13", "A14"]
    assert actions["has_more"] is False
    assert [item["ticker"] for item in trades["items"]] == ["T11", "T10", "T9"]
    assert trades["has_more"] is True


def test_decision_diagnostics_includes_drilldown_items_with_charts() -> None:
    diagnostics = build_decision_diagnostics(
        [
            {
                "ticker": "AAA",
                "company_name": "Alpha Apps",
                "sector": "Technology",
                "action": "WATCH_READY",
                "setup_type": "VWAP Reclaim",
                "score": 0.61,
                "current_price_usd": 101,
                "buy_zone_low": 100,
                "buy_zone_high": 102,
                "stop_loss": 96,
                "target_1": 108,
                "target_2": 115,
                "chart_url": "/agent-results/charts/aaa.png",
                "reason": "WATCH_READY: entry confirmation missing; waiting for regular-session confirmation",
                "decision_json": {
                    "market_regime": "NEUTRAL",
                    "sector_regime": "STRONG",
                    "setup_score": 0.61,
                    "minimum_net_rr_required": 2.0,
                    "weighted_net_rr": 2.22,
                    "net_rr_1": 1.1,
                    "net_rr_2": 3.0,
                    "entry_confirmation_passed": False,
                    "warnings": ["WATCH_READY: staged for confirmation"],
                },
            }
        ]
    )

    watch_ready = diagnostics["drilldowns"]["WATCH_READY"][0]
    assert diagnostics["watch_ready_count"] == 1
    assert watch_ready["ticker"] == "AAA"
    assert watch_ready["company_name"] == "Alpha Apps"
    assert watch_ready["chart_url"] == "/agent-results/charts/aaa.png"
    assert watch_ready["weighted_net_rr"] == 2.22
    assert watch_ready["entry_confirmation_passed"] is False
    assert diagnostics["why_no_buys"][0]["label"] == "Entry confirmation missing"
    assert diagnostics["watch_ready_funnel"]["unique_detected"] == 1
    assert diagnostics["watch_ready_funnel"]["confirmation_passed_unique"] == 0
    assert diagnostics["watch_ready_funnel"]["rr_passed_unique"] == 1
    assert diagnostics["entry_blockers_summary"][0]["label"] == "Entry confirmation missing"
    assert diagnostics["closest_to_entry"][0]["ticker"] == "AAA"
    assert diagnostics["closest_to_entry"][0]["entry_readiness_score"] > 50
    assert diagnostics["closest_to_entry"][0]["missing_conditions"][0]["key"] == "entry_confirmation"


def test_decision_diagnostics_ranks_closest_to_entry_above_weak_candidates() -> None:
    diagnostics = build_decision_diagnostics(
        [
            {
                "ticker": "CLOSE",
                "action": "WATCH_READY",
                "setup_type": "Breakout + Retest",
                "score": 0.62,
                "reason": "WATCH_READY: needs completed entry confirmation",
                "decision_json": {
                    "market_regime": "BULL",
                    "sector_regime": "STRONG",
                    "setup_score": 0.62,
                    "minimum_setup_score_required": 0.45,
                    "minimum_net_rr_required": 2.0,
                    "weighted_net_rr": 1.92,
                    "net_rr_1": 1.05,
                    "net_rr_2": 3.2,
                    "entry_confirmation_passed": False,
                },
            },
            {
                "ticker": "WEAK",
                "action": "SKIP",
                "setup_type": "Fib 61.8",
                "score": 0.38,
                "reason": "SKIP: sector regime is weak and weighted risk/reward is below minimum",
                "decision_json": {
                    "market_regime": "BULL",
                    "sector_regime": "WEAK",
                    "setup_score": 0.38,
                    "minimum_setup_score_required": 0.45,
                    "minimum_net_rr_required": 2.0,
                    "weighted_net_rr": 0.8,
                    "entry_confirmation_passed": False,
                },
            },
        ]
    )

    assert diagnostics["closest_to_entry"][0]["ticker"] == "CLOSE"
    assert diagnostics["closest_to_entry"][0]["entry_readiness_score"] > diagnostics["closest_to_entry"][1]["entry_readiness_score"]
    blocker_labels = [item["label"] for item in diagnostics["entry_blockers_summary"]]
    assert "Weak sector" in blocker_labels
    assert "Weighted/net R/R below gate" in blocker_labels


def test_risk_dashboard_groups_sector_factor_and_capacity() -> None:
    positions = [
        {
            "ticker": "NVDA",
            "sector": "Semiconductors",
            "exposure_ils": 6_000,
            "decision_json": {"factor_tags": ["AI / Semiconductors", "Mega Cap Tech"]},
        },
        {
            "ticker": "MSFT",
            "sector": "Technology",
            "exposure_ils": 4_000,
            "decision_json": {"factor_tags": ["Mega Cap Tech"]},
        },
    ]
    summary = {
        "starting_capital_ils": 100_000,
        "cash_ils": 90_000,
        "exposure_ils": 10_000,
        "open_risk_ils": 800,
    }

    dashboard = build_risk_dashboard(positions, summary, [{"market_regime": "NEUTRAL"}])

    assert dashboard["max_total_exposure"] == 20_000
    assert dashboard["remaining_new_trade_budget"] == 10_000
    assert dashboard["open_risk_pct"] == 0.8
    assert dashboard["sector_exposure"][0]["name"] == "Semiconductors"
    assert dashboard["sector_exposure"][0]["pct_of_exposure"] == 60
    assert dashboard["factor_exposure"][0]["name"] == "Mega Cap Tech"
    assert dashboard["factor_exposure"][0]["count"] == 2


def test_position_timeline_marks_tp1_and_breakeven_stop() -> None:
    positions = [
        with_position_calculations(
            {
                "ticker": "BA",
                "company_name": "Boeing",
                "sector": "Industrials",
                "entry_date": "2026-08-10T14:00:00",
                "entry_price_usd": 200,
                "current_price_usd": 214,
                "quantity": 5,
                "stop_loss": 200,
                "target_1": 212,
                "target_2": 230,
                "notes": "Partial profit taken; stop moved to breakeven.",
            }
        )
    ]

    timeline = build_position_timeline(positions)

    assert timeline[0]["partial_taken"] is True
    assert timeline[0]["breakeven_stop"] is True
    statuses = {step["label"]: step["status"] for step in timeline[0]["steps"]}
    assert statuses["Entry"] == "complete"
    assert statuses["TP1 partial"] == "complete"
    assert statuses["Stop to entry"] == "complete"
    assert statuses["Current"] == "active"


def test_dashboard_section_payload_filters_and_sorts_diagnostics() -> None:
    dashboard = {
        "status": "ok",
        "decision_diagnostics": {
            "drilldowns": {
                "WATCH_READY": [
                    {
                        "ticker": "FAR",
                        "sector": "Technology",
                        "setup_type": "VWAP Reclaim",
                        "setup_score": 0.80,
                        "weighted_net_rr": 2.5,
                        "entry_confirmation_passed": False,
                        "current_price_usd": 120,
                        "buy_zone_low": 100,
                        "buy_zone_high": 104,
                        "chart_url": "",
                    },
                    {
                        "ticker": "NEAR",
                        "sector": "Technology",
                        "setup_type": "Breakout",
                        "setup_score": 0.55,
                        "weighted_net_rr": 2.0,
                        "entry_confirmation_passed": True,
                        "current_price_usd": 101,
                        "buy_zone_low": 100,
                        "buy_zone_high": 104,
                        "chart_url": "/agent-results/charts/near.png",
                    },
                    {
                        "ticker": "HLTH",
                        "sector": "Healthcare",
                        "setup_type": "Breakout",
                        "setup_score": 0.90,
                        "weighted_net_rr": 3.0,
                        "entry_confirmation_passed": True,
                        "current_price_usd": 90,
                        "buy_zone_low": 88,
                        "buy_zone_high": 92,
                        "chart_url": "/agent-results/charts/hlth.png",
                    },
                ]
            }
        },
    }

    payload = dashboard_section_payload(
        dashboard,
        section="diagnostics",
        diagnostic_key="WATCH_READY",
        sector="Technology",
        chart_filter="all",
        confirmation="all",
        sort="closest",
        limit=10,
    )

    assert payload["status"] == "ok"
    assert payload["total"] == 2
    assert [item["ticker"] for item in payload["items"]] == ["NEAR", "FAR"]
    assert payload["facets"]["sectors"][0] == {"value": "Technology", "count": 2}


def test_write_diagnostic_snapshot_creates_run_file(tmp_path) -> None:
    dashboard = {
        "status": "ok",
        "snapshot": {"selected_date": ""},
        "latest_run": {
            "run_id": "run:123",
            "timestamp": "2026-08-09T12:00:00",
            "tickers": ["AAA"],
            "valid_setups": 1,
            "trade_ready_setups": 1,
            "action_counts": {"WATCH_READY": 1},
            "market_regime": "BULL",
        },
        "decision_diagnostics": {"watch_ready_count": 1},
    }

    path = write_diagnostic_snapshot(tmp_path, dashboard)

    assert path is not None
    assert path.name == "diagnostics_run_123.json"
    assert path.exists()
    assert '"watch_ready_count":1' in path.read_text(encoding="utf-8")


def test_position_attention_flags_near_target_one_before_partial_profit() -> None:
    position = with_position_calculations(
        {
            "ticker": "LLY",
            "company_name": "Eli Lilly and Company",
            "sector": "Healthcare",
            "quantity": 4,
            "entry_price_usd": 1180.0,
            "current_price_usd": 1231.94,
            "stop_loss": 1160.0,
            "target_1": 1233.41,
            "target_2": 1290.0,
            "notes": "",
        }
    )

    attention = position["position_attention"]

    assert attention["level"] == "high"
    assert attention["event"] == "TAKE_PARTIAL_PROFIT"
    assert attention["label"] == "Target 1"
    assert attention["distance_pct"] < 0.2


def test_position_attention_uses_target_two_after_partial_profit() -> None:
    position = with_position_calculations(
        {
            "ticker": "ADBE",
            "company_name": "Adobe Inc.",
            "sector": "Technology",
            "quantity": 5,
            "entry_price_usd": 350.0,
            "current_price_usd": 379.5,
            "stop_loss": 350.0,
            "target_1": 360.0,
            "target_2": 382.0,
            "notes": "Partial profit taken; stop moved to breakeven.",
        }
    )

    attention = position["position_attention"]

    assert position["partial_taken"] is True
    assert attention["level"] == "medium"
    assert attention["event"] == "TAKE_PROFIT"
    assert attention["label"] == "Target 2"


def test_build_position_attention_filters_low_priority_and_sorts() -> None:
    positions = [
        with_position_calculations(
            {
                "ticker": "FAR",
                "quantity": 1,
                "entry_price_usd": 100.0,
                "current_price_usd": 101.0,
                "stop_loss": 90.0,
                "target_1": 120.0,
                "target_2": 130.0,
                "notes": "",
            }
        ),
        with_position_calculations(
            {
                "ticker": "HIGH",
                "quantity": 1,
                "entry_price_usd": 100.0,
                "current_price_usd": 104.8,
                "stop_loss": 95.0,
                "target_1": 105.0,
                "target_2": 115.0,
                "notes": "",
            }
        ),
        with_position_calculations(
            {
                "ticker": "MED",
                "quantity": 1,
                "entry_price_usd": 100.0,
                "current_price_usd": 103.0,
                "stop_loss": 95.0,
                "target_1": 104.5,
                "target_2": 115.0,
                "notes": "",
            }
        ),
    ]

    attention = build_position_attention(positions)

    assert [item["ticker"] for item in attention] == ["HIGH", "MED"]
    assert attention[0]["attention"]["level"] == "high"
    assert attention[1]["attention"]["level"] == "medium"
